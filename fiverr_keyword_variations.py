#!/usr/bin/env python3
"""
Fiverr Keyword Variation Finder
─────────────────────────────────
For every "Positive" niche in the Excel, generates alternative keyword
phrasings and tests each on Fiverr to find lower-competition entry points.

Strategy:
  1. Take the base keyword from col E
  2. Generate up to 10 variations using smart prefixes/suffixes
  3. Search each variation on Fiverr — record gig count
  4. Flag variations with gig count < threshold as new opportunities

Excel output:
  New sheet "Keyword Variations" with columns:
    Sub-Niche | Base Keyword | Variation | Gig Count | Status | Opportunity

Also saves a CSV: keyword_variations_results.csv

Reads hub_config.json for threshold/delay settings if present.
"""

import os, re, time, random, json, csv, subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "Excel and Images")
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")
CONFIG_FILE = os.path.join(BASE_DIR, "hub_config.json")
CSV_OUT     = os.path.join(OUTPUT_DIR, "keyword_variations_results.csv")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def _load_cfg():
    d = {"threshold": 1800, "min_delay": 3.5, "max_delay": 6.5, "slow_down": 15}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                return {**d, **json.load(f)}
        except Exception:
            pass
    return d

_cfg      = _load_cfg()
THRESHOLD = _cfg["threshold"]
MIN_DELAY = _cfg["min_delay"]
MAX_DELAY = _cfg["max_delay"]
SLOW_DOWN = _cfg["slow_down"]

# ── Keyword variation modifiers ───────────────────────────────────────────────
# Prefixes that tend to narrow the niche (less competition)
PREFIXES = [
    "professional", "expert", "custom", "affordable",
    "premium", "certified", "dedicated",
]

# Suffixes that add specificity
SUFFIXES = [
    "for small business", "for startup", "for ecommerce",
    "for local business", "with support", "from scratch",
    "for beginners", "for agencies",
]

# Category-specific modifiers detected by keyword content
CATEGORY_HINTS = {
    "wordpress":  ["child theme", "woocommerce", "elementor", "multisite"],
    "shopify":    ["dropshipping", "print on demand", "conversion optimization"],
    "wix":        ["wix studio", "wix store", "wix seo"],
    "ai":         ["for business", "automation", "chatgpt integration", "workflow"],
    "chatbot":    ["whatsapp", "customer support", "lead generation"],
    "seo":        ["google ranking", "organic traffic", "technical seo"],
    "social":     ["content creation", "engagement growth", "brand awareness"],
    "email":      ["automation", "newsletter", "cold outreach", "drip campaign"],
    "tiktok":     ["viral content", "tiktok shop", "influencer"],
    "facebook":   ["facebook ads", "retargeting", "lead generation"],
    "app":        ["ios app", "android app", "cross platform", "react native"],
    "game":       ["2d game", "3d game", "mobile game", "multiplayer"],
    "blockchain": ["smart contract", "nft", "defi", "web3"],
    "python":     ["automation script", "data scraping", "api integration"],
    "react":      ["next.js", "typescript", "redux", "component library"],
}

def get_variations(keyword):
    """
    Generate up to 10 smart keyword variations.
    Returns list of (variation_str, strategy_label).
    """
    kw_lower = keyword.lower()
    variations = []

    # 3 best prefix variations
    for prefix in PREFIXES[:3]:
        variations.append((f"{prefix} {keyword}", f"prefix:{prefix}"))

    # 3 best suffix variations
    for suffix in SUFFIXES[:3]:
        variations.append((f"{keyword} {suffix}", f"suffix:{suffix}"))

    # Category-specific (up to 4)
    added = 0
    for cat_key, hints in CATEGORY_HINTS.items():
        if cat_key in kw_lower and added < 4:
            for hint in hints[:2]:
                if hint not in kw_lower:
                    variations.append((f"{keyword} {hint}", f"niche:{hint}"))
                    added += 1

    # Remove duplicates / too-similar entries, cap at 10
    seen = set()
    unique = []
    for var, label in variations:
        key = var.lower().strip()
        if key not in seen and key != keyword.lower():
            seen.add(key)
            unique.append((var, label))
        if len(unique) >= 10:
            break

    return unique

# ── macOS helpers ─────────────────────────────────────────────────────────────
def notify(title, msg):
    try:
        subprocess.run(["osascript", "-e",
            f'display notification "{msg}" with title "{title}" sound name "Glass"'], timeout=4)
    except Exception:
        pass

def bring_chrome_front():
    try:
        subprocess.run(["osascript", "-e",
            'tell application "Google Chrome" to activate'], timeout=4)
    except Exception:
        pass

# ── CAPTCHA ───────────────────────────────────────────────────────────────────
def is_captcha_page(html, url=""):
    if any(s in url for s in ["/search/gigs", "/categories/"]):
        return False
    hl = html.lower()
    return (("it needs a human touch" in hl or "complete the task and we" in hl)
            and ('id="px-captcha"' in html or 'class="px-loader-wrapper"' in html))

def safe_content(page, retries=3):
    for i in range(retries):
        try:
            time.sleep(1.5)
            return page.content()
        except Exception:
            if i < retries - 1:
                time.sleep(2)
    return ""

def wait_for_captcha_clear(page, timeout_minutes=5):
    notify("Fiverr CAPTCHA", "Please solve the CAPTCHA in Chrome")
    bring_chrome_front()
    for i in range(timeout_minutes * 12):
        time.sleep(5)
        try:
            cur_url = page.url
        except Exception:
            cur_url = ""
        html = safe_content(page)
        if html and not is_captcha_page(html, cur_url):
            print("    CAPTCHA cleared.")
            notify("Fiverr Scraper", "CAPTCHA solved — resuming")
            time.sleep(2)
            return True
        if i % 12 == 0 and i > 0:
            print(f"    Waiting for CAPTCHA... ({i * 5 // 60} min)")
            bring_chrome_front()
    return False

# ── Gig count extraction ─────────────────────────────────────────────────────
def get_gig_count(html, page):
    for pat in [
        r'([\d,]+)\+?\s+[Rr]esults',
        r'([\d,]+)\+?\s+[Ss]ervices',
        r'([\d,]+)\+?\s+[Gg]igs',
    ]:
        m = re.search(pat, html)
        if m:
            raw = m.group(1).replace(",", "")
            if raw.isdigit() and int(raw) > 0:
                return int(raw)

    for pat in [r'"total"\s*:\s*(\d+)', r'"total_count"\s*:\s*(\d+)']:
        for m in re.finditer(pat, html):
            v = int(m.group(1))
            if 10 <= v <= 500000:
                return v

    try:
        count = page.evaluate("""() => {
            try {
                const s = window.__INITIAL_STATE__;
                if (s && s.search && s.search.total_count) return s.search.total_count;
            } catch(e) {}
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while ((node = walker.nextNode())) {
                const t = node.textContent.trim();
                const m = t.match(/^(\\d[\\d,]*)\\+?\\s+(results|services|gigs)$/i);
                if (m) return parseInt(m[1].replace(/,/g,''));
            }
            return null;
        }""")
        if count and isinstance(count, (int, float)) and count > 0:
            return int(count)
    except Exception:
        pass

    return None

# ── Excel sheet helpers ───────────────────────────────────────────────────────
SHEET_NAME = "Keyword Variations"
HEADERS    = ["Sub-Niche", "Niche", "Base Keyword", "Variation",
              "Strategy", "Gig Count", "vs Base", "Status", "Opportunity"]

thin   = Side(style="thin", color="BDBDBD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def ensure_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    ws2 = wb.create_sheet(SHEET_NAME)
    # Headers
    for col, h in enumerate(HEADERS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill      = fill("1DBF73")
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    widths = [25, 20, 30, 35, 20, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    return ws2

def write_variation_row(ws2, data_row, sub, niche, base_kw, variation,
                        strategy, count, base_count, threshold):
    """Append one variation result row."""
    if count is not None and base_count is not None:
        diff = count - base_count
        diff_str = f"{diff:+,}"
    else:
        diff_str = "N/A"

    if count is None:
        status = opp = "N/A"
        row_fg = "FFFDE7"
    elif count <= threshold:
        status = "Positive"
        row_fg = "C8E6C9"
        if base_count and count < base_count * 0.7:
            opp = "Better than base"
        else:
            opp = "Alternative entry"
    else:
        status = "Negative"
        row_fg = "FFCDD2"
        opp    = "Too competitive"

    vals = [sub, niche, base_kw, variation, strategy,
            count if count else "N/A", diff_str, status, opp]

    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=data_row, column=col, value=val)
        c.border    = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col in (4, 3):  # variation and base keyword — left-align
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = fill(row_fg)

    return status

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws2 = ensure_sheet(wb)

    # Collect Positive rows
    positive_rows = []
    for row in range(2, ws.max_row + 1):
        status  = ws.cell(row=row, column=4).value
        keyword = ws.cell(row=row, column=5).value
        sub     = ws.cell(row=row, column=3).value
        niche   = ws.cell(row=row, column=2).value
        if (status and str(status).strip().lower() == "positive"
                and keyword and str(keyword).strip()):
            positive_rows.append({
                "row":     row,
                "keyword": str(keyword).strip(),
                "sub":     str(sub).strip() if sub else "",
                "niche":   str(niche).strip() if niche else "",
            })

    total = len(positive_rows)
    print(f"\n{'='*65}")
    print(f"  Fiverr Keyword Variation Finder")
    print(f"  Positive niches: {total}  |  Threshold: {THRESHOLD:,} gigs")
    print(f"  Testing up to 10 variations per niche")
    print(f"{'='*65}\n")

    if total == 0:
        print("  No Positive niches found. Run the Niche Scraper first.")
        return

    # CSV writer
    csv_rows = []
    next_sheet_row = ws2.max_row + 1

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(
                channel="chrome", headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled",
                      "--no-first-run", "--no-default-browser-check", "--disable-extensions"])
            print("  Browser: system Chrome")
        except Exception:
            browser = pw.chromium.launch(
                headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled", "--no-sandbox"])
            print("  Browser: Playwright Chromium")

        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1400, "height": 900},
            locale="en-US", timezone_id="America/New_York",
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
            window.chrome = { runtime: {} };
        """)
        page = ctx.new_page()

        notify("Fiverr Scraper", "Opening Fiverr for keyword variation testing...")
        print("Opening Fiverr homepage...")
        try:
            page.goto("https://www.fiverr.com", wait_until="load", timeout=40000)
        except Exception as e:
            print(f"  Note: {e}")
        time.sleep(4)
        bring_chrome_front()

        html = safe_content(page)
        try:
            cur_url = page.url
        except Exception:
            cur_url = ""
        if is_captcha_page(html, cur_url):
            print("\n  CAPTCHA — solve in Chrome...")
            wait_for_captcha_clear(page)
        else:
            print(f"  Fiverr loaded ({cur_url[:60]})")

        print("\nStarting in 3 seconds...\n")
        time.sleep(3)

        total_searches = 0
        consecutive    = 0

        for idx, info in enumerate(positive_rows, 1):
            keyword = info["keyword"]
            sub     = info["sub"]
            niche   = info["niche"]

            variations = get_variations(keyword)
            print(f"\n  [{idx}/{total}] {sub[:55]}")
            print(f"    Base keyword : {keyword}")
            print(f"    Variations   : {len(variations)}")

            # First: get base keyword gig count
            base_count = None
            base_q = keyword.replace(" ", "+").replace("&", "%26").replace("/", "+")
            base_url = (f"https://www.fiverr.com/search/gigs?query={base_q}"
                        f"&source=top-bar&search_in=everywhere")
            try:
                page.goto(base_url, wait_until="load", timeout=40000)
                time.sleep(random.uniform(2.5, 3.5))
                html = safe_content(page)
                if is_captcha_page(html):
                    wait_for_captcha_clear(page)
                    page.goto(base_url, wait_until="load", timeout=40000)
                    time.sleep(3)
                    html = safe_content(page)
                base_count = get_gig_count(html, page)
                total_searches += 1
                consecutive    += 1
                print(f"    Base count   : {base_count:,}" if base_count else "    Base count: N/A")
            except Exception as e:
                print(f"    Base search error: {e}")

            # Now test each variation
            for variation, strategy in variations:
                q   = variation.replace(" ", "+").replace("&", "%26").replace("/", "+")
                url = (f"https://www.fiverr.com/search/gigs?query={q}"
                       f"&source=top-bar&search_in=everywhere")

                count = None
                try:
                    page.goto(url, wait_until="load", timeout=40000)
                    time.sleep(random.uniform(2.0, 3.5))
                    html = safe_content(page)
                    if is_captcha_page(html):
                        wait_for_captcha_clear(page)
                        page.goto(url, wait_until="load", timeout=40000)
                        time.sleep(3)
                        html = safe_content(page)
                    count = get_gig_count(html, page)
                    total_searches += 1
                    consecutive    += 1
                except Exception as e:
                    print(f"      [Error: {e}]")

                status = write_variation_row(ws2, next_sheet_row, sub, niche,
                                             keyword, variation, strategy,
                                             count, base_count, THRESHOLD)
                next_sheet_row += 1

                flag = ""
                if count is not None and base_count is not None and count < base_count * 0.7:
                    flag = " ← BETTER OPPORTUNITY"
                diff = (count - base_count) if (count and base_count) else None
                print(f"      {variation[:50]:50s} | {str(count):>8} gigs | {status}{flag}")

                csv_rows.append([sub, niche, keyword, variation, strategy,
                                 count, diff, status])

                if consecutive >= SLOW_DOWN:
                    wait = random.uniform(20, 28)
                    print(f"\n  Rate-limit break: {wait:.0f}s...\n")
                    time.sleep(wait)
                    consecutive = 0
                else:
                    time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

            wb.save(EXCEL_PATH)

        browser.close()

    # Write CSV
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Sub-Niche", "Niche", "Base Keyword", "Variation",
                         "Strategy", "Gig Count", "vs Base", "Status"])
        writer.writerows(csv_rows)

    print(f"\n{'='*65}")
    print(f"  Done! Ran {total_searches} searches across {total} niches.")
    print(f"  Excel: {EXCEL_PATH}  (sheet '{SHEET_NAME}')")
    print(f"  CSV  : {CSV_OUT}")
    print(f"{'='*65}\n")
    notify("Fiverr Scraper", f"Keyword variations done! {total_searches} searches completed.")

if __name__ == "__main__":
    main()
