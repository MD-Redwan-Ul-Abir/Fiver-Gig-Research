#!/usr/bin/env python3
"""
Fiverr Top Sellers Analyzer
────────────────────────────
For each "Positive" niche, searches Fiverr (best_selling sort) and reads
seller level badges on the top 10–12 gig cards.

Seller levels detected:
  TRS  = Top Rated Seller   (hardest to compete with)
  L2   = Level 2 Seller
  L1   = Level 1 Seller
  New  = No badge           (easier to enter alongside)

Entry Opportunity written to Excel:
  High   (≥50% New sellers) — low barrier, good for new sellers
  Medium (25–49% New)
  Low    (<25% New)         — dominated by experienced sellers

Excel output:
  col L → Seller Mix   (e.g. "TRS:2  L2:3  L1:2  New:5")
  col M → Entry Level  (High / Medium / Low)

Reads hub_config.json for threshold/delay settings if present.
"""

import os, re, time, random, subprocess, json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "Excel and Images")
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")
CONFIG_FILE = os.path.join(BASE_DIR, "hub_config.json")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def _load_cfg():
    d = {"threshold": 1800, "min_delay": 4.0, "max_delay": 7.5, "slow_down": 15}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                return {**d, **json.load(f)}
        except Exception:
            pass
    return d

_cfg      = _load_cfg()
MIN_DELAY = _cfg["min_delay"]
MAX_DELAY = _cfg["max_delay"]
SLOW_DOWN = _cfg["slow_down"]

COL_SELLER_MIX  = 12   # L
COL_ENTRY_LEVEL = 13   # M

# ── Styles ────────────────────────────────────────────────────────────────────
thin   = Side(style="thin", color="BDBDBD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def style_cell(cell, val, align="center", bold=False, fg=None):
    cell.value     = val
    cell.border    = border
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bold:
        cell.font = Font(bold=True, size=10)
    if fg:
        cell.fill = fill(fg)

# ── macOS helpers ─────────────────────────────────────────────────────────────
def notify(title, msg):
    try:
        subprocess.run(["osascript", "-e",
            f'display notification "{msg}" with title "{title}" sound name "Glass"'], timeout=4)
    except Exception:
        pass

def bring_chrome_front():
    try:
        subprocess.run(["osascript", "-e",
            'tell application "Google Chrome" to activate'], timeout=4)
    except Exception:
        pass

# ── CAPTCHA ───────────────────────────────────────────────────────────────────
def is_captcha_page(html, url=""):
    if any(s in url for s in ["/search/gigs", "/categories/", "/gigs/"]):
        return False
    hl = html.lower()
    return (("it needs a human touch" in hl or "complete the task and we" in hl)
            and ('id="px-captcha"' in html or 'class="px-loader-wrapper"' in html))

def safe_content(page, retries=3):
    for i in range(retries):
        try:
            time.sleep(1.5)
            return page.content()
        except Exception:
            if i < retries - 1:
                time.sleep(2)
    return ""

def wait_for_captcha_clear(page, timeout_minutes=5):
    notify("Fiverr CAPTCHA", "Please solve the CAPTCHA in Chrome")
    bring_chrome_front()
    for i in range(timeout_minutes * 12):
        time.sleep(5)
        cur_url = ""
        try:
            cur_url = page.url
        except Exception:
            pass
        html = safe_content(page)
        if html and not is_captcha_page(html, cur_url):
            print("    CAPTCHA cleared. Resuming...")
            notify("Fiverr Scraper", "CAPTCHA solved — resuming")
            time.sleep(2)
            return True
        if i % 12 == 0 and i > 0:
            mins = i * 5 // 60
            print(f"    Waiting for CAPTCHA... ({mins} min)")
            bring_chrome_front()
    return False

# ── Seller level extraction ───────────────────────────────────────────────────
def extract_seller_levels(page):
    """
    Inspect gig cards on the current Fiverr search page.
    Returns a list of level strings: 'top_rated', 'level_2', 'level_1', 'new'
    """
    try:
        levels = page.evaluate(r"""() => {
            const CARD_SELS = [
                'li[class*="gig"]', 'div[class*="gig-card"]',
                'div[class*="GigCard"]', 'div[class*="gigCard"]',
                '[data-testid*="gig"]', 'div[class*="basic-gig"]',
            ];
            let cardEls = [];
            for (const sel of CARD_SELS) {
                const found = document.querySelectorAll(sel);
                if (found.length > 0) { cardEls = Array.from(found); break; }
            }
            if (cardEls.length === 0) {
                // Fallback: articles with gig links
                cardEls = Array.from(document.querySelectorAll('article')).filter(el => {
                    const a = el.querySelector('a[href]');
                    return a && /^\/([\w-]+)\/([\w-]+)/.test(a.getAttribute('href') || '');
                });
            }

            const results = [];
            for (const card of cardEls.slice(0, 12)) {
                const text  = card.innerText.toLowerCase();
                const html  = card.innerHTML.toLowerCase();
                let level   = 'new';

                // Check text + class name patterns for each level
                if (text.includes('top rated') ||
                    html.includes('top_rated') ||
                    html.includes('top-rated') ||
                    html.includes('toprated')) {
                    level = 'top_rated';
                } else if (text.includes('level two') || text.includes('level 2') ||
                           html.includes('level_two') || html.includes('level-two') ||
                           html.includes('seller_level_two')) {
                    level = 'level_2';
                } else if (text.includes('level one') || text.includes('level 1') ||
                           html.includes('level_one') || html.includes('level-one') ||
                           html.includes('seller_level_one')) {
                    level = 'level_1';
                }

                // Double-check badge/img alt attributes
                for (const img of card.querySelectorAll('img[alt]')) {
                    const alt = (img.alt || '').toLowerCase();
                    if (alt.includes('top rated'))   { level = 'top_rated'; break; }
                    if (alt.includes('level 2'))      { level = 'level_2';   break; }
                    if (alt.includes('level 1'))      { level = 'level_1';   break; }
                }

                results.push(level);
            }
            return results;
        }""")
        return levels or []
    except Exception as e:
        print(f"    [Seller extract error: {e}]")
        return []

def already_done(ws, row):
    v = ws.cell(row=row, column=COL_SELLER_MIX).value
    return v is not None and str(v).strip() not in ("", "N/A", "ERROR")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Headers
    for col, name in [(COL_SELLER_MIX, "Seller Mix"), (COL_ENTRY_LEVEL, "Entry Level")]:
        cell = ws.cell(row=1, column=col)
        if not cell.value:
            cell.value     = name
            cell.fill      = fill("1DBF73")
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            ws.column_dimensions[get_column_letter(col)].width = 20

    # Collect Positive rows
    positive_rows = []
    for row in range(2, ws.max_row + 1):
        status  = ws.cell(row=row, column=4).value
        keyword = ws.cell(row=row, column=5).value
        if (status and str(status).strip().lower() == "positive"
                and keyword and str(keyword).strip()):
            positive_rows.append((row, str(keyword).strip()))

    total     = len(positive_rows)
    already   = sum(1 for (r, _) in positive_rows if already_done(ws, r))
    remaining = total - already

    print(f"\n{'='*60}")
    print(f"  Fiverr Top Sellers Analyzer")
    print(f"  Positive niches: {total}  |  Already done: {already}  |  Remaining: {remaining}")
    print(f"{'='*60}\n")

    if remaining == 0:
        print("  All positive rows already analyzed!")
        return

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(
                channel="chrome", headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled",
                      "--no-first-run", "--no-default-browser-check", "--disable-extensions"])
            print("  Browser: system Chrome")
        except Exception:
            browser = pw.chromium.launch(
                headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled", "--no-sandbox"])
            print("  Browser: Playwright Chromium")

        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1400, "height": 900},
            locale="en-US", timezone_id="America/New_York",
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en'] });
            window.chrome = { runtime: {} };
        """)
        page = ctx.new_page()

        notify("Fiverr Scraper", "Opening Fiverr — solve CAPTCHA if prompted")
        print("Opening Fiverr homepage...")
        try:
            page.goto("https://www.fiverr.com", wait_until="load", timeout=40000)
        except Exception as e:
            print(f"  Note: {e}")
        time.sleep(4)
        bring_chrome_front()

        html = safe_content(page)
        cur_url = ""
        try:
            cur_url = page.url
        except Exception:
            pass
        if is_captcha_page(html, cur_url):
            print("\n  CAPTCHA on homepage — solve in Chrome...")
            wait_for_captcha_clear(page)
        else:
            print(f"  Fiverr loaded  ({cur_url[:60]})")

        print("\nStarting in 3 seconds...\n")
        time.sleep(3)

        done_count  = 0
        consecutive = 0

        for idx, (row, keyword) in enumerate(positive_rows, 1):
            if already_done(ws, row):
                sub = ws.cell(row=row, column=3).value or ""
                print(f"  [{idx:>3}/{total}] SKIP: {sub[:50]}")
                continue

            sub_niche = ws.cell(row=row, column=3).value or keyword
            print(f"  [{idx:>3}/{total}] Analyzing: {sub_niche[:55]}", end=" ", flush=True)

            query = keyword.replace(" ", "+").replace("&", "%26").replace("/", "+")
            url   = (f"https://www.fiverr.com/search/gigs?query={query}"
                     f"&source=top-bar&search_in=everywhere&sort=best_selling")

            try:
                page.goto(url, wait_until="load", timeout=40000)
                time.sleep(random.uniform(2.5, 4.0))
            except Exception as e:
                print(f"\n    Navigation error: {e}")
                style_cell(ws.cell(row=row, column=COL_SELLER_MIX),  "ERROR", fg="FFB6C1")
                style_cell(ws.cell(row=row, column=COL_ENTRY_LEVEL), "N/A")
                wb.save(EXCEL_PATH)
                continue

            html = safe_content(page)
            if is_captcha_page(html):
                print("\n    CAPTCHA — solve in Chrome...")
                wait_for_captcha_clear(page)
                try:
                    page.goto(url, wait_until="load", timeout=40000)
                    time.sleep(4)
                except Exception:
                    pass

            levels = extract_seller_levels(page)

            if not levels:
                print("N/A")
                style_cell(ws.cell(row=row, column=COL_SELLER_MIX),  "N/A", fg="FFFDE7")
                style_cell(ws.cell(row=row, column=COL_ENTRY_LEVEL), "N/A", fg="FFFDE7")
            else:
                total_cards = len(levels)
                counts = {
                    "top_rated": levels.count("top_rated"),
                    "level_2":   levels.count("level_2"),
                    "level_1":   levels.count("level_1"),
                    "new":       levels.count("new"),
                }
                pct_new = counts["new"] / total_cards * 100 if total_cards else 0

                mix_str = (f"TRS:{counts['top_rated']}  L2:{counts['level_2']}"
                           f"  L1:{counts['level_1']}  New:{counts['new']}")

                if pct_new >= 50:
                    entry, entry_fg = "High",   "90EE90"
                elif pct_new >= 25:
                    entry, entry_fg = "Medium", "FFFACD"
                else:
                    entry, entry_fg = "Low",    "FFB6C1"

                print(f"→ {mix_str}  | {pct_new:.0f}% new | Entry: {entry}")

                style_cell(ws.cell(row=row, column=COL_SELLER_MIX),  mix_str, fg="E3F2FD")
                style_cell(ws.cell(row=row, column=COL_ENTRY_LEVEL), entry, bold=True, fg=entry_fg)

            wb.save(EXCEL_PATH)
            done_count  += 1
            consecutive += 1

            if consecutive >= SLOW_DOWN:
                wait = random.uniform(20, 30)
                print(f"\n  Rate-limit break: {wait:.0f}s...\n")
                time.sleep(wait)
                consecutive = 0
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    print(f"\n{'='*60}")
    print(f"  Done! Analyzed {done_count} niches.")
    print(f"  Results saved to: {EXCEL_PATH}")
    print(f"{'='*60}\n")
    notify("Fiverr Scraper", f"Top seller analysis complete! {done_count} niches done.")

if __name__ == "__main__":
    main()
