#!/usr/bin/env python3
"""
Create Comprehensive Excel — AI-Powered (Multi-Call Edition)
────────────────────────────────────────────────────────────────────
Strategy:
  Phase 1 — Use Fiverr's real category list as the seed (hardcoded from
             Fiverr's actual category tree, 2024-2025).
  Phase 2 — For EACH niche, make a dedicated AI call asking for every
             possible sub-niche. One focused call per niche = no token
             limits, no lazy "top 5" answers.
  Phase 3 — Build / update the Excel file.

Run from the GUI:  Scrapers → Create / Regenerate Excel
"""

import os, sys, json, time

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "Excel and Images")
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(os.path.join(OUTPUT_DIR, "gig_images"), exist_ok=True)


sys.path.insert(0, BASE_DIR)
from api_manager import (
    load_api_config, call_ai, extract_json, has_api_key,
    FIVERR_NICHES, SUBNICHE_PROMPT_TEMPLATE,
)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("  ERROR: openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)

PROVIDERS_LABELS = {
    "gemini": "Google Gemini",
    "openai": "OpenAI ChatGPT",
}

SECTION_COLORS = {
    "Programming & Tech": "1565C0",
    "Digital Marketing":  "2E7D32",
}
DEFAULT_COLOR = "455A64"

HEADERS    = ["Section", "Niche", "Sub-Niche", "Fiverr +/-",
              "Keyword",  "Gig Title", "Gig Description", "Image"]
COL_WIDTHS = [22, 24, 32, 13, 34, 38, 55, 16]


# ── Excel styling helpers ──────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _brd():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _lighten(hex_color, factor=0.88):
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"{int(r+(255-r)*factor):02X}{int(g+(255-g)*factor):02X}{int(b+(255-b)*factor):02X}"

def _hdr(cell, text, bg="1DBF73"):
    cell.value     = text
    cell.fill      = _fill(bg)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _brd()

def _dat(cell, text="", bg="FFFFFF", fg="111111", bold=False, center=False):
    cell.value     = text if text is not None else ""
    cell.fill      = _fill(bg)
    cell.font      = Font(bold=bold, color=fg, size=10)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True)
    cell.border    = _brd()


# ── Excel builder ──────────────────────────────────────────────────────────────

def build_excel(categories: list) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title        = "Niche Research"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        _hdr(ws.cell(1, col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    dv = DataValidation(type="list", formula1='"Positive,Negative"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)

    row = 2
    for cat in categories:
        section  = str(cat.get("section", "Other")).strip()
        niche    = str(cat.get("niche",   "General")).strip()
        sub_list = cat.get("sub_niches", [])
        if not isinstance(sub_list, list) or not sub_list:
            continue

        sec_hex  = SECTION_COLORS.get(section, DEFAULT_COLOR)
        sec_bg   = _lighten(sec_hex, 0.88)
        niche_bg = _lighten(sec_hex, 0.93)

        for sub in sub_list:
            if isinstance(sub, str):
                name, keyword = sub.strip(), sub.strip().lower()
            elif isinstance(sub, dict):
                name    = str(sub.get("name", "")).strip()
                keyword = str(sub.get("keyword", name.lower())).strip()
            else:
                continue
            if not name:
                continue

            _dat(ws.cell(row, 1), section, bg=sec_bg,   fg=sec_hex, bold=True, center=True)
            _dat(ws.cell(row, 2), niche,   bg=niche_bg, fg="333333")
            _dat(ws.cell(row, 3), name)
            c = ws.cell(row, 4)
            _dat(c, "", center=True)
            dv.add(c)
            _dat(ws.cell(row, 5), keyword)
            for col in range(6, 9):
                _dat(ws.cell(row, col))
            ws.row_dimensions[row].height = 18
            row += 1

    wb.save(EXCEL_PATH)
    return row - 2


# ── Per-niche AI call ──────────────────────────────────────────────────────────

def fetch_sub_niches(section: str, niche: str, attempt: int = 1) -> list:
    """
    Ask the AI for every possible sub-niche under one specific niche.
    Returns a list of {"name": ..., "keyword": ...} dicts.
    Retries once on failure.
    """
    prompt = SUBNICHE_PROMPT_TEMPLATE.format(section=section, niche=niche)
    try:
        raw  = call_ai(prompt, max_tokens=4096)
        data = extract_json(raw)
        subs = data.get("sub_niches", [])
        if isinstance(subs, list) and subs:
            return subs
        raise ValueError("Empty sub_niches list")
    except Exception as e:
        if attempt == 1:
            time.sleep(3)
            return fetch_sub_niches(section, niche, attempt=2)
        print(f"    ✗ Failed after retry: {e}")
        return []


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    ensure_output_dir()

    print(f"\n{'='*62}")
    print(f"  Fiverr Excel Creator  —  Exhaustive Sub-Niche Generation")
    print(f"{'='*62}\n")

    if not has_api_key():
        print("  ERROR: No AI API key configured.")
        print("  Open the Hub → Settings → AI Configuration → add your key.")
        sys.exit(1)

    cfg      = load_api_config()
    provider = PROVIDERS_LABELS.get(cfg.get("provider", ""), cfg.get("provider", "Unknown"))

    print(f"  AI Provider  : {provider}")
    print(f"  Output file  : {EXCEL_PATH}")
    print(f"  Niches       : {len(FIVERR_NICHES)} (based on Fiverr's real category tree)")
    print(f"  Strategy     : 1 dedicated AI call per niche → no sub-niche limits\n")

    # ── Phase 2: one AI call per niche ────────────────────────────────────────
    categories  = []
    total_subs  = 0
    failed      = []

    print(f"  {'─'*58}")
    print(f"  {'#':<4} {'Niche':<34} {'Sub-niches':>10}")
    print(f"  {'─'*58}")

    for idx, (section, niche) in enumerate(FIVERR_NICHES, 1):
        label = f"{niche[:33]:<33}"
        print(f"  {idx:<4} {label}  …", end="", flush=True)

        subs = fetch_sub_niches(section, niche)

        if subs:
            categories.append({"section": section, "niche": niche, "sub_niches": subs})
            total_subs += len(subs)
            print(f"\r  {idx:<4} {label}  {len(subs):>5} ✓")
        else:
            failed.append(niche)
            print(f"\r  {idx:<4} {label}  {'FAIL':>5} ✗")

        # Small pause between calls to avoid rate-limiting
        if idx < len(FIVERR_NICHES):
            time.sleep(1.5)

    print(f"  {'─'*58}")
    print(f"\n  Niches processed  : {len(FIVERR_NICHES)}")
    print(f"  Successful        : {len(categories)}")
    print(f"  Total sub-niches  : {total_subs}")
    if failed:
        print(f"  Failed niches     : {', '.join(failed)}")

    if not categories:
        print("\n  ERROR: No data collected. Check your API key and try again.")
        sys.exit(1)

    # ── Phase 3: Build Excel ───────────────────────────────────────────────────
    print(f"\n  Building Excel…")
    written = build_excel(categories)

    print(f"\n{'='*62}")
    print(f"  Complete!")
    print(f"  Rows written  : {written}")
    print(f"  Saved to      : {EXCEL_PATH}")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    main()
