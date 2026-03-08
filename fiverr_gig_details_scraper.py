#!/usr/bin/env python3
"""
Fiverr Gig Details Scraper  (v3 — gig URL + pricing)
──────────────────────────────────────────────────────
For every row in the Excel where "Fiverr +/-" = "Positive":
  1. Takes the keyword from col E
  2. Searches Fiverr, picks the gig with the highest score (reviews + orders×20)
  3. Opens the gig page
  4. Extracts: Gig Title, Gig Description, Gig URL, Starting Price
  5. Downloads ALL gallery images (NOT profile photos):
       First image  → {row}.jpg     (e.g. 3.jpg)
       Extra images → {row}_2.jpg … (e.g. 3_2.jpg)
  6. Writes to Excel:
       col F → Gig Title
       col G → Gig Description
       col H → image embedded + all local paths
       col I → Gig URL  (clickable link to the gig)
       col J → Pricing  (starting / Basic package price)
  7. Saves after every row

Dependencies: playwright  openpyxl  requests  Pillow
  pip install playwright openpyxl requests Pillow
  playwright install chromium
"""

import json, os, re, subprocess, sys, time, random
import requests
import openpyxl
from openpyxl.styles          import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils           import get_column_letter
from openpyxl.drawing.image   import Image as XLImage
from PIL                      import Image as PILImage
from io                       import BytesIO
from playwright.sync_api      import sync_playwright, TimeoutError as PWTimeout

# ── Config ────────────────────────────────────────────────────────────────────
_BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
_OUTPUT_DIR = os.path.join(_BASE_DIR, "Excel and Images")
os.makedirs(_OUTPUT_DIR, exist_ok=True)
EXCEL_PATH  = os.path.join(_OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")
IMAGES_DIR  = os.path.join(_OUTPUT_DIR, "gig_images")
os.makedirs(IMAGES_DIR, exist_ok=True)
MIN_DELAY      = 4.0
MAX_DELAY      = 8.0
GIG_PAGE_DELAY = 6.0
SLOW_DOWN      = 12
IMG_WIDTH      = 200    # embedded image display width  in Excel (px)
IMG_HEIGHT     = 133    # embedded image display height in Excel (px)
MAX_GIG_IMGS   = 5      # max gallery images to download per gig

# ── Column positions (update here if Excel layout changes) ────────────────────
COL_TITLE   = 6   # F  Gig Title
COL_DESC    = 7   # G  Gig Description
COL_IMAGE   = 8   # H  Image (embedded + paths)
COL_GIG_URL = 9   # I  Gig URL  ← NEW
COL_PRICING = 10  # J  Pricing  ← NEW


# ── macOS helpers ─────────────────────────────────────────────────────────────
def notify(title, message):
    try:
        subprocess.run(["osascript", "-e",
            f'display notification "{message}" with title "{title}" sound name "Glass"'],
            timeout=4)
    except Exception:
        pass

def bring_chrome_front():
    try:
        subprocess.run(["osascript", "-e",
            'tell application "Google Chrome" to activate'], timeout=4)
    except Exception:
        pass

# ── Excel styles ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

_thin   = Side(style="thin", color="BDBDBD")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def style_cell(cell, val, align="left", bold=False, fg=None):
    cell.value     = val
    cell.border    = _border
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    if bold:
        cell.font = Font(bold=True, size=10)
    if fg:
        cell.fill = fill(fg)

# ── CAPTCHA ───────────────────────────────────────────────────────────────────
def is_captcha_page(html, url=""):
    if any(s in url for s in ["/search/gigs", "/categories/", "/gigs/", "/users/"]):
        return False
    html_l = html.lower()
    challenge = "it needs a human touch" in html_l or "complete the task and we" in html_l
    widget    = 'id="px-captcha"' in html or 'class="px-loader-wrapper"' in html
    return challenge and widget

def safe_content(page, retries=3):
    for i in range(retries):
        try:
            time.sleep(1.5)
            return page.content()
        except Exception:
            if i < retries - 1:
                time.sleep(2)
    return ""

def wait_for_captcha_clear(page, timeout_minutes=6):
    notify("Fiverr CAPTCHA", "Please solve the CAPTCHA in Chrome")
    bring_chrome_front()
    for i in range(timeout_minutes * 12):
        time.sleep(5)
        cur_url = ""
        try:    cur_url = page.url
        except: pass
        html = safe_content(page)
        if html and not is_captcha_page(html, cur_url):
            print(f"    CAPTCHA cleared. Resuming…")
            notify("Fiverr Scraper", "CAPTCHA solved — resuming")
            time.sleep(2)
            return True
        if i % 12 == 0 and i > 0:
            mins = i * 5 // 60
            print(f"    Waiting for CAPTCHA… ({mins} min)")
            bring_chrome_front()
    return False

# ── Gig-card extraction from search results ───────────────────────────────────
def extract_gig_cards(page, html):
    """Return gig list sorted by score (reviews + orders×20), highest first."""
    cards = []

    try:
        raw = page.evaluate(r"""() => {
            const results = [];
            const CARD_SELS = [
                'li[class*="gig"]', 'div[class*="gig-card"]',
                'div[class*="GigCard"]', 'div[class*="gigCard"]',
                '[data-testid*="gig"]', '[data-testid*="GigCard"]',
                'div[class*="basic-gig"]',
            ];
            let cardEls = [];
            for (const sel of CARD_SELS) {
                const found = document.querySelectorAll(sel);
                if (found.length > 0) { cardEls = Array.from(found); break; }
            }
            if (cardEls.length === 0) {
                cardEls = Array.from(document.querySelectorAll('article, li')).filter(el => {
                    const a = el.querySelector('a[href]');
                    return a && /^\/[^/]+\/[^/]+/.test(a.getAttribute('href') || '');
                });
            }
            for (const card of cardEls.slice(0, 24)) {
                try {
                    let gigUrl = null, gigTitle = null;
                    for (const a of card.querySelectorAll('a[href]')) {
                        const href = a.getAttribute('href') || '';
                        if (/^\/[^/]+\/[^/?#]+($|\?)/.test(href) &&
                            !href.startsWith('/search') &&
                            !href.startsWith('/categories') &&
                            !href.startsWith('/pro')) {
                            gigUrl   = href.startsWith('http') ? href
                                       : 'https://www.fiverr.com' + href.split('?')[0];
                            gigTitle = a.innerText.trim() || a.title || '';
                            break;
                        }
                    }
                    if (!gigUrl) continue;

                    let reviews = 0;
                    for (const el of card.querySelectorAll('*')) {
                        if (el.children.length > 0) continue;
                        const m = el.textContent.trim().match(/^\((\d[\d,]*)\)$/);
                        if (m) { reviews = parseInt(m[1].replace(/,/g, '')); break; }
                    }
                    let orders = 0;
                    for (const el of card.querySelectorAll('*')) {
                        if (el.children.length > 0) continue;
                        const m = el.textContent.trim().match(/^(\d+)\s+[Oo]rders? in [Qq]ueue$/);
                        if (m) { orders = parseInt(m[1]); break; }
                    }
                    results.push({ title: gigTitle, url: gigUrl, reviews, orders });
                } catch (_) {}
            }
            return results;
        }""")
        for item in (raw or []):
            if item.get("url"):
                score = (item.get("reviews") or 0) + (item.get("orders") or 0) * 20
                cards.append({
                    "title":   item.get("title", ""),
                    "url":     item.get("url",   ""),
                    "reviews": item.get("reviews") or 0,
                    "orders":  item.get("orders")  or 0,
                    "score":   score,
                })
    except Exception as e:
        print(f"    [card-extract error: {e}]")

    # Fallback: __NEXT_DATA__ JSON
    if not cards:
        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
        if m:
            try:
                def _dig_gigs(obj, depth=0):
                    if depth > 8: return []
                    if isinstance(obj, list):
                        out = []
                        for x in obj: out.extend(_dig_gigs(x, depth+1))
                        return out
                    if isinstance(obj, dict):
                        if "gig_url" in obj and "title" in obj:
                            return [obj]
                        out = []
                        for v in obj.values(): out.extend(_dig_gigs(v, depth+1))
                        return out
                    return []
                gigs = _dig_gigs(json.loads(m.group(1)))
                for g in gigs[:20]:
                    url = g.get("gig_url", "")
                    if url and not url.startswith("http"):
                        url = "https://www.fiverr.com" + url
                    reviews = int(g.get("review_count") or g.get("reviews_count") or 0)
                    orders  = int(g.get("orders_in_queue") or 0)
                    cards.append({
                        "title":   g.get("title", ""),
                        "url":     url,
                        "reviews": reviews,
                        "orders":  orders,
                        "score":   reviews + orders * 20,
                    })
            except Exception:
                pass

    cards.sort(key=lambda x: x["score"], reverse=True)
    return cards

# ── Gig IMAGE extraction — the core fix ──────────────────────────────────────
def extract_gig_images(page, html):
    """
    Extract ONLY the gig portfolio/gallery images from a gig page.
    Profile photos and avatars are explicitly excluded.

    Returns a list of image URLs (landscape gig images, NOT profile pics).
    """
    img_urls = []

    # ── Strategy 1: __NEXT_DATA__ JSON (most reliable source) ────────────────
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _find_images(obj, depth=0):
                """Recursively collect all image URLs that look like gig gallery images."""
                found = []
                if depth > 12: return found
                if isinstance(obj, list):
                    for x in obj:
                        found.extend(_find_images(x, depth+1))
                elif isinstance(obj, dict):
                    # Common keys that hold gig gallery images
                    for key in ("gallery", "images", "attachments", "portfolio",
                                "gig_images", "gigImages", "media", "assets",
                                "items", "photos"):
                        if key in obj:
                            found.extend(_find_images(obj[key], depth+1))
                    # Direct URL fields
                    for key in ("url", "src", "image_url", "thumbnail",
                                "cover_image", "main_image", "original"):
                        val = obj.get(key, "")
                        if isinstance(val, str) and val.startswith("http"):
                            # Keep only real gig images — skip profile/avatar URLs
                            low = val.lower()
                            if ("profile_photo" not in low and
                                "avatars"        not in low and
                                "avatar"         not in low and
                                "user_"          not in low and
                                "fiverr" in low):
                                found.append(val)
                    # Recurse into remaining values
                    for k, v in obj.items():
                        if k not in ("gallery","images","attachments","portfolio",
                                     "gig_images","gigImages","media","assets",
                                     "items","photos","url","src","image_url",
                                     "thumbnail","cover_image","main_image","original"):
                            found.extend(_find_images(v, depth+1))
                return found

            candidates = _find_images(data)
            # De-duplicate while preserving order
            seen = set()
            for u in candidates:
                if u not in seen:
                    seen.add(u)
                    img_urls.append(u)
        except Exception:
            pass

    # ── Strategy 2: Live DOM — gallery container specifically ─────────────────
    if not img_urls:
        try:
            dom_imgs = page.evaluate(r"""() => {
                const results = [];
                const BAD_PATTERNS = [
                    'profile_photo', 'avatars', 'avatar',
                    'user_', 'icon', 'logo', 'badge', 'level_'
                ];
                const isBad = src => BAD_PATTERNS.some(p => src.toLowerCase().includes(p));

                // Look for the gallery/slider container first
                const GALLERY_SELS = [
                    '[class*="gallery"]',
                    '[class*="Gallery"]',
                    '[class*="slider"]',
                    '[class*="Slider"]',
                    '[class*="ImageGallery"]',
                    '[class*="GigGallery"]',
                    '[class*="gig-gallery"]',
                    '[class*="media-container"]',
                    '[class*="MediaContainer"]',
                    '[class*="gig-page-overview"]',
                    '[class*="GigPageOverview"]',
                ];

                let galleryImgs = [];
                for (const sel of GALLERY_SELS) {
                    const container = document.querySelector(sel);
                    if (!container) continue;
                    const imgs = container.querySelectorAll('img');
                    for (const img of imgs) {
                        const src = img.src || img.dataset.src || img.getAttribute('data-lazy-src') || '';
                        if (src.startsWith('http') && !isBad(src)) {
                            galleryImgs.push(src);
                        }
                    }
                    if (galleryImgs.length > 0) break;
                }

                // Also try: all imgs with 'attachment' in src (Fiverr CDN pattern for gig files)
                if (galleryImgs.length === 0) {
                    for (const img of document.querySelectorAll('img')) {
                        const src = img.src || '';
                        if ((src.includes('attachment') || src.includes('t2/attachments') ||
                             src.includes('fiverr-res.cloudinary')) && !isBad(src)) {
                            galleryImgs.push(src);
                        }
                    }
                }

                // Fallback: any large-ish image (skip tiny icons and squares ≤250px)
                if (galleryImgs.length === 0) {
                    for (const img of document.querySelectorAll('img')) {
                        const src = img.src || '';
                        const w = img.naturalWidth  || parseInt(img.getAttribute('width')  || '0');
                        const h = img.naturalHeight || parseInt(img.getAttribute('height') || '0');
                        // Skip obvious profile pics (250×250 square) and tiny icons
                        if (src.startsWith('http') && !isBad(src) &&
                            !(w === 250 && h === 250) &&
                            !(w > 0 && w === h && w <= 250) &&
                            (w === 0 || w > 300)) {
                            galleryImgs.push(src);
                        }
                    }
                }

                // De-duplicate
                return [...new Set(galleryImgs)];
            }""")
            if dom_imgs:
                img_urls = dom_imgs
        except Exception as e:
            print(f"    [DOM img error: {e}]")

    # ── Strategy 3: regex on raw HTML for Fiverr CDN URLs ─────────────────────
    if not img_urls:
        cdn_patterns = [
            r'(https://fiverr-res\.cloudinary\.com/(?!.*profile_photo)(?!.*avatars)[^\s"\'<>]+\.(?:jpg|jpeg|png|webp))',
            r'(https://media\.fiverr\.com/(?!.*profile_photo)(?!.*avatars)[^\s"\'<>]+\.(?:jpg|jpeg|png|webp))',
        ]
        seen = set()
        for pat in cdn_patterns:
            for mu in re.findall(pat, html, re.IGNORECASE):
                if mu not in seen and "profile_photo" not in mu.lower() and "avatars" not in mu.lower():
                    seen.add(mu)
                    img_urls.append(mu)

    # ── OG image as last resort ───────────────────────────────────────────────
    if not img_urls:
        m2 = re.search(r'<meta[^>]+property="og:image"[^>]+content="([^"]+)"', html)
        if m2:
            img_urls = [m2.group(1)]

    # Final filter: remove profile photos regardless of source
    BAD = {"profile_photo", "avatars", "avatar", "user_", "badge", "level_"}
    img_urls = [u for u in img_urls if not any(b in u.lower() for b in BAD)]

    # De-duplicate
    seen = set()
    clean = []
    for u in img_urls:
        if u not in seen:
            seen.add(u)
            clean.append(u)

    return clean[:MAX_GIG_IMGS]


# ── Gig title + description extraction ───────────────────────────────────────
def extract_gig_text(page, html):
    """Return dict: { title, description }"""
    result = {"title": None, "description": None}

    # __NEXT_DATA__ first
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _find(obj, *keys, depth=0):
                if depth > 10 or not isinstance(obj, (dict, list)): return None
                if isinstance(obj, list):
                    for x in obj:
                        v = _find(x, *keys, depth=depth+1)
                        if v: return v
                else:
                    for k in keys:
                        if k in obj and obj[k]: return obj[k]
                    for v in obj.values():
                        r = _find(v, *keys, depth=depth+1)
                        if r: return r
                return None

            t = _find(data, "title", "gig_title", "gigTitle")
            d = _find(data, "description", "gig_description", "full_description", "overview")
            if t: result["title"]       = str(t).strip()
            if d: result["description"] = str(d).strip()[:1500]
        except Exception:
            pass

    # DOM fallback
    if not result["title"] or not result["description"]:
        try:
            dom = page.evaluate(r"""() => {
                let title = null, desc = null;
                const h1 = document.querySelector('h1');
                if (h1) title = h1.innerText.trim();

                const DSELS = [
                    '[class*="description-content"]',
                    '[class*="Description"] p',
                    '[class*="overview-description"]',
                    '[data-testid*="description"]',
                    '[class*="Overview"] p',
                    '.description p',
                ];
                for (const sel of DSELS) {
                    const els = document.querySelectorAll(sel);
                    if (els.length > 0) {
                        const t = Array.from(els)
                            .map(e => e.innerText.trim())
                            .filter(t => t.length > 30)
                            .join('\n\n').slice(0, 1500);
                        if (t.length > 50) { desc = t; break; }
                    }
                }
                if (!desc) {
                    let longest = '';
                    for (const p of document.querySelectorAll('p')) {
                        const t = p.innerText.trim();
                        if (t.length > longest.length) longest = t;
                    }
                    if (longest.length > 60) desc = longest.slice(0, 1500);
                }
                return { title, desc };
            }""")
            if dom:
                if not result["title"]       and dom.get("title"):
                    result["title"]       = dom["title"].strip()
                if not result["description"] and dom.get("desc"):
                    result["description"] = dom["desc"].strip()
        except Exception as e:
            print(f"    [DOM text error: {e}]")

    # HTML regex fallback
    if not result["title"]:
        m2 = re.search(r'<h1[^>]*>([^<]{10,200})</h1>', html)
        if m2: result["title"] = re.sub(r'\s+', ' ', m2.group(1)).strip()

    return result


# ── Pricing extraction ────────────────────────────────────────────────────────
def extract_gig_price(page, html):
    """
    Extract the starting (Basic package) price from an open gig page.
    Returns a string like "$15" or "N/A".

    Strategy order:
      1. __NEXT_DATA__ JSON  → packages[0].price  (most reliable)
      2. Live DOM            → smallest visible price element
      3. HTML regex          → smallest "$XX" near pricing keywords
    """

    # ── Strategy 1: __NEXT_DATA__ ────────────────────────────────────────────
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _dig_price(obj, depth=0):
                if depth > 12: return None
                if isinstance(obj, list):
                    for x in obj:
                        v = _dig_price(x, depth + 1)
                        if v: return v
                elif isinstance(obj, dict):
                    # packages array — take the first (Basic) package price
                    if "packages" in obj and isinstance(obj["packages"], list):
                        for pkg in obj["packages"]:
                            if isinstance(pkg, dict):
                                for key in ("price", "usd_price", "starting_price",
                                            "amount", "base_price"):
                                    p = pkg.get(key)
                                    if p and isinstance(p, (int, float)) and 1 <= p <= 50000:
                                        return f"${int(p) if p == int(p) else p}"
                    # direct price keys on any dict
                    for key in ("price", "usd_price", "starting_price",
                                "min_price", "base_price", "amount"):
                        p = obj.get(key)
                        if p and isinstance(p, (int, float)) and 1 <= p <= 50000:
                            return f"${int(p) if p == int(p) else p}"
                    for v in obj.values():
                        r = _dig_price(v, depth + 1)
                        if r: return r
                return None

            price = _dig_price(data)
            if price:
                return price
        except Exception:
            pass

    # ── Strategy 2: DOM ───────────────────────────────────────────────────────
    try:
        dom_price = page.evaluate(r"""() => {
            const PRICE_SELS = [
                '[class*="price"]', '[class*="Price"]',
                '[data-testid*="price"]', '[class*="package-price"]',
                '[class*="PackagePrice"]', '[class*="starting-price"]',
            ];
            let prices = [];
            for (const sel of PRICE_SELS) {
                for (const el of document.querySelectorAll(sel)) {
                    const t = el.innerText.trim();
                    const m = t.match(/\$\s*(\d+(?:\.\d{1,2})?)/);
                    if (m) {
                        const v = parseFloat(m[1]);
                        if (v >= 1 && v <= 50000) prices.push(v);
                    }
                }
            }
            // Also walk all text nodes for "$XX" patterns
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while ((node = walker.nextNode())) {
                const t = node.textContent.trim();
                const m = t.match(/^\$\s*(\d+(?:\.\d{1,2})?)$/);
                if (m) {
                    const v = parseFloat(m[1]);
                    if (v >= 1 && v <= 50000) prices.push(v);
                }
            }
            if (prices.length === 0) return null;
            const lowest = Math.min(...prices);
            return `$${Number.isInteger(lowest) ? lowest : lowest.toFixed(2)}`;
        }""")
        if dom_price:
            return dom_price
    except Exception:
        pass

    # ── Strategy 3: HTML regex ────────────────────────────────────────────────
    # Find all "$XX" values and return the smallest (= Basic package price)
    prices_found = re.findall(r'\$\s*(\d+(?:\.\d{1,2})?)', html)
    valid = [float(p) for p in prices_found if 1 <= float(p) <= 50000]
    if valid:
        lowest = min(valid)
        return f"${int(lowest) if lowest == int(lowest) else lowest:.2f}"

    return "N/A"


# ── Image download ────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.fiverr.com/",
}

def download_image(url, filepath, max_size=(800, 533)):
    """
    Download one image. Skip if it looks like a profile photo
    (250×250 square → profile pic, not gig image).
    Returns True on success.
    """
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return False

        img = PILImage.open(BytesIO(resp.content))
        w, h = img.size

        # Reject 250×250 (or any square ≤ 256px) — these are profile pictures
        if w == h and w <= 256:
            print(f"    [Skipped {w}×{h} square — profile photo]")
            return False

        img.thumbnail(max_size, PILImage.LANCZOS)
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        img.save(filepath, "JPEG", quality=85)
        return True
    except Exception as e:
        print(f"    [Image dl error: {e}]")
        return False


def download_gig_images(img_urls, row_num):
    """
    Download all gig images for a row.
    Names: {row}.jpg, {row}_2.jpg, {row}_3.jpg …
    Returns list of local file paths that were successfully saved.
    """
    saved = []
    for i, url in enumerate(img_urls, 1):
        filename = f"{row_num}.jpg" if i == 1 else f"{row_num}_{i}.jpg"
        filepath = os.path.join(IMAGES_DIR, filename)
        if download_image(url, filepath):
            saved.append(filepath)
            print(f"    Image {i}: {filename}  ✓")
        else:
            print(f"    Image {i}: skipped/failed")
    return saved


# ── Excel image embedding ──────────────────────────────────────────────────────
class SafeXLImage(XLImage):
    """
    openpyxl Image subclass that fixes the 'I/O operation on closed file' bug.

    Root cause in openpyxl 3.1.x: _data() calls fp.close() after reading,
    so every wb.save() after the first one crashes because PIL can no longer
    seek the now-closed file pointer.

    Fix: read ALL bytes from the file once, store as a plain bytes object,
    and return them directly from _data() — no file handles ever involved.
    """
    def __init__(self, img_path):
        with open(img_path, "rb") as f:
            self._raw_bytes = f.read()

        pil = PILImage.open(BytesIO(self._raw_bytes))
        self.width, self.height = pil.size
        try:
            self.format = pil.format.lower()
        except AttributeError:
            self.format = "jpeg"
        pil.close()

        # openpyxl required attributes (bypass parent __init__ to avoid file handles)
        self.ref    = None
        self.anchor = "A1"
        self._id    = 1   # overwritten by openpyxl when added to the sheet

    def _data(self):
        """Return raw bytes every time — no PIL file handles, no close() issues."""
        return self._raw_bytes


def embed_image_excel(ws, row_num, col_num, img_path, w=IMG_WIDTH, h=IMG_HEIGHT):
    """Embed the first gig image into the Excel cell using SafeXLImage."""
    try:
        if not os.path.exists(img_path):
            return False

        col_letter = get_column_letter(col_num)
        xl_img        = SafeXLImage(img_path)
        xl_img.width  = w
        xl_img.height = h
        xl_img.anchor = f"{col_letter}{row_num}"
        ws.add_image(xl_img)
        ws.row_dimensions[row_num].height = max(
            ws.row_dimensions[row_num].height or 0,
            h * 0.75 + 8
        )
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0,
            int(w / 7) + 2
        )
        return True
    except Exception as e:
        print(f"    [Excel embed error: {e}]")
        return False


# ── Already-done check ────────────────────────────────────────────────────────
def already_done(ws, row):
    """
    Row is done when the Gig URL column (col I) has a URL.
    This means title, description, URL, pricing, and images are all saved.
    Rows with only images (from an earlier partial run) will be re-processed
    so they get their Gig URL and Pricing filled in.
    """
    val = ws.cell(row=row, column=COL_GIG_URL).value
    return (val is not None and
            str(val).strip() not in ("", "N/A", "ERROR", "None"))


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(IMAGES_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # ── Convert any pre-loaded XLImage objects to SafeXLImage ────────────────
    # When openpyxl loads a workbook that already has embedded images, it creates
    # XLImage objects whose BytesIO refs get closed after the FIRST wb.save().
    # Every subsequent save then crashes with "I/O operation on closed file".
    # Fix: call _data() on each loaded image RIGHT NOW (before any save) to read
    # the bytes while the handle is still open, then wrap in SafeXLImage.
    if ws._images:
        safe = []
        for old_img in ws._images:
            try:
                raw = old_img._data()           # read bytes while handle is fresh
                new_img          = SafeXLImage.__new__(SafeXLImage)
                new_img._raw_bytes = raw
                pil = PILImage.open(BytesIO(raw))
                new_img.width  = old_img.width
                new_img.height = old_img.height
                try:    new_img.format = pil.format.lower()
                except: new_img.format = "jpeg"
                pil.close()
                new_img.ref    = None
                new_img.anchor = old_img.anchor
                new_img._id    = 1
                safe.append(new_img)
            except Exception as e:
                print(f"  Warning: skipping unreadable embedded image ({e})")
        ws._images = safe
        print(f"  Converted {len(safe)} existing image(s) to SafeXLImage")

    if not ws.cell(row=1, column=COL_IMAGE).value:
        ws.cell(row=1, column=COL_IMAGE).value = "Image"

    # Collect positive rows
    positive_rows = []
    for row in range(2, ws.max_row + 1):
        status  = ws.cell(row=row, column=4).value
        keyword = ws.cell(row=row, column=5).value
        if (status and str(status).strip().lower() == "positive"
                and keyword and str(keyword).strip()):
            positive_rows.append((row, str(keyword).strip()))

    total     = len(positive_rows)
    already   = sum(1 for (r, _) in positive_rows if already_done(ws, r))
    remaining = total - already

    print(f"\n{'='*62}")
    print(f"  Fiverr Gig Details Scraper  (v2 — gig images only)")
    print(f"  Positive rows : {total}")
    print(f"  Already done  : {already}")
    print(f"  To process    : {remaining}")
    print(f"{'='*62}\n")

    if remaining == 0:
        print("  All positive rows already done!")
        return

    # ── Launch browser ────────────────────────────────────────────────────────
    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(
                channel="chrome", headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled",
                      "--no-first-run", "--no-default-browser-check", "--disable-extensions"])
            print("  Browser: system Chrome")
        except Exception:
            browser = pw.chromium.launch(
                headless=False,
                args=["--start-maximized", "--disable-blink-features=AutomationControlled", "--no-sandbox"])
            print("  Browser: Playwright Chromium")

        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en'] });
            window.chrome = { runtime: {} };
        """)
        page = ctx.new_page()

        notify("Fiverr Scraper", "Opening Fiverr — solve CAPTCHA if prompted")
        print("Opening Fiverr homepage…")
        try:
            page.goto("https://www.fiverr.com", wait_until="load", timeout=40000)
        except Exception as e:
            print(f"  Note: {e}")
        time.sleep(4)
        bring_chrome_front()

        html = safe_content(page)
        cur_url = ""
        try:   cur_url = page.url
        except: pass
        if is_captcha_page(html, cur_url):
            print("\n  CAPTCHA on homepage — solve in Chrome, script auto-resumes…\n")
            wait_for_captcha_clear(page)
        else:
            print(f"  Fiverr loaded  ({cur_url[:70]})")

        print("\nStarting in 3 seconds…\n")
        time.sleep(3)

        done_count  = 0
        consecutive = 0

        for idx, (row, keyword) in enumerate(positive_rows, 1):
            if already_done(ws, row):
                sub = ws.cell(row=row, column=3).value or ""
                print(f"  [{idx:>3}/{total}] SKIP: {sub[:45]}")
                continue

            sub_niche = ws.cell(row=row, column=3).value or keyword
            print(f"\n  [{idx:>3}/{total}] ── Row {row}  |  {sub_niche[:55]}")
            print(f"    keyword : {keyword[:70]}")

            # ── Search ────────────────────────────────────────────────────────
            query      = keyword.replace(" ", "+").replace("&", "%26").replace("/", "+")
            search_url = (
                f"https://www.fiverr.com/search/gigs"
                f"?query={query}&source=top-bar&search_in=everywhere&sort=best_selling"
            )
            try:
                page.goto(search_url, wait_until="load", timeout=40000)
                time.sleep(random.uniform(3, 5))
            except (PWTimeout, Exception) as e:
                print(f"    Search error: {e} — skipping")
                ws.cell(row=row, column=COL_GIG_URL).value = "ERROR - search failed"
                wb.save(EXCEL_PATH)
                continue

            html = safe_content(page)
            if is_captcha_page(html):
                print("    CAPTCHA — solve in Chrome…")
                wait_for_captcha_clear(page)
                try:
                    page.goto(search_url, wait_until="load", timeout=40000)
                    time.sleep(4)
                    html = safe_content(page)
                except Exception:
                    pass

            # ── Pick best gig ─────────────────────────────────────────────────
            cards = extract_gig_cards(page, html)
            if not cards:
                print(f"    No gig cards found — skipping")
                ws.cell(row=row, column=COL_GIG_URL).value = "ERROR - no gigs found"
                wb.save(EXCEL_PATH)
                continue

            best = cards[0]
            print(f"    Best gig  : {(best['title'] or 'N/A')[:65]}")
            print(f"    Reviews: {best['reviews']}  Orders: {best['orders']}  Score: {best['score']}")
            print(f"    URL       : {best['url'][:80]}")

            # ── Open gig page ─────────────────────────────────────────────────
            try:
                page.goto(best["url"], wait_until="load", timeout=45000)
                time.sleep(GIG_PAGE_DELAY)
            except (PWTimeout, Exception) as e:
                print(f"    Gig page error: {e} — skipping")
                ws.cell(row=row, column=COL_GIG_URL).value = "ERROR - gig page failed"
                wb.save(EXCEL_PATH)
                continue

            html = safe_content(page)
            if is_captcha_page(html):
                print("    CAPTCHA on gig page — solve in Chrome…")
                wait_for_captcha_clear(page)
                try:
                    page.goto(best["url"], wait_until="load", timeout=45000)
                    time.sleep(4)
                    html = safe_content(page)
                except Exception:
                    pass

            # ── Extract title + description ───────────────────────────────────
            text_data  = extract_gig_text(page, html)
            gig_title  = text_data["title"]       or best["title"] or "N/A"
            gig_desc   = text_data["description"] or "N/A"
            gig_url    = best["url"]
            print(f"    Title     : {gig_title[:70]}")
            print(f"    Desc      : {(gig_desc[:100] + '…') if len(gig_desc) > 100 else gig_desc}")
            print(f"    Gig URL   : {gig_url}")

            # ── Extract pricing ───────────────────────────────────────────────
            gig_price = extract_gig_price(page, html)
            print(f"    Pricing   : {gig_price}")

            # ── Extract & download gig gallery images ─────────────────────────
            img_urls    = extract_gig_images(page, html)
            print(f"    Gallery   : {len(img_urls)} image URL(s) found")
            for u in img_urls:
                print(f"      → {u[:80]}")

            saved_paths = download_gig_images(img_urls, row)

            if not saved_paths:
                print(f"    WARNING: No gig images downloaded for row {row}")

            # ── Write to Excel ────────────────────────────────────────────────
            # Col F — Gig Title
            style_cell(ws.cell(row=row, column=COL_TITLE),   gig_title, fg="FFF9C4", bold=True)
            # Col G — Gig Description
            style_cell(ws.cell(row=row, column=COL_DESC),    gig_desc,  fg="F1F8E9")
            # Col H — Image (embedded + paths)
            if saved_paths:
                ws.cell(row=row, column=COL_IMAGE).value = "\n".join(saved_paths)
                embed_image_excel(ws, row, COL_IMAGE, saved_paths[0])
                print(f"    Images    : {len(saved_paths)} saved  [{row}.jpg …]")
            elif img_urls:
                ws.cell(row=row, column=COL_IMAGE).value = "\n".join(img_urls[:3])
                print(f"    Images    : URLs stored (download failed)")
            else:
                ws.cell(row=row, column=COL_IMAGE).value = "N/A"

            # Col I — Gig URL
            url_cell = ws.cell(row=row, column=COL_GIG_URL)
            url_cell.value     = gig_url
            url_cell.font      = Font(color="0563C1", underline="single", size=10)
            url_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            url_cell.border    = _border

            # Col J — Pricing
            style_cell(ws.cell(row=row, column=COL_PRICING), gig_price,
                       align="center", fg="E8F5E9", bold=True)

            wb.save(EXCEL_PATH)
            done_count  += 1
            consecutive += 1
            print(f"    Saved ✓  (URL + Price stored)")

            # Rate limiting
            if consecutive >= SLOW_DOWN:
                wait = random.uniform(20, 35)
                print(f"\n  Rate-limit break: {wait:.0f}s…\n")
                time.sleep(wait)
                consecutive = 0
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    print(f"\n{'='*62}")
    print(f"  Done!  Processed {done_count} rows this run.")
    print(f"  Images : {IMAGES_DIR}")
    print(f"  Excel  : {EXCEL_PATH}")
    print(f"{'='*62}\n")
    notify("Fiverr Scraper", f"Complete! {done_count} gigs scraped.")


if __name__ == "__main__":
    main()
