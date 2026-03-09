#!/usr/bin/env python3
"""
Fiverr Submission Analyzer
──────────────────────────
Analyzes the top 5 gigs for each specified submission (sub-niche from the Excel).

Usage (run from GUI or command line):
  python fiverr_submission_analyzer.py "10"
      → Analyze top 10 Positive submissions by opportunity score

  python fiverr_submission_analyzer.py "3,7,12"
      → Analyze specific rows by Excel row number

  python fiverr_submission_analyzer.py "Corporate Website, AI SaaS Platform"
      → Analyze submissions whose sub-niche name contains the given text

Output:
  CSV  : Excel and Images/submission_top_gigs.csv  (single shared file, append mode)
  Images: Excel and Images/gig_images/sa_{row}_{rank}_{img}.jpg

CSV columns:
  Submission, Row, Rank, Gig_Title, Gig_Description, Pricing,
  Image_Paths, Gig_URL, Reviews, Orders, Score

Dependencies: playwright  openpyxl  requests  Pillow
"""

import json, os, re, subprocess, sys, time, random, csv
from datetime import datetime
import requests
import openpyxl
from PIL                 import Image as PILImage
from io                  import BytesIO
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ─────────────────────────────────────────────────────────────────────
import os as _os, json as _json

_BASE_DIR    = _os.path.dirname(_os.path.abspath(__file__))
_OUTPUT_DIR  = _os.path.join(_BASE_DIR, "Excel and Images")
_CONFIG_FILE = _os.path.join(_BASE_DIR, "hub_config.json")
_os.makedirs(_OUTPUT_DIR, exist_ok=True)

def _load_cfg():
    d = {"threshold": 1800, "min_delay": 3.5, "max_delay": 6.5,
         "slow_down": 18,
         "excel_path":  _os.path.join(_OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx"),
         "images_dir":  _os.path.join(_OUTPUT_DIR, "gig_images")}
    if _os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE) as f:
                return {**d, **_json.load(f)}
        except Exception:
            pass
    return d

_cfg       = _load_cfg()
THRESHOLD  = int(_cfg["threshold"])
MIN_DELAY  = float(_cfg["min_delay"])
MAX_DELAY  = float(_cfg["max_delay"])
SLOW_DOWN  = int(_cfg["slow_down"])
EXCEL_PATH = _cfg["excel_path"]
IMAGES_DIR = _cfg["images_dir"]

_os.makedirs(IMAGES_DIR, exist_ok=True)

TOP_GIGS_PER_SUB = 5   # analyze this many gigs per submission
MAX_GIG_IMGS     = 3   # download up to this many images per gig
GIG_PAGE_DELAY   = 6.0

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.fiverr.com/",
}


# ── macOS helpers ──────────────────────────────────────────────────────────────
def notify(title, message):
    try:
        subprocess.run(["osascript", "-e",
            f'display notification "{message}" with title "{title}" sound name "Glass"'],
            timeout=4)
    except Exception:
        pass

def bring_chrome_front():
    try:
        subprocess.run(["osascript", "-e",
            'tell application "Google Chrome" to activate'], timeout=4)
    except Exception:
        pass


# ── CAPTCHA ────────────────────────────────────────────────────────────────────
def is_captcha_page(html, url=""):
    if any(s in url for s in ["/search/gigs", "/categories/", "/gigs/", "/users/"]):
        return False
    html_l = html.lower()
    challenge = "it needs a human touch" in html_l or "complete the task and we" in html_l
    widget    = 'id="px-captcha"' in html or 'class="px-loader-wrapper"' in html
    return challenge and widget

def safe_content(page, retries=3):
    for i in range(retries):
        try:
            time.sleep(1.5)
            return page.content()
        except Exception:
            if i < retries - 1:
                time.sleep(2)
    return ""

def wait_for_captcha_clear(page, timeout_minutes=6):
    notify("Fiverr CAPTCHA", "Please solve the CAPTCHA in Chrome")
    bring_chrome_front()
    for i in range(timeout_minutes * 12):
        time.sleep(5)
        cur_url = ""
        try:    cur_url = page.url
        except: pass
        html = safe_content(page)
        if html and not is_captcha_page(html, cur_url):
            print(f"    CAPTCHA cleared. Resuming…")
            notify("Fiverr Analyzer", "CAPTCHA solved — resuming")
            time.sleep(2)
            return True
        if i % 12 == 0 and i > 0:
            mins = i * 5 // 60
            print(f"    Waiting for CAPTCHA… ({mins} min)")
            bring_chrome_front()
    return False


# ── Gig card extraction ────────────────────────────────────────────────────────
def extract_gig_cards(page, html, limit=TOP_GIGS_PER_SUB):
    """Return up to `limit` gigs sorted by score (reviews + orders×20), highest first."""
    cards = []

    try:
        raw = page.evaluate(r"""() => {
            const results = [];
            const CARD_SELS = [
                'li[class*="gig"]', 'div[class*="gig-card"]',
                'div[class*="GigCard"]', 'div[class*="gigCard"]',
                '[data-testid*="gig"]', '[data-testid*="GigCard"]',
                'div[class*="basic-gig"]',
            ];
            let cardEls = [];
            for (const sel of CARD_SELS) {
                const found = document.querySelectorAll(sel);
                if (found.length > 0) { cardEls = Array.from(found); break; }
            }
            if (cardEls.length === 0) {
                cardEls = Array.from(document.querySelectorAll('article, li')).filter(el => {
                    const a = el.querySelector('a[href]');
                    return a && /^\/[^/]+\/[^/]+/.test(a.getAttribute('href') || '');
                });
            }
            for (const card of cardEls.slice(0, 30)) {
                try {
                    let gigUrl = null, gigTitle = null;
                    for (const a of card.querySelectorAll('a[href]')) {
                        const href = a.getAttribute('href') || '';
                        if (/^\/[^/]+\/[^/?#]+($|\?)/.test(href) &&
                            !href.startsWith('/search') &&
                            !href.startsWith('/categories') &&
                            !href.startsWith('/pro')) {
                            gigUrl   = href.startsWith('http') ? href
                                       : 'https://www.fiverr.com' + href.split('?')[0];
                            gigTitle = a.innerText.trim() || a.title || '';
                            break;
                        }
                    }
                    if (!gigUrl) continue;

                    let reviews = 0;
                    for (const el of card.querySelectorAll('*')) {
                        if (el.children.length > 0) continue;
                        const m = el.textContent.trim().match(/^\((\d[\d,]*)\)$/);
                        if (m) { reviews = parseInt(m[1].replace(/,/g, '')); break; }
                    }
                    let orders = 0;
                    for (const el of card.querySelectorAll('*')) {
                        if (el.children.length > 0) continue;
                        const m = el.textContent.trim().match(/^(\d+)\s+[Oo]rders? in [Qq]ueue$/);
                        if (m) { orders = parseInt(m[1]); break; }
                    }
                    results.push({ title: gigTitle, url: gigUrl, reviews, orders });
                } catch (_) {}
            }
            return results;
        }""")
        for item in (raw or []):
            if item.get("url"):
                score = (item.get("reviews") or 0) + (item.get("orders") or 0) * 20
                cards.append({
                    "title":   item.get("title", ""),
                    "url":     item.get("url",   ""),
                    "reviews": item.get("reviews") or 0,
                    "orders":  item.get("orders")  or 0,
                    "score":   score,
                })
    except Exception as e:
        print(f"    [card-extract error: {e}]")

    # Fallback: __NEXT_DATA__
    if not cards:
        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
        if m:
            try:
                def _dig_gigs(obj, depth=0):
                    if depth > 8: return []
                    if isinstance(obj, list):
                        out = []
                        for x in obj: out.extend(_dig_gigs(x, depth+1))
                        return out
                    if isinstance(obj, dict):
                        if "gig_url" in obj and "title" in obj:
                            return [obj]
                        out = []
                        for v in obj.values(): out.extend(_dig_gigs(v, depth+1))
                        return out
                    return []
                gigs = _dig_gigs(json.loads(m.group(1)))
                for g in gigs[:30]:
                    url = g.get("gig_url", "")
                    if url and not url.startswith("http"):
                        url = "https://www.fiverr.com" + url
                    reviews = int(g.get("review_count") or g.get("reviews_count") or 0)
                    orders  = int(g.get("orders_in_queue") or 0)
                    cards.append({
                        "title":   g.get("title", ""),
                        "url":     url,
                        "reviews": reviews,
                        "orders":  orders,
                        "score":   reviews + orders * 20,
                    })
            except Exception:
                pass

    cards.sort(key=lambda x: x["score"], reverse=True)
    return cards[:limit]


# ── Gig text extraction ────────────────────────────────────────────────────────
def extract_gig_text(page, html):
    result = {"title": None, "description": None}

    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _find(obj, *keys, depth=0):
                if depth > 10 or not isinstance(obj, (dict, list)): return None
                if isinstance(obj, list):
                    for x in obj:
                        v = _find(x, *keys, depth=depth+1)
                        if v: return v
                else:
                    for k in keys:
                        if k in obj and obj[k]: return obj[k]
                    for v in obj.values():
                        r = _find(v, *keys, depth=depth+1)
                        if r: return r
                return None

            t = _find(data, "title", "gig_title", "gigTitle")
            d = _find(data, "description", "gig_description", "full_description", "overview")
            if t: result["title"]       = str(t).strip()
            if d: result["description"] = str(d).strip()[:2000]
        except Exception:
            pass

    if not result["title"] or not result["description"]:
        try:
            dom = page.evaluate(r"""() => {
                let title = null, desc = null;
                const h1 = document.querySelector('h1');
                if (h1) title = h1.innerText.trim();
                const DSELS = [
                    '[class*="description-content"]',
                    '[class*="Description"] p',
                    '[class*="overview-description"]',
                    '[data-testid*="description"]',
                    '[class*="Overview"] p',
                    '.description p',
                ];
                for (const sel of DSELS) {
                    const els = document.querySelectorAll(sel);
                    if (els.length > 0) {
                        const t = Array.from(els)
                            .map(e => e.innerText.trim())
                            .filter(t => t.length > 30)
                            .join('\n\n').slice(0, 2000);
                        if (t.length > 50) { desc = t; break; }
                    }
                }
                if (!desc) {
                    let longest = '';
                    for (const p of document.querySelectorAll('p')) {
                        const t = p.innerText.trim();
                        if (t.length > longest.length) longest = t;
                    }
                    if (longest.length > 60) desc = longest.slice(0, 2000);
                }
                return { title, desc };
            }""")
            if dom:
                if not result["title"]       and dom.get("title"):
                    result["title"]       = dom["title"].strip()
                if not result["description"] and dom.get("desc"):
                    result["description"] = dom["desc"].strip()
        except Exception as e:
            print(f"    [DOM text error: {e}]")

    if not result["title"]:
        m2 = re.search(r'<h1[^>]*>([^<]{10,200})</h1>', html)
        if m2: result["title"] = re.sub(r'\s+', ' ', m2.group(1)).strip()

    return result


# ── Pricing extraction ─────────────────────────────────────────────────────────
def extract_gig_price(page, html):
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _dig_price(obj, depth=0):
                if depth > 12: return None
                if isinstance(obj, list):
                    for x in obj:
                        v = _dig_price(x, depth + 1)
                        if v: return v
                elif isinstance(obj, dict):
                    if "packages" in obj and isinstance(obj["packages"], list):
                        for pkg in obj["packages"]:
                            if isinstance(pkg, dict):
                                for key in ("price", "usd_price", "starting_price",
                                            "amount", "base_price"):
                                    p = pkg.get(key)
                                    if p and isinstance(p, (int, float)) and 1 <= p <= 50000:
                                        return f"${int(p) if p == int(p) else p}"
                    for key in ("price", "usd_price", "starting_price",
                                "min_price", "base_price", "amount"):
                        p = obj.get(key)
                        if p and isinstance(p, (int, float)) and 1 <= p <= 50000:
                            return f"${int(p) if p == int(p) else p}"
                    for v in obj.values():
                        r = _dig_price(v, depth + 1)
                        if r: return r
                return None

            price = _dig_price(data)
            if price:
                return price
        except Exception:
            pass

    try:
        dom_price = page.evaluate(r"""() => {
            const PRICE_SELS = [
                '[class*="price"]', '[class*="Price"]',
                '[data-testid*="price"]', '[class*="package-price"]',
                '[class*="PackagePrice"]', '[class*="starting-price"]',
            ];
            let prices = [];
            for (const sel of PRICE_SELS) {
                for (const el of document.querySelectorAll(sel)) {
                    const t = el.innerText.trim();
                    const m = t.match(/\$\s*(\d+(?:\.\d{1,2})?)/);
                    if (m) {
                        const v = parseFloat(m[1]);
                        if (v >= 1 && v <= 50000) prices.push(v);
                    }
                }
            }
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while ((node = walker.nextNode())) {
                const t = node.textContent.trim();
                const m = t.match(/^\$\s*(\d+(?:\.\d{1,2})?)$/);
                if (m) {
                    const v = parseFloat(m[1]);
                    if (v >= 1 && v <= 50000) prices.push(v);
                }
            }
            if (prices.length === 0) return null;
            const lowest = Math.min(...prices);
            return `$${Number.isInteger(lowest) ? lowest : lowest.toFixed(2)}`;
        }""")
        if dom_price:
            return dom_price
    except Exception:
        pass

    prices_found = re.findall(r'\$\s*(\d+(?:\.\d{1,2})?)', html)
    valid = [float(p) for p in prices_found if 1 <= float(p) <= 50000]
    if valid:
        lowest = min(valid)
        return f"${int(lowest) if lowest == int(lowest) else lowest:.2f}"

    return "N/A"


# ── Image extraction ───────────────────────────────────────────────────────────
def extract_gig_images(page, html):
    img_urls = []

    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group(1))

            def _find_images(obj, depth=0):
                found = []
                if depth > 12: return found
                if isinstance(obj, list):
                    for x in obj:
                        found.extend(_find_images(x, depth+1))
                elif isinstance(obj, dict):
                    for key in ("gallery", "images", "attachments", "portfolio",
                                "gig_images", "gigImages", "media", "assets",
                                "items", "photos"):
                        if key in obj:
                            found.extend(_find_images(obj[key], depth+1))
                    for key in ("url", "src", "image_url", "thumbnail",
                                "cover_image", "main_image", "original"):
                        val = obj.get(key, "")
                        if isinstance(val, str) and val.startswith("http"):
                            low = val.lower()
                            if ("profile_photo" not in low and
                                "avatars"        not in low and
                                "avatar"         not in low and
                                "user_"          not in low and
                                "fiverr" in low):
                                found.append(val)
                    for k, v in obj.items():
                        if k not in ("gallery","images","attachments","portfolio",
                                     "gig_images","gigImages","media","assets",
                                     "items","photos","url","src","image_url",
                                     "thumbnail","cover_image","main_image","original"):
                            found.extend(_find_images(v, depth+1))
                return found

            candidates = _find_images(data)
            seen = set()
            for u in candidates:
                if u not in seen:
                    seen.add(u)
                    img_urls.append(u)
        except Exception:
            pass

    if not img_urls:
        try:
            dom_imgs = page.evaluate(r"""() => {
                const results = [];
                const BAD_PATTERNS = [
                    'profile_photo', 'avatars', 'avatar',
                    'user_', 'icon', 'logo', 'badge', 'level_'
                ];
                const isBad = src => BAD_PATTERNS.some(p => src.toLowerCase().includes(p));
                const GALLERY_SELS = [
                    '[class*="gallery"]', '[class*="Gallery"]',
                    '[class*="slider"]', '[class*="Slider"]',
                    '[class*="ImageGallery"]', '[class*="GigGallery"]',
                    '[class*="gig-gallery"]', '[class*="media-container"]',
                    '[class*="MediaContainer"]', '[class*="gig-page-overview"]',
                    '[class*="GigPageOverview"]',
                ];
                let galleryImgs = [];
                for (const sel of GALLERY_SELS) {
                    const container = document.querySelector(sel);
                    if (!container) continue;
                    const imgs = container.querySelectorAll('img');
                    for (const img of imgs) {
                        const src = img.src || img.dataset.src || img.getAttribute('data-lazy-src') || '';
                        if (src.startsWith('http') && !isBad(src)) {
                            galleryImgs.push(src);
                        }
                    }
                    if (galleryImgs.length > 0) break;
                }
                if (galleryImgs.length === 0) {
                    for (const img of document.querySelectorAll('img')) {
                        const src = img.src || '';
                        if ((src.includes('attachment') || src.includes('t2/attachments') ||
                             src.includes('fiverr-res.cloudinary')) && !isBad(src)) {
                            galleryImgs.push(src);
                        }
                    }
                }
                if (galleryImgs.length === 0) {
                    for (const img of document.querySelectorAll('img')) {
                        const src = img.src || '';
                        const w = img.naturalWidth  || parseInt(img.getAttribute('width')  || '0');
                        const h = img.naturalHeight || parseInt(img.getAttribute('height') || '0');
                        if (src.startsWith('http') && !isBad(src) &&
                            !(w === 250 && h === 250) &&
                            !(w > 0 && w === h && w <= 250) &&
                            (w === 0 || w > 300)) {
                            galleryImgs.push(src);
                        }
                    }
                }
                return [...new Set(galleryImgs)];
            }""")
            if dom_imgs:
                img_urls = dom_imgs
        except Exception as e:
            print(f"    [DOM img error: {e}]")

    if not img_urls:
        cdn_patterns = [
            r'(https://fiverr-res\.cloudinary\.com/(?!.*profile_photo)(?!.*avatars)[^\s"\'<>]+\.(?:jpg|jpeg|png|webp))',
            r'(https://media\.fiverr\.com/(?!.*profile_photo)(?!.*avatars)[^\s"\'<>]+\.(?:jpg|jpeg|png|webp))',
        ]
        seen = set()
        for pat in cdn_patterns:
            for mu in re.findall(pat, html, re.IGNORECASE):
                if mu not in seen and "profile_photo" not in mu.lower() and "avatars" not in mu.lower():
                    seen.add(mu)
                    img_urls.append(mu)

    if not img_urls:
        m2 = re.search(r'<meta[^>]+property="og:image"[^>]+content="([^"]+)"', html)
        if m2:
            img_urls = [m2.group(1)]

    BAD = {"profile_photo", "avatars", "avatar", "user_", "badge", "level_"}
    img_urls = [u for u in img_urls if not any(b in u.lower() for b in BAD)]

    seen = set()
    clean = []
    for u in img_urls:
        if u not in seen:
            seen.add(u)
            clean.append(u)

    return clean[:MAX_GIG_IMGS]


# ── Image download ─────────────────────────────────────────────────────────────
def download_image(url, filepath):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return False
        img = PILImage.open(BytesIO(resp.content))
        w, h = img.size
        if w == h and w <= 256:
            print(f"    [Skipped {w}×{h} square — profile photo]")
            return False
        img.thumbnail((800, 533), PILImage.LANCZOS)
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        img.save(filepath, "JPEG", quality=85)
        return True
    except Exception as e:
        print(f"    [Image dl error: {e}]")
        return False


def download_gig_images(img_urls, row_num, rank):
    """
    Download images for one gig.
    Naming: sa_{row}_{rank}_{imgnum}.jpg
    Returns list of local file paths saved.
    """
    saved = []
    for i, url in enumerate(img_urls, 1):
        filename = f"sa_{row_num}_{rank}_{i}.jpg"
        filepath = os.path.join(IMAGES_DIR, filename)
        if download_image(url, filepath):
            saved.append(filepath)
            print(f"      Image {i}: {filename}  ✓")
        else:
            print(f"      Image {i}: skipped/failed")
    return saved


# ── Excel loading & opportunity score ─────────────────────────────────────────
def calc_opp_score(gig_count, avg_queue, avg_reviews, threshold=None):
    threshold = threshold or THRESHOLD
    score = 0.0
    if gig_count is not None:
        score += max(0.0, 1.0 - gig_count / threshold) * 40
    if avg_queue is not None:
        score += min(float(avg_queue) / 3.0, 1.0) * 35
    if avg_reviews is not None:
        r = float(avg_reviews)
        score += (8 if r < 20 else 16 if r < 50 else 25 if r < 150 else 15 if r < 300 else 6)
    return round(score, 1)


def load_submissions_from_excel():
    """
    Load all rows from the Excel file.
    Returns list of dicts with keys: row, section, niche, sub_niche, status, keyword,
                                      gig_count, avg_reviews, avg_queue, opp_score
    """
    if not os.path.exists(EXCEL_PATH):
        print(f"  ERROR: Excel not found at {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        sub_niche = ws.cell(r, 3).value
        keyword   = ws.cell(r, 5).value
        if not sub_niche or not keyword:
            continue
        status     = str(ws.cell(r, 4).value or "").strip()
        gig_count  = ws.cell(r, 9).value
        avg_reviews= ws.cell(r, 10).value
        avg_queue  = ws.cell(r, 11).value

        try: gig_count   = int(gig_count)   if gig_count   is not None else None
        except: gig_count = None
        try: avg_reviews = float(avg_reviews) if avg_reviews is not None else None
        except: avg_reviews = None
        try: avg_queue   = float(avg_queue)   if avg_queue   is not None else None
        except: avg_queue = None

        opp_score = calc_opp_score(gig_count, avg_queue, avg_reviews)

        rows.append({
            "row":       r,
            "section":   str(ws.cell(r, 1).value or "").strip(),
            "niche":     str(ws.cell(r, 2).value or "").strip(),
            "sub_niche": str(sub_niche).strip(),
            "status":    status,
            "keyword":   str(keyword).strip(),
            "gig_count": gig_count,
            "avg_reviews": avg_reviews,
            "avg_queue":   avg_queue,
            "opp_score":   opp_score,
        })
    return rows


def parse_query(query_str, all_rows):
    """
    Parse the input query and return the list of matching submission dicts.

    Modes:
      "10"          → top 10 Positive submissions by opportunity score
      "3,7,12"      → specific Excel row numbers
      "Corp Web, AI" → substring match on sub_niche (case-insensitive)
    """
    query_str = query_str.strip()
    parts = [p.strip() for p in query_str.split(",") if p.strip()]

    if not parts:
        return []

    # All parts are integers → row-number mode
    if all(p.isdigit() for p in parts):
        if len(parts) == 1:
            n = int(parts[0])
            # If n looks like a row number (≥ 2) AND there's a matching row, use row mode
            row_match = [r for r in all_rows if r["row"] == n]
            if row_match:
                print(f"  Mode: row-number  (row {n})")
                return row_match
            # Otherwise treat as "top N"
            print(f"  Mode: top-N  (top {n} by opportunity score)")
            positive = [r for r in all_rows if r["status"].lower() == "positive"]
            positive.sort(key=lambda x: x["opp_score"], reverse=True)
            return positive[:n]
        else:
            row_nums = {int(p) for p in parts}
            selected = [r for r in all_rows if r["row"] in row_nums]
            print(f"  Mode: row-numbers  ({sorted(row_nums)})")
            if not selected:
                print("  WARNING: No matching rows found in Excel.")
            return selected

    # Text mode → substring match on sub_niche
    print(f"  Mode: name-search  ({parts})")
    selected = []
    for r in all_rows:
        sn_lower = r["sub_niche"].lower()
        if any(p.lower() in sn_lower for p in parts):
            selected.append(r)
    if not selected:
        print("  WARNING: No sub-niches matched the given names.")
    return selected


# ── CSV writer ─────────────────────────────────────────────────────────────────
CSV_FIELDNAMES = [
    "Scraped_At", "Submission", "Row", "Rank", "Gig_Title", "Gig_Description",
    "Pricing", "Image_Paths", "Gig_URL", "Reviews", "Orders", "Score",
]

# Fixed filename — all runs append to this single file
SUBMISSION_CSV = os.path.join(_OUTPUT_DIR, "submission_top_gigs.csv")

def open_csv_writer(csv_path):
    """Open in append mode; write header only when creating the file for the first time."""
    file_exists = os.path.exists(csv_path)
    f = open(csv_path, "a", newline="", encoding="utf-8")
    writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
    if not file_exists:
        writer.writeheader()
    return f, writer


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2 or not sys.argv[1].strip():
        print("Usage: python fiverr_submission_analyzer.py \"<query>\"")
        print("  <query> = row number(s) (e.g. '3,7,12'), a count (e.g. '10'),")
        print("            or sub-niche name(s) (e.g. 'Corporate Website, AI SaaS')")
        sys.exit(1)

    query_str = sys.argv[1].strip()

    print(f"\n{'='*62}")
    print(f"  Fiverr Submission Analyzer")
    print(f"  Query   : {query_str}")
    print(f"  Excel   : {EXCEL_PATH}")
    print(f"  Images  : {IMAGES_DIR}")
    print(f"{'='*62}\n")

    # ── Load Excel & resolve submissions ──────────────────────────────────────
    print("  Loading Excel…")
    all_rows  = load_submissions_from_excel()
    print(f"  Found {len(all_rows)} rows in Excel.")

    submissions = parse_query(query_str, all_rows)
    if not submissions:
        print("\n  No submissions to analyze. Exiting.")
        sys.exit(0)

    print(f"  Submissions to analyze: {len(submissions)}")
    for s in submissions:
        print(f"    Row {s['row']:>3}: {s['sub_niche'][:55]}  (score: {s['opp_score']})")

    # ── Prepare CSV output ─────────────────────────────────────────────────────
    # All runs share one file; rows are tagged with batch_ts so the Gig Creator
    # can distinguish the most recent scan for each submission.
    batch_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    csv_file, csv_writer = open_csv_writer(SUBMISSION_CSV)
    print(f"\n  CSV     : {SUBMISSION_CSV}  (appending batch {batch_ts})\n")

    # ── Launch browser ────────────────────────────────────────────────────────
    os.makedirs(IMAGES_DIR, exist_ok=True)

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(
                channel="chrome", headless=False,
                args=["--start-maximized",
                      "--disable-blink-features=AutomationControlled",
                      "--no-first-run", "--no-default-browser-check",
                      "--disable-extensions"])
            print("  Browser: system Chrome")
        except Exception:
            browser = pw.chromium.launch(
                headless=False,
                args=["--start-maximized",
                      "--disable-blink-features=AutomationControlled",
                      "--no-sandbox"])
            print("  Browser: Playwright Chromium")

        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en'] });
            window.chrome = { runtime: {} };
        """)
        page = ctx.new_page()

        notify("Fiverr Analyzer", "Opening Fiverr — solve CAPTCHA if prompted")
        print("Opening Fiverr homepage…")
        try:
            page.goto("https://www.fiverr.com", wait_until="load", timeout=40000)
        except Exception as e:
            print(f"  Note: {e}")
        time.sleep(4)
        bring_chrome_front()

        html = safe_content(page)
        cur_url = ""
        try:   cur_url = page.url
        except: pass
        if is_captcha_page(html, cur_url):
            print("\n  CAPTCHA on homepage — solve in Chrome, script auto-resumes…\n")
            wait_for_captcha_clear(page)
        else:
            print(f"  Fiverr loaded  ({cur_url[:70]})")

        print("\nStarting in 3 seconds…\n")
        time.sleep(3)

        total_subs  = len(submissions)
        consecutive = 0
        total_gigs_written = 0

        for sub_idx, sub in enumerate(submissions, 1):
            row      = sub["row"]
            keyword  = sub["keyword"]
            sub_name = sub["sub_niche"]

            print(f"\n{'─'*62}")
            print(f"  [{sub_idx:>2}/{total_subs}]  Row {row}  |  {sub_name[:55]}")
            print(f"    keyword : {keyword}")
            print(f"    score   : {sub['opp_score']}")

            # ── Search ──────────────────────────────────────────────────────
            query      = keyword.replace(" ", "+").replace("&", "%26").replace("/", "+")
            search_url = (
                f"https://www.fiverr.com/search/gigs"
                f"?query={query}&source=top-bar&search_in=everywhere&sort=best_selling"
            )
            try:
                page.goto(search_url, wait_until="load", timeout=40000)
                time.sleep(random.uniform(3, 5))
            except (PWTimeout, Exception) as e:
                print(f"    Search error: {e} — skipping submission")
                continue

            html = safe_content(page)
            if is_captcha_page(html):
                print("    CAPTCHA — solve in Chrome…")
                wait_for_captcha_clear(page)
                try:
                    page.goto(search_url, wait_until="load", timeout=40000)
                    time.sleep(4)
                    html = safe_content(page)
                except Exception:
                    pass

            # ── Top 5 gig cards ─────────────────────────────────────────────
            cards = extract_gig_cards(page, html, limit=TOP_GIGS_PER_SUB)
            if not cards:
                print(f"    No gig cards found — skipping")
                continue

            print(f"    Found {len(cards)} gig(s) to analyze")

            # ── Process each gig ─────────────────────────────────────────────
            for rank, card in enumerate(cards, 1):
                print(f"\n    Gig #{rank}  {(card['title'] or 'N/A')[:60]}")
                print(f"      Reviews: {card['reviews']}  Orders: {card['orders']}  Score: {card['score']}")
                print(f"      URL: {card['url'][:80]}")

                try:
                    page.goto(card["url"], wait_until="load", timeout=45000)
                    time.sleep(GIG_PAGE_DELAY)
                except (PWTimeout, Exception) as e:
                    print(f"      Gig page error: {e} — skipping gig")
                    continue

                html = safe_content(page)
                if is_captcha_page(html):
                    print("      CAPTCHA on gig page — solve in Chrome…")
                    wait_for_captcha_clear(page)
                    try:
                        page.goto(card["url"], wait_until="load", timeout=45000)
                        time.sleep(4)
                        html = safe_content(page)
                    except Exception:
                        pass

                # Extract data
                text_data = extract_gig_text(page, html)
                gig_title = text_data["title"]       or card["title"] or "N/A"
                gig_desc  = text_data["description"] or "N/A"
                gig_price = extract_gig_price(page, html)
                img_urls  = extract_gig_images(page, html)

                print(f"      Title   : {gig_title[:65]}")
                print(f"      Price   : {gig_price}")
                print(f"      Desc    : {(gig_desc[:100] + '…') if len(gig_desc) > 100 else gig_desc}")
                print(f"      Images  : {len(img_urls)} URL(s) found")

                # Download images
                saved_paths = download_gig_images(img_urls, row, rank)
                img_paths_str = " | ".join(saved_paths) if saved_paths else ""

                # Write CSV row
                csv_writer.writerow({
                    "Scraped_At":     batch_ts,
                    "Submission":     sub_name,
                    "Row":            row,
                    "Rank":           rank,
                    "Gig_Title":      gig_title,
                    "Gig_Description": gig_desc,
                    "Pricing":        gig_price,
                    "Image_Paths":    img_paths_str,
                    "Gig_URL":        card["url"],
                    "Reviews":        card["reviews"],
                    "Orders":         card["orders"],
                    "Score":          card["score"],
                })
                csv_file.flush()
                total_gigs_written += 1
                print(f"      Saved to CSV ✓  (rank {rank})")

                consecutive += 1
                if consecutive >= SLOW_DOWN:
                    wait = random.uniform(20, 35)
                    print(f"\n  Rate-limit break: {wait:.0f}s…\n")
                    time.sleep(wait)
                    consecutive = 0
                else:
                    if rank < len(cards):  # delay between gigs in same submission
                        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

            # Delay between submissions
            if sub_idx < total_subs:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()
        csv_file.close()

    print(f"\n{'='*62}")
    print(f"  Done!")
    print(f"  Submissions analyzed : {total_subs}")
    print(f"  Gig records written  : {total_gigs_written}")
    print(f"  CSV                  : {SUBMISSION_CSV}")
    print(f"  Images               : {IMAGES_DIR}")
    print(f"{'='*62}\n")
    notify("Fiverr Analyzer", f"Complete! {total_gigs_written} gigs analyzed.")


if __name__ == "__main__":
    main()
