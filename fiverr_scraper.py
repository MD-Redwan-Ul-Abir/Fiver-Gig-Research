#!/usr/bin/env python3
"""
Fiverr Sub-Niche Scraper
────────────────────────
• Opens a VISIBLE Chromium window (bypasses bot detection)
• User solves any CAPTCHA once, then the script runs automatically
• For each sub-niche: extracts Gig Count, Avg Reviews, Avg Queue
• Updates Excel:
    - Col D  → "Positive" (≤1800 gigs) or "Negative" (>1800 gigs)
    - Col I  → Gig Count
    - Col J  → Avg Review Count (from first page of gig cards)
    - Col K  → Avg Orders in Queue (from first page of gig cards)
• Saves progress after every row so you can Ctrl+C and resume later
"""

import time, random, re, sys, subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ──────────────────────────────────────────────────────────────────
import os as _os, json as _json
_BASE_DIR    = _os.path.dirname(_os.path.abspath(__file__))
_OUTPUT_DIR  = _os.path.join(_BASE_DIR, "Excel and Images")
_CONFIG_FILE = _os.path.join(_BASE_DIR, "hub_config.json")
_os.makedirs(_OUTPUT_DIR, exist_ok=True)
EXCEL_PATH   = _os.path.join(_OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")

def _load_cfg():
    d = {"threshold": 1800, "min_delay": 3.5, "max_delay": 6.5, "slow_down": 18}
    if _os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE) as f:
                return {**d, **_json.load(f)}
        except Exception:
            pass
    return d

_cfg      = _load_cfg()
THRESHOLD = int(_cfg["threshold"])   # max gigs → Positive (read from Settings)
MIN_DELAY = _cfg["min_delay"]
MAX_DELAY = _cfg["max_delay"]
SLOW_DOWN = int(_cfg["slow_down"])

# ── Notifications & Chrome focus ────────────────────────────────────────────
def notify(title, message):
    try:
        subprocess.run(["osascript","-e",
            f'display notification "{message}" with title "{title}" sound name "Glass"'], timeout=4)
    except Exception: pass

def bring_chrome_front():
    try:
        subprocess.run(["osascript","-e",'tell application "Google Chrome" to activate'], timeout=4)
    except Exception: pass

# ── Styles ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

thin = Side(style="thin", color="BDBDBD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, val, align="center", bold=False, fg=None):
    cell.value = val
    cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bold:
        cell.font = Font(bold=True, size=10)
    if fg:
        cell.fill = fill(fg)

# ── Excel helpers ────────────────────────────────────────────────────────────
def prepare_extra_columns(ws):
    """Add Gig Count / Avg Reviews / Avg Queue headers if not present."""
    headers_needed = {9: "Gig Count", 10: "Avg Reviews", 11: "Avg Queue"}
    for col, name in headers_needed.items():
        cell = ws.cell(row=1, column=col)
        if cell.value != name:
            cell.value = name
            cell.fill = fill("1DBF73")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 16
    return headers_needed

def already_done(ws, row):
    """Return True if this row already has gig-count data."""
    v = ws.cell(row=row, column=9).value
    return v is not None and str(v).strip() not in ("", "N/A", "ERROR")

# ── Data extraction ──────────────────────────────────────────────────────────
CAPTCHA_SIGNALS = [
    "it needs a human touch",
    "complete the task",
    "px-captcha",
    "perimeterx",
    "robot",
    "cloudflare",
    "please verify",
    "access denied",
]

def is_captcha_page(html, url=""):
    """Return True only when Fiverr is actively showing a bot challenge."""
    if any(u in url for u in ["/search/gigs", "/categories/", "/gigs/"]):
        return False
    html_l = html.lower()
    is_challenge = ("it needs a human touch" in html_l or "complete the task and we" in html_l)
    has_widget   = ('id="px-captcha"' in html or 'class="px-loader-wrapper"' in html)
    return is_challenge and has_widget

def safe_content(page, retries=3):
    """Get page HTML safely."""
    for i in range(retries):
        try:
            time.sleep(1.5)
            return page.content()
        except Exception:
            if i < retries - 1:
                time.sleep(2)
    return ""

def wait_for_captcha_clear(page, label="", timeout_minutes=6):
    """Poll until CAPTCHA is gone. Send macOS notifications + bring Chrome to front."""
    notify("Fiverr CAPTCHA", "Please solve the CAPTCHA in the Chrome window")
    bring_chrome_front()
    for i in range(timeout_minutes * 12):
        time.sleep(5)
        cur_url = ""
        try: cur_url = page.url
        except: pass
        html = safe_content(page)
        if html and not is_captcha_page(html, cur_url):
            print(f"  CAPTCHA solved! URL: {cur_url[:60]}")
            notify("Fiverr Scraper", "CAPTCHA solved — resuming searches")
            time.sleep(2)
            return True
        if i % 12 == 0 and i > 0:
            mins = i * 5 // 60
            print(f"  Still waiting ({mins} min)... URL: {cur_url[:50]}")
            notify("Fiverr CAPTCHA", f"Still waiting — solve CAPTCHA in Chrome ({mins} min)")
            bring_chrome_front()
    print("  CAPTCHA wait timed out.")
    return False

def extract_gig_count(html, page):
    """Try several strategies to get the total gig count from a Fiverr search page."""
    # Strategy 1: Visible text — Fiverr shows "84,000+ results" or "1,234 results"
    # Must handle the optional "+" between number and word
    for pat in [
        r'([\d,]+)\+?\s+[Rr]esults',       # "84,000+ results" / "1,234 results"
        r'([\d,]+)\+?\s+[Ss]ervices',       # "1,234 services"
        r'([\d,]+)\+?\s+[Gg]igs',           # "1,234 gigs"
        r'Results\s+for[^<]{0,80}?(\d[\d,]+)',  # near "Results for ..."
    ]:
        m = re.search(pat, html)
        if m:
            raw = m.group(1).replace(",", "")
            if raw.isdigit():
                val = int(raw)
                if val > 0:
                    return val

    # Strategy 2: JSON "total" key in embedded page data (Fiverr uses "total":84073)
    for pat in [
        r'"total"\s*:\s*(\d+)',
        r'"total_count"\s*:\s*(\d+)',
        r'"gigs_count"\s*:\s*(\d+)',
    ]:
        for m in re.finditer(pat, html):
            val = int(m.group(1))
            if 10 <= val <= 500000:
                return val

    # Strategy 3: JavaScript DOM text walk — look for "X+ results" visible text nodes
    try:
        count = page.evaluate("""() => {
            try {
                const s = window.__INITIAL_STATE__;
                if (s && s.search && s.search.total_count) return s.search.total_count;
                if (s && s.searchResults && s.searchResults.total_count) return s.searchResults.total_count;
            } catch(e) {}
            // Walk text nodes looking for patterns like "84,000+ results"
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while ((node = walker.nextNode())) {
                const t = node.textContent.trim();
                const m = t.match(/^(\\d[\\d,]*)\\+?\\s+(results|services|gigs)$/i);
                if (m) return parseInt(m[1].replace(/,/g,''));
            }
            return null;
        }""")
        if count and isinstance(count, (int, float)) and count > 0:
            return int(count)
    except Exception:
        pass

    # Strategy 4: Try CSS selectors for count-like elements
    try:
        for sel in [
            "[class*='count']",
            "[class*='result']",
            "[class*='total']",
            "span[class*='Count']",
            "div[class*='Count']",
        ]:
            els = page.query_selector_all(sel)
            for el in els[:10]:
                try:
                    txt = el.inner_text().strip()
                    m = re.match(r'^([\d,]+)$', txt)
                    if m:
                        val = int(m.group(1).replace(",", ""))
                        if 1 <= val <= 500000:
                            return val
                except Exception:
                    pass
    except Exception:
        pass

    return None

def extract_card_stats(page, html):
    """
    Extract review counts and queue counts from gig cards on the search results page.
    Returns (avg_reviews, avg_queue) as floats, or (None, None).
    """
    review_counts = []
    queue_counts  = []

    # Strategy A: JavaScript to read gig card data
    try:
        data = page.evaluate("""() => {
            const results = [];
            // Fiverr renders gig cards — look for review counts and queue
            const allText = document.querySelectorAll('*');
            for (const el of allText) {
                if (el.children.length > 0) continue; // leaf nodes only
                const t = el.textContent.trim();
                const rev = t.match(/^\\((\\d[\\d,]*)\\)$/);
                if (rev) results.push({type:'review', val: parseInt(rev[1].replace(/,/g,''))});
                const q = t.match(/^(\\d+)\\s+[Oo]rders? in [Qq]ueue$/);
                if (q) results.push({type:'queue', val: parseInt(q[1])});
            }
            return results;
        }""")
        if data:
            for item in data:
                if item["type"] == "review" and 0 < item["val"] < 100000:
                    review_counts.append(item["val"])
                elif item["type"] == "queue" and 0 < item["val"] < 10000:
                    queue_counts.append(item["val"])
    except Exception:
        pass

    # Strategy B: Regex on raw HTML
    if not review_counts:
        for m in re.finditer(r'\((\d[\d,]*)\)', html):
            val = int(m.group(1).replace(",", ""))
            if 1 <= val <= 50000:
                review_counts.append(val)
        # de-dup and cap
        review_counts = list(set(review_counts))[:30]

    if not queue_counts:
        for m in re.finditer(r'(\d+)\s+[Oo]rders? in [Qq]ueue', html):
            queue_counts.append(int(m.group(1)))

    avg_rev   = round(sum(review_counts) / len(review_counts), 1) if review_counts else None
    avg_queue = round(sum(queue_counts)  / len(queue_counts),  1) if queue_counts  else None
    return avg_rev, avg_queue

# ── Main scraper ─────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    prepare_extra_columns(ws)

    # Collect all rows that need processing
    rows_to_process = []
    for row in range(2, ws.max_row + 1):
        sub_niche = ws.cell(row=row, column=3).value
        if sub_niche and str(sub_niche).strip():
            rows_to_process.append((row, str(sub_niche).strip()))

    total = len(rows_to_process)
    print(f"\n{'='*60}")
    print(f"  Fiverr Sub-Niche Scraper")
    print(f"  Total sub-niches to search: {total}")
    print(f"  Gig count threshold: {THRESHOLD}")
    print(f"{'='*60}\n")

    # Count already done
    already = sum(1 for (row, _) in rows_to_process if already_done(ws, row))
    print(f"  Already completed: {already}")
    print(f"  Remaining: {total - already}")
    print()

    with sync_playwright() as p:
        # ── Open a VISIBLE browser (uses system Chrome for best fingerprint) ─
        try:
            browser = p.chromium.launch(
                channel="chrome",
                headless=False,
                args=[
                    "--start-maximized",
                    "--disable-blink-features=AutomationControlled",
                    "--no-first-run",
                    "--no-default-browser-check",
                    "--disable-extensions",
                ]
            )
            print("  Using system Chrome browser")
        except Exception:
            browser = p.chromium.launch(
                headless=False,
                args=[
                    "--start-maximized",
                    "--disable-blink-features=AutomationControlled",
                    "--no-first-run",
                    "--no-sandbox",
                ]
            )
            print("  Using Playwright Chromium browser")

        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1400, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1,2,3,4,5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US','en'] });
            window.chrome = { runtime: {} };
        """)
        page = ctx.new_page()

        # ── Step 1: Go to Fiverr homepage to establish session ─────────────
        notify("Fiverr Scraper", "Opening Fiverr — please solve CAPTCHA if Chrome shows one")
        print("Opening Fiverr homepage in the browser...")
        try:
            page.goto("https://www.fiverr.com", wait_until="load", timeout=40000)
        except Exception as e:
            print(f"  Note: {e}")
        time.sleep(4)
        bring_chrome_front()

        html = safe_content(page)
        cur_url = ""
        try: cur_url = page.url
        except: pass
        if is_captcha_page(html, cur_url):
            print("\n⚠️  CAPTCHA! Please solve it in the Chrome window.")
            print("  Script auto-continues once you clear the CAPTCHA...\n")
            wait_for_captcha_clear(page, "homepage")
        else:
            print(f"  Fiverr homepage loaded (URL: {cur_url[:60]})")

        print("\nStarting searches in 3 seconds...\n")
        time.sleep(3)

        done_count = 0
        consecutive = 0

        for idx, (row, sub_niche) in enumerate(rows_to_process, 1):
            # Skip rows already done
            if already_done(ws, row):
                print(f"  [{idx}/{total}] SKIP (done): {sub_niche}")
                continue

            print(f"  [{idx}/{total}] Searching: {sub_niche} ...", end=" ", flush=True)

            # Build search URL
            query = sub_niche.replace(" ", "+").replace("&", "%26").replace("/", "+")
            url = f"https://www.fiverr.com/search/gigs?query={query}&source=top-bar&search_in=everywhere"

            # Navigate
            retries = 0
            success = False
            while retries < 3:
                try:
                    page.goto(url, wait_until="load", timeout=40000)
                    time.sleep(random.uniform(2.5, 4))
                    success = True
                    break
                except PWTimeout:
                    retries += 1
                    print(f"  [timeout, retry {retries}]", end=" ", flush=True)
                    time.sleep(5)
                except Exception as e:
                    retries += 1
                    print(f"  [error: {e}, retry {retries}]", end=" ", flush=True)
                    time.sleep(5)

            if not success:
                print("FAILED (navigation) — skipping")
                style_cell(ws.cell(row=row, column=9),  "ERROR", fg="FFB6C1")
                style_cell(ws.cell(row=row, column=10), "N/A")
                style_cell(ws.cell(row=row, column=11), "N/A")
                wb.save(EXCEL_PATH)
                continue

            html = safe_content(page)

            # Handle CAPTCHA mid-scrape
            if is_captcha_page(html):
                print("\n  ⚠️  CAPTCHA! Please solve it in the Chrome window...")
                wait_for_captcha_clear(page)
                # Retry current search after CAPTCHA is solved
                try:
                    page.goto(url, wait_until="load", timeout=40000)
                    time.sleep(4)
                    html = safe_content(page)
                except Exception:
                    pass

            # Extract data
            gig_count = extract_gig_count(html, page)
            avg_rev, avg_queue = extract_card_stats(page, html)

            # Determine +/-
            if gig_count is not None:
                positive = gig_count <= THRESHOLD
                verdict = "Positive" if positive else "Negative"
                verdict_fg = "90EE90" if positive else "FFB6C1"
                count_display = f"{gig_count:,}"
                print(f"{gig_count:,} gigs → {verdict}", end="")
            else:
                verdict = "N/A"
                verdict_fg = "FFFACD"
                count_display = "N/A"
                print("count=N/A", end="")

            if avg_rev:
                print(f" | Avg Reviews: {avg_rev}", end="")
            if avg_queue:
                print(f" | Avg Queue: {avg_queue}", end="")
            print()

            # Write to Excel
            # Col D (4): Fiverr +/-
            d_cell = ws.cell(row=row, column=4)
            d_cell.value = verdict
            d_cell.fill  = fill(verdict_fg)
            d_cell.border = border
            d_cell.alignment = Alignment(horizontal="center", vertical="center")
            d_cell.font = Font(bold=True, size=10)

            # Col I (9): Gig Count
            style_cell(ws.cell(row=row, column=9),
                       gig_count if gig_count is not None else "N/A",
                       fg="C8E6C9" if (gig_count and gig_count <= THRESHOLD) else
                          "FFCDD2" if (gig_count and gig_count > THRESHOLD) else "FFFDE7")

            # Col J (10): Avg Reviews
            style_cell(ws.cell(row=row, column=10),
                       avg_rev if avg_rev is not None else "N/A",
                       fg="E3F2FD")

            # Col K (11): Avg Queue
            style_cell(ws.cell(row=row, column=11),
                       avg_queue if avg_queue is not None else "N/A",
                       fg="F3E5F5")

            # Save after every row
            wb.save(EXCEL_PATH)
            done_count += 1
            consecutive += 1

            # Take a longer break every SLOW_DOWN searches to avoid rate limiting
            if consecutive >= SLOW_DOWN:
                wait = random.uniform(20, 35)
                print(f"\n  Taking a {wait:.0f}s break to avoid rate limiting...\n")
                time.sleep(wait)
                consecutive = 0
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    print(f"\n{'='*60}")
    print(f"  Scraping complete!")
    print(f"  Processed {done_count} sub-niches this run.")
    print(f"  Results saved to: {EXCEL_PATH}")
    print(f"{'='*60}\n")

    # Summary
    positive = negative = na = 0
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=4).value
        if v == "Positive":   positive += 1
        elif v == "Negative": negative += 1
        else:                 na += 1
    print(f"  Positive (≤{THRESHOLD:,} gigs): {positive}")
    print(f"  Negative (>{THRESHOLD:,} gigs): {negative}")
    print(f"  Not retrieved (N/A / Error): {na}")

if __name__ == "__main__":
    main()
