#!/usr/bin/env python3
"""
Fiverr Research Hub  —  Unified platform for Fiverr niche research.
Double-click FiverrResearchHub.command to launch.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess, threading, os, sys, json, csv, queue, math, platform, webbrowser
from datetime import datetime

# ── Platform detection ─────────────────────────────────────────────────────────
_PLATFORM = platform.system()   # 'Darwin', 'Windows', 'Linux'
if _PLATFORM == 'Windows':
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "Excel and Images")
os.makedirs(OUTPUT_DIR, exist_ok=True)
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "Fiverr_Comprehensive_Niche_Research.xlsx")
IMAGES_DIR  = os.path.join(OUTPUT_DIR, "gig_images")
os.makedirs(IMAGES_DIR, exist_ok=True)
SNAP_DIR    = os.path.join(BASE_DIR, "snapshots")
CONFIG_FILE = os.path.join(BASE_DIR, "hub_config.json")
PYTHON      = sys.executable
APP_VERSION = "2.0"

# ── API Manager (same folder) ─────────────────────────────────────────────────
sys.path.insert(0, BASE_DIR)
try:
    from api_manager import (has_api_key, save_api_config, load_api_config,
                             test_api_key, get_masked_key, get_provider_label)
    _API_MANAGER_OK = True
except ImportError:
    _API_MANAGER_OK = False
    def has_api_key():       return True
    def get_masked_key():    return "—"
    def get_provider_label(): return "Not configured"

DEFAULT_SETTINGS = {
    "threshold": 1800, "min_delay": 3.5, "max_delay": 6.5,
    "slow_down": 18,   "excel_path": EXCEL_PATH, "images_dir": IMAGES_DIR,
}

# ── Premium Theme ─────────────────────────────────────────────────────────────
C_BG       = "#0A0F1C"   # deep navy
C_PANEL    = "#111827"   # sidebar / header
C_CARD     = "#1A2436"   # card surface
C_CARD_H   = "#1F2D44"   # card hover
C_BORDER   = "#1E3050"   # subtle border
C_LINE     = "#243554"   # separator line
C_GREEN    = "#1DBF73"   # Fiverr brand
C_GREEN_L  = "#25D980"   # lighter green (hover)
C_GREEN_DK = "#158F56"   # darker
C_GREEN_BG = "#0A2218"   # dark green tint
C_TEXT     = "#EEF4FB"   # primary text
C_TEXT2    = "#BDD1E4"   # secondary
C_MUTED    = "#87AABF"   # muted — lightened to pass WCAG AA on all dark surfaces
C_YELLOW   = "#F4C842"   # warning
C_YELLOW_D = "#C49A0E"
C_RED      = "#FF4757"   # danger
C_RED_D    = "#CC2233"
C_BLUE     = "#4D9DE0"   # info
C_BLUE_L   = "#6AB2EC"
C_PURPLE   = "#9B72CF"   # keyword accent
C_PURPLE_L = "#AD87DC"
C_SUCCESS  = "#2ECC71"   # high score
C_LOG_BG   = "#050912"   # log terminal bg
C_LOG_FG   = "#CBD5E1"   # log text

# ── Button-specific color pairs (bg + WCAG AA-compliant fg) ───────────────────
B_GREEN    = "#1DBF73";  B_GREEN_FG  = "#063319"   # dark on Fiverr green  7.97:1
B_BLUE     = "#2878C8";  B_BLUE_FG   = "#FFFFFF"   # white on darkened blue 4.55:1
B_PURPLE   = "#7A52B8";  B_PURPLE_FG = "#FFFFFF"   # white on deep purple   5.64:1
B_RED      = "#C8303F";  B_RED_FG    = "#FFFFFF"   # white on darkened red  5.22:1
B_YELLOW   = "#F4C842";  B_YELLOW_FG = "#2A1D00"   # dark on amber          8.39:1
B_GHOST    = "#1F2D44";  B_GHOST_FG  = "#BDD1E4"   # light on dark ghost    6.50:1

if _PLATFORM == 'Windows':
    _FF, _FM, _FS = 'Segoe UI', 'Consolas', 9
elif _PLATFORM == 'Darwin':
    _FF, _FM, _FS = 'SF Pro Display', 'Menlo', 11
else:  # Linux and others
    _FF, _FM, _FS = 'DejaVu Sans', 'DejaVu Sans Mono', 10

FT_BRAND  = (_FF, _FS + 8, 'bold')
FT_H1     = (_FF, _FS + 4, 'bold')
FT_H2     = (_FF, _FS + 2, 'bold')
FT_H3     = (_FF, _FS,     'bold')
FT_BODY   = (_FF, _FS)
FT_SMALL  = (_FF, _FS - 1)
FT_TINY   = (_FF, _FS - 2)
FT_MONO   = (_FM, _FS - 1)
FT_MONO_S = (_FM, _FS - 2)

# ── Tool registry ─────────────────────────────────────────────────────────────
TOOLS = [
    {"id": "create_excel",  "title": "Create / Regenerate Excel",
     "icon": "📊", "group": "Setup",
     "desc": "Uses AI to generate a fresh, dynamic list of Fiverr sub-niches and builds the research spreadsheet. WARNING: overwrites existing file.",
     "script": "create_comprehensive_excel.py", "color": B_BLUE, "btn_fg": B_BLUE_FG,
     "warn": "This will OVERWRITE the existing Excel file. All scraping progress will be lost. Continue?"},
    {"id": "niche_scraper", "title": "Niche Scraper  (Full Run)",
     "icon": "🔍", "group": "Scrapers",
     "desc": "Searches Fiverr for every sub-niche. Extracts gig count, avg reviews, avg queue. Marks Positive (≤ threshold) or Negative. Auto-skips completed rows.",
     "script": "fiverr_scraper.py", "color": B_GREEN, "btn_fg": B_GREEN_FG, "warn": None},
    {"id": "gig_details",   "title": "Gig Details Scraper",
     "icon": "🎯", "group": "Scrapers",
     "desc": "For every Positive row: picks the best-scoring gig and downloads title, description, URL, pricing, and gallery images into the Excel.",
     "script": "fiverr_gig_details_scraper.py", "color": B_GREEN, "btn_fg": B_GREEN_FG, "warn": None},
    {"id": "top_sellers",   "title": "Top Sellers Analyzer",
     "icon": "🏆", "group": "Scrapers",
     "desc": "Inspects the top 10 gig cards per Positive niche to measure seller level mix (Top Rated / L2 / New). Calculates your entry difficulty.",
     "script": "fiverr_top_sellers.py", "color": B_GREEN, "btn_fg": B_GREEN_FG, "warn": None},
    {"id": "keyword_vars",  "title": "Keyword Variation Finder",
     "icon": "🔑", "group": "Scrapers",
     "desc": "Tests up to 10 alternative keyword phrasings per Positive niche to find lower-competition entry points. Saves results to a new Excel sheet + CSV.",
     "script": "fiverr_keyword_variations.py", "color": B_PURPLE, "btn_fg": B_PURPLE_FG, "warn": None},
    {"id": "test_scraper",  "title": "Quick Test  (3 Niches)",
     "icon": "🧪", "group": "Debug",
     "desc": "Runs a 3-niche test to confirm the scraper is working. No Excel changes — results shown in the log only.",
     "script": "test_scraper_3rows.py", "color": B_YELLOW, "btn_fg": B_YELLOW_FG, "warn": None},
    {"id": "debug_page",    "title": "Debug Page Inspector",
     "icon": "🛠", "group": "Debug",
     "desc": "Opens a live browser window to inspect Fiverr page HTML structure. Useful for diagnosing scraper failures.",
     "script": "debug_page.py", "color": B_GHOST, "btn_fg": B_GHOST_FG, "warn": None},
    {"id": "niche_discovery", "title": "Niche Discovery (from Scraper page)",
     "icon": "🔭", "group": "Discovery",
     "desc": "Launch from Niche Manager → Discover. Searches Fiverr autocomplete and category pages for new sub-niche opportunities.",
     "script": "fiverr_niche_discovery.py", "color": B_PURPLE, "btn_fg": B_PURPLE_FG, "warn": None},
    {"id": "trending_scan",  "title": "Trending Services Scanner",
     "icon": "📡", "group": "Discovery",
     "desc": "Scans Fiverr homepage and category pages to capture currently popular/trending services. Results appear in the Trends page.",
     "script": "fiverr_trending.py", "color": B_PURPLE, "btn_fg": B_PURPLE_FG, "warn": None},
]

GROUP_COLORS = {"Setup": B_BLUE, "Scrapers": B_GREEN, "Debug": C_MUTED, "Discovery": B_PURPLE}

# ── Colour helpers ────────────────────────────────────────────────────────────
def _adj(hex_col, factor):
    """Lighten (>1) or darken (<1) a hex colour."""
    h = hex_col.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = min(255, max(0, int(r * factor)))
    g = min(255, max(0, int(g * factor)))
    b = min(255, max(0, int(b * factor)))
    return f"#{r:02x}{g:02x}{b:02x}"

def score_color(s):
    if s is None: return C_MUTED
    if s >= 70:   return C_SUCCESS
    if s >= 50:   return "#6CCB5A"
    if s >= 35:   return C_YELLOW
    return C_RED

# ── Widget helpers ────────────────────────────────────────────────────────────
def _btn(parent, text, cmd, bg, fg="white", padx=12, pady=6, font=None, **kw):
    font = font or FT_SMALL
    hover = _adj(bg, 1.15)
    b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg, font=font,
                  relief="flat", padx=padx, pady=pady, cursor="hand2",
                  activebackground=hover, activeforeground=fg,
                  bd=0, highlightthickness=0, **kw)
    b.bind("<Enter>", lambda e: b.config(bg=hover))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

def _label(parent, text, font=None, fg=None, bg=None, **kw):
    return tk.Label(parent, text=text, font=font or FT_BODY,
                    fg=fg or C_TEXT, bg=bg or C_CARD, **kw)

def _sep(parent, bg=C_LINE, height=1, pady=6):
    f = tk.Frame(parent, bg=bg, height=height)
    f.pack(fill="x", pady=pady)
    return f

def _card(parent, bg=C_CARD, pad_x=14, pad_y=12, **kw):
    outer = tk.Frame(parent, bg=C_BORDER, **kw)
    inner = tk.Frame(outer, bg=bg)
    inner.pack(fill="both", expand=True, padx=1, pady=1)
    return outer, inner

# ── Business logic helpers ────────────────────────────────────────────────────
def load_settings():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                return {**DEFAULT_SETTINGS, **json.load(f)}
        except Exception:
            pass
    return dict(DEFAULT_SETTINGS)

def save_settings(settings):
    with open(CONFIG_FILE, "w") as f:
        json.dump(settings, f, indent=2)

def calc_opp_score(gig_count, avg_queue, avg_reviews, threshold=1800):
    score = 0.0
    if gig_count   is not None: score += max(0.0, 1.0 - gig_count / threshold) * 40
    if avg_queue   is not None: score += min(float(avg_queue) / 3.0, 1.0) * 35
    if avg_reviews is not None:
        r = float(avg_reviews)
        score += (8 if r < 20 else 16 if r < 50 else 25 if r < 150 else 15 if r < 300 else 6)
    return round(score, 1)

def read_excel_data():
    if not os.path.exists(EXCEL_PATH):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        threshold = 1800
        try:
            with open(CONFIG_FILE) as cf:
                threshold = json.load(cf).get("threshold", 1800)
        except Exception:
            pass
        rows = []
        for r in range(2, ws.max_row + 1):
            sub = ws.cell(r, 3).value
            if not sub or not str(sub).strip():
                continue
            def _v(c): v = ws.cell(r, c).value; return str(v).strip() if v is not None else ""
            def _n(c):
                v = ws.cell(r, c).value
                try:    return float(v)
                except: return None
            gig_count = _n(9); avg_reviews = _n(10); avg_queue = _n(11)
            verdict   = _v(4)
            score     = calc_opp_score(gig_count, avg_queue, avg_reviews, threshold) \
                        if verdict.lower() == "positive" else None
            rows.append({
                "row": r, "section": _v(1), "niche": _v(2), "sub": str(sub).strip(),
                "verdict": verdict, "gig": int(gig_count) if gig_count is not None else None,
                "reviews": avg_reviews, "queue": avg_queue, "score": score,
                "title": _v(6), "price": _v(10), "entry": _v(13),
                "url": _v(9) if ws.cell(r, 9).value and "http" in str(ws.cell(r,9).value) else "",
            })
        wb.close()
        return rows
    except Exception:
        return None

def get_quick_stats(rows):
    if rows is None: return None
    total   = len(rows)
    pos     = sum(1 for r in rows if r["verdict"].lower() == "positive")
    neg     = sum(1 for r in rows if r["verdict"].lower() == "negative")
    na      = total - pos - neg
    details = sum(1 for r in rows if r["url"])
    top5    = sorted([r for r in rows if r["score"] is not None],
                     key=lambda x: x["score"], reverse=True)[:5]
    return dict(total=total, pos=pos, neg=neg, na=na, details=details, top_opps=top5)


# ═══════════════════════════════════════════════════════════════════════════════
#  Main Application
# ═══════════════════════════════════════════════════════════════════════════════
class FiverrHub(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Fiverr Research Hub")
        self.configure(bg=C_BG)
        self.geometry("1260x830")
        self.minsize(1050, 680)

        self._settings      = load_settings()
        self._proc          = None
        self._log_queue     = queue.Queue()
        self._active_tool   = None
        self._analysis_data = []
        self._current_page  = None
        self._nav_items     = {}  # page → {outer, accent, btn}

        self._configure_ttk_styles()
        self._build_ui()
        self._show_page("dashboard")
        self._poll_log()
        self.after(400, self._check_api_key_on_startup)

    # ── TTK styling (cross-platform) ───────────────────────────────────────────
    def _configure_ttk_styles(self):
        style = ttk.Style()
        if _PLATFORM != 'Darwin':
            style.theme_use('clam')

        style.configure("Hub.Treeview",
                        background=C_CARD, foreground=C_TEXT2, rowheight=26,
                        fieldbackground=C_CARD, font=FT_SMALL, borderwidth=0)
        style.configure("Hub.Treeview.Heading",
                        background=C_PANEL, foreground=C_GREEN,
                        font=FT_H3, relief="flat", borderwidth=0)
        style.map("Hub.Treeview",
                  background=[("selected", _adj(C_BORDER, 1.3))],
                  foreground=[("selected", C_TEXT)])

        for orient in ("Vertical", "Horizontal"):
            style.configure(f"{orient}.TScrollbar",
                            background=C_CARD_H, troughcolor=C_BG,
                            arrowcolor=C_MUTED, bordercolor=C_BG,
                            darkcolor=C_CARD, lightcolor=C_CARD)
            style.map(f"{orient}.TScrollbar",
                      background=[("active", C_BORDER), ("pressed", C_BORDER)])

        style.configure("TCombobox",
                        fieldbackground=C_BG, background=C_CARD_H,
                        foreground=C_TEXT, selectbackground=C_BORDER,
                        selectforeground=C_TEXT, bordercolor=C_BORDER,
                        arrowcolor=C_MUTED)
        style.map("TCombobox",
                  fieldbackground=[("readonly", C_BG)],
                  foreground=[("readonly", C_TEXT)],
                  selectbackground=[("readonly", C_BORDER)])

        self.option_add('*TCombobox*Listbox.background',      C_CARD)
        self.option_add('*TCombobox*Listbox.foreground',      C_TEXT)
        self.option_add('*TCombobox*Listbox.selectBackground', C_BORDER)
        self.option_add('*TCombobox*Listbox.selectForeground', C_TEXT)

    # ── UI skeleton ────────────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        body = tk.Frame(self, bg=C_BG)
        body.pack(fill="both", expand=True)
        self._build_sidebar(body)
        self._content = tk.Frame(body, bg=C_BG)
        self._content.pack(side="right", fill="both", expand=True)
        self._pages = {
            "dashboard":     self._build_dashboard_page(),
            "scrapers":      self._build_scrapers_page(),
            "analysis":      self._build_analysis_page(),
            "niche_manager": self._build_niche_manager_page(),
            "trends":        self._build_trends_page(),
            "settings":      self._build_settings_page(),
        }
        for frame in self._pages.values():
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            frame.lower()
        self._build_status_bar()

    # ── Header ─────────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self, bg=C_PANEL, height=58)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        # Green left accent stripe
        tk.Frame(hdr, bg=C_GREEN, width=4).pack(side="left", fill="y")

        # Brand
        brand = tk.Frame(hdr, bg=C_PANEL)
        brand.pack(side="left", padx=(14, 0))
        tk.Label(brand, text="Fiverr", font=(_FF, 19, "bold"),
                 bg=C_PANEL, fg=C_GREEN).pack(side="left")
        tk.Label(brand, text=" Research Hub", font=(_FF, 19, "bold"),
                 bg=C_PANEL, fg=C_TEXT).pack(side="left")
        tk.Label(hdr, text="  Niche Intelligence Platform",
                 font=FT_SMALL, bg=C_PANEL, fg=C_MUTED).pack(side="left")

        # Header buttons
        btn_row = tk.Frame(hdr, bg=C_PANEL)
        btn_row.pack(side="right", padx=16)
        for text, bg, fg, cmd in [
            ("⟳  Refresh",    C_PANEL,  C_TEXT2,     self._refresh_all),
            ("🖼  Gig Images", B_BLUE,   B_BLUE_FG,   self._open_images),
            ("📄  Open Excel", B_GREEN,  B_GREEN_FG,  self._open_excel),
        ]:
            b = _btn(btn_row, text, cmd, bg=bg, fg=fg, padx=14, pady=7, font=FT_SMALL)
            b.pack(side="right", padx=3)

    # ── Sidebar ────────────────────────────────────────────────────────────────
    def _build_sidebar(self, parent):
        sb = tk.Frame(parent, bg=C_PANEL, width=195)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        tk.Frame(sb, bg=C_LINE, height=1).pack(fill="x")

        nav_items = [
            ("dashboard",     "🏠", "Dashboard"),
            ("scrapers",      "⚙️", "Scrapers"),
            ("analysis",      "📊", "Analysis"),
            ("niche_manager", "🗃", "Niche Manager"),
            ("trends",        "📈", "Trends"),
            ("settings",      "⚙",  "Settings"),
        ]

        for page, icon, label in nav_items:
            outer = tk.Frame(sb, bg=C_PANEL)
            outer.pack(fill="x")

            accent = tk.Frame(outer, bg=C_PANEL, width=4)
            accent.pack(side="left", fill="y")

            btn = tk.Button(outer, text=f"  {icon}  {label}", font=FT_H3,
                            bg=C_PANEL, fg=C_MUTED, relief="flat",
                            anchor="w", padx=10, pady=12, cursor="hand2",
                            activebackground=C_CARD_H, activeforeground=C_TEXT,
                            command=lambda p=page: self._show_page(p))
            btn.pack(side="left", fill="x", expand=True)

            def _enter(e, b=btn, p=page):
                if self._current_page != p:
                    b.config(bg=C_CARD, fg=C_TEXT2)
            def _leave(e, b=btn, p=page):
                if self._current_page != p:
                    b.config(bg=C_PANEL, fg=C_MUTED)
            btn.bind("<Enter>", _enter)
            btn.bind("<Leave>", _leave)

            self._nav_items[page] = {"outer": outer, "accent": accent, "btn": btn}

        tk.Frame(sb, bg=C_LINE, height=1).pack(fill="x", pady=(8, 4))

        # Live mini-stats
        self._sb_stats = tk.Label(sb, text="Loading…", font=FT_TINY,
                                  bg=C_PANEL, fg=C_MUTED,
                                  justify="left", wraplength=170, anchor="w")
        self._sb_stats.pack(padx=16, anchor="w")

        # Version footer
        tk.Frame(sb, bg=C_PANEL).pack(fill="both", expand=True)
        tk.Frame(sb, bg=C_LINE, height=1).pack(fill="x")
        tk.Label(sb, text=f"v{APP_VERSION}  ·  Fiverr Research Hub",
                 font=FT_TINY, bg=C_PANEL, fg=C_MUTED).pack(pady=6)

    def _show_page(self, name):
        # Deactivate previous nav item
        if self._current_page and self._current_page in self._nav_items:
            old = self._nav_items[self._current_page]
            old["accent"].config(bg=C_PANEL)
            old["btn"].config(bg=C_PANEL, fg=C_MUTED)

        # Activate new nav item
        if name in self._nav_items:
            new = self._nav_items[name]
            new["accent"].config(bg=C_GREEN)
            new["btn"].config(bg=C_CARD_H, fg=C_GREEN, font=(_FF, 11, "bold"))

        # Swap page
        for pname, frame in self._pages.items():
            if pname == name: frame.lift()
            else:             frame.lower()

        self._current_page = name

        # On-show hooks
        if name == "dashboard":     self._refresh_dashboard()
        if name == "analysis":      self._load_analysis_data()
        if name == "niche_manager": self._load_niche_data()
        if name == "trends":        self._refresh_trends_page()

    # ── Status bar ─────────────────────────────────────────────────────────────
    def _build_status_bar(self):
        bar = tk.Frame(self, bg=C_PANEL, height=28)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Frame(bar, bg=C_LINE, height=1).pack(fill="x", side="top")
        self._status_var = tk.StringVar(value="Ready")
        tk.Label(bar, textvariable=self._status_var,
                 font=FT_TINY, bg=C_PANEL, fg=C_MUTED, anchor="w"
                 ).pack(side="left", padx=14)
        self._running_lbl = tk.Label(bar, text="", font=FT_TINY,
                                     bg=C_PANEL, fg=C_GREEN)
        self._running_lbl.pack(side="right", padx=14)

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    def _build_dashboard_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        # Title
        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 10))
        tk.Label(th, text="Dashboard", font=FT_H1, bg=C_BG, fg=C_TEXT).pack(side="left")
        _btn(th, "↻  Refresh", self._refresh_dashboard,
             bg=B_GHOST, fg=B_GHOST_FG, padx=12, pady=5).pack(side="right")

        # ── Stat cards row ──────────────────────────────────────────────────
        stat_row = tk.Frame(frame, bg=C_BG)
        stat_row.pack(fill="x", padx=20, pady=(0, 12))
        self._dash_stat_vars = {}

        stat_defs = [
            ("total",   "Total Niches",   C_BLUE,    "All research rows"),
            ("pos",     "Positive",        C_SUCCESS, "≤ threshold gigs"),
            ("neg",     "Negative",        C_RED,     "> threshold gigs"),
            ("na",      "Pending",         C_YELLOW,  "Not yet scraped"),
            ("details", "With Details",    C_PURPLE,  "Gig data collected"),
        ]
        for key, label, color, sub_lbl in stat_defs:
            outer, inner = _card(stat_row)
            outer.pack(side="left", fill="both", expand=True, padx=(0, 8))

            # Colored top accent bar
            tk.Frame(inner, bg=color, height=4).pack(fill="x")

            body = tk.Frame(inner, bg=C_CARD)
            body.pack(fill="both", expand=True, padx=16, pady=(12, 14))

            v = tk.StringVar(value="—")
            self._dash_stat_vars[key] = v
            tk.Label(body, textvariable=v, font=(_FF, 30, "bold"),
                     bg=C_CARD, fg=color, anchor="w").pack(anchor="w")
            tk.Label(body, text=label, font=FT_H3,
                     bg=C_CARD, fg=C_TEXT, anchor="w").pack(anchor="w")
            tk.Label(body, text=sub_lbl, font=FT_TINY,
                     bg=C_CARD, fg=C_MUTED, anchor="w").pack(anchor="w")

        # ── Bottom: left column (opportunities + progress) ──────────────────
        bottom = tk.Frame(frame, bg=C_BG)
        bottom.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        left_col = tk.Frame(bottom, bg=C_BG)
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Progress panel
        out_p, in_p = _card(left_col)
        out_p.pack(fill="x", pady=(0, 10))
        tk.Label(in_p, text="Scraping Progress", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT, anchor="w").pack(anchor="w", padx=14, pady=(12, 8))

        self._prog_bars = {}
        for key, label, color in [
            ("scraped",   "Niches Scraped",    C_GREEN),
            ("details",   "Details Collected", C_BLUE),
            ("positive",  "Positive Rate",     C_YELLOW),
        ]:
            row = tk.Frame(in_p, bg=C_CARD)
            row.pack(fill="x", padx=14, pady=3)
            tk.Label(row, text=label, font=FT_SMALL, bg=C_CARD, fg=C_TEXT2,
                     width=20, anchor="w").pack(side="left")
            track = tk.Canvas(row, height=8, bg=C_BG, highlightthickness=0)
            track.pack(side="left", fill="x", expand=True, padx=(4, 8))
            pct_lbl = tk.Label(row, text="0%", font=FT_TINY,
                               bg=C_CARD, fg=C_MUTED, width=4)
            pct_lbl.pack(side="right")
            self._prog_bars[key] = (track, pct_lbl, color)
        tk.Frame(in_p, bg=C_BG, height=8).pack()

        # Top opportunities
        out_o, in_o = _card(left_col)
        out_o.pack(fill="both", expand=True)
        hdr_o = tk.Frame(in_o, bg=C_CARD)
        hdr_o.pack(fill="x", padx=14, pady=(12, 6))
        tk.Label(hdr_o, text="Top Opportunities", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")
        tk.Label(hdr_o, text="by score →", font=FT_TINY,
                 bg=C_CARD, fg=C_MUTED).pack(side="left", padx=6)
        _btn(hdr_o, "→ Open Analysis", self._go_analysis,
             bg=C_GREEN_BG, fg=C_GREEN, padx=10, pady=3).pack(side="right")

        self._opp_frame = tk.Frame(in_o, bg=C_CARD)
        self._opp_frame.pack(fill="both", expand=True, padx=14, pady=(0, 12))

        # ── Right column: chart + next step ────────────────────────────────
        right_col = tk.Frame(bottom, bg=C_BG, width=290)
        right_col.pack(side="right", fill="y")
        right_col.pack_propagate(False)

        out_c, in_c = _card(right_col)
        out_c.pack(fill="x")
        tk.Label(in_c, text="Market Distribution", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT, anchor="w").pack(anchor="w", padx=14, pady=(12, 4))
        self._chart_canvas = tk.Canvas(in_c, bg=C_CARD, highlightthickness=0,
                                       width=270, height=230)
        self._chart_canvas.pack(padx=14, pady=(0, 12))

        # Next step
        out_n, in_n = _card(right_col)
        out_n.pack(fill="x", pady=(10, 0))
        tk.Label(in_n, text="Recommended Next Step", font=FT_H3,
                 bg=C_CARD, fg=C_TEXT, anchor="w").pack(anchor="w", padx=14, pady=(10, 2))
        self._next_lbl = tk.Label(in_n, text="", font=FT_SMALL, bg=C_CARD,
                                  fg=C_TEXT2, wraplength=262, justify="left")
        self._next_lbl.pack(anchor="w", padx=14, pady=(0, 12))

        return frame

    def _go_analysis(self):
        self._show_page("analysis")

    def _refresh_dashboard(self):
        rows  = read_excel_data()
        stats = get_quick_stats(rows)

        if stats is None:
            for v in self._dash_stat_vars.values(): v.set("—")
            self._sb_stats.config(text="Excel not found\nRun 'Create Excel'")
            self._next_lbl.config(
                text="Step 1 → Go to Scrapers and run 'Create / Regenerate Excel' first.")
            return

        self._dash_stat_vars["total"].set(str(stats["total"]))
        self._dash_stat_vars["pos"].set(str(stats["pos"]))
        self._dash_stat_vars["neg"].set(str(stats["neg"]))
        self._dash_stat_vars["na"].set(str(stats["na"]))
        self._dash_stat_vars["details"].set(str(stats["details"]))
        self._sb_stats.config(
            text=f"Total: {stats['total']}\nPos: {stats['pos']}\n"
                 f"Neg: {stats['neg']}\nPending: {stats['na']}")

        # Progress bars
        total = max(stats["total"], 1)
        pcts  = {
            "scraped":  round((stats["pos"] + stats["neg"]) / total * 100),
            "details":  round(stats["details"] / max(stats["pos"], 1) * 100),
            "positive": round(stats["pos"] / max(stats["pos"] + stats["neg"], 1) * 100),
        }
        for key, (track, lbl, color) in self._prog_bars.items():
            track.update_idletasks()
            w = max(track.winfo_width(), 150)
            h = 8
            pct = pcts[key]
            track.delete("all")
            track.create_rectangle(0, 0, w, h, fill=_adj(C_BG, 1.5), outline="")
            bar_w = int(w * pct / 100)
            if bar_w > 0:
                track.create_rectangle(0, 0, bar_w, h, fill=color, outline="")
            lbl.config(text=f"{pct}%", fg=color if pct > 0 else C_MUTED)

        # Top opportunities table
        for c in self._opp_frame.winfo_children():
            c.destroy()

        if not stats["top_opps"]:
            tk.Label(self._opp_frame, text="Run the Niche Scraper to see opportunities.",
                     font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(anchor="w", pady=8)
        else:
            # Header row
            hr = tk.Frame(self._opp_frame, bg=C_CARD)
            hr.pack(fill="x", pady=(0, 4))
            for h, w in [("#", 3), ("Sub-Niche", 28), ("Gigs", 9), ("Score", 7)]:
                tk.Label(hr, text=h, font=FT_TINY, bg=C_CARD, fg=C_MUTED,
                         width=w, anchor="w").pack(side="left")

            for i, r in enumerate(stats["top_opps"], 1):
                row_bg = _adj(C_CARD, 1.06) if i % 2 == 0 else C_CARD
                rrow = tk.Frame(self._opp_frame, bg=row_bg)
                rrow.pack(fill="x")
                sc  = r["score"] or 0
                col = score_color(sc)
                vals = [
                    (f"{i}", 3, C_GREEN),
                    (r["sub"][:32], 28, C_TEXT),
                    (f"{r['gig']:,}" if r["gig"] else "—", 9, C_TEXT2),
                    (f"{sc:.0f}", 7, col),
                ]
                for val, w, fg in vals:
                    tk.Label(rrow, text=val, font=FT_SMALL, bg=row_bg,
                             fg=fg, width=w, anchor="w").pack(side="left", padx=1)

        # Donut chart
        self._draw_donut(stats)

        # Next step
        if stats["na"] > 0:
            self._next_lbl.config(
                text=f"Run Niche Scraper → {stats['na']} niches still need gig count data.")
        elif stats["details"] < stats["pos"]:
            self._next_lbl.config(
                text=f"Run Gig Details Scraper → {stats['pos'] - stats['details']} "
                     f"Positive niches need title / price / images.")
        else:
            self._next_lbl.config(
                text="All data collected!  Check the Analysis page for opportunity scores and recommendations.")

    def _draw_donut(self, stats):
        c  = self._chart_canvas
        c.delete("all")
        cx, cy = 135, 105
        ro, ri = 88, 50   # outer / inner radius

        segs = [
            (stats["pos"], C_SUCCESS,  "Positive"),
            (stats["neg"], C_RED,      "Negative"),
            (stats["na"],  C_YELLOW,   "Pending"),
        ]
        total = max(sum(v for v, _, _ in segs), 1)

        # Outer ring background
        c.create_oval(cx-ro, cy-ro, cx+ro, cy+ro,
                      fill=_adj(C_CARD, 0.6), outline=C_BORDER, width=1)

        start = -90.0
        for value, color, label in segs:
            if value == 0:
                continue
            extent = 360.0 * value / total
            c.create_arc(cx-ro, cy-ro, cx+ro, cy+ro,
                         start=start, extent=extent,
                         fill=color, outline=C_CARD, width=3, style=tk.PIESLICE)
            start += extent

        # Inner hole
        c.create_oval(cx-ri, cy-ri, cx+ri, cy+ri, fill=C_CARD, outline=C_CARD)

        # Centre text
        c.create_text(cx, cy-10, text=str(stats["total"]),
                      font=(_FF, 26, "bold"), fill=C_TEXT)
        c.create_text(cx, cy+14, text="niches",
                      font=FT_TINY, fill=C_MUTED)

        # Legend
        ly = cy + ro + 16
        for value, color, label in segs:
            pct = round(value / total * 100)
            c.create_rectangle(cx-76, ly, cx-66, ly+9, fill=color, outline="")
            c.create_text(cx-60, ly+4.5,
                          text=f"{label}  {value}  ({pct}%)",
                          font=FT_TINY, fill=C_TEXT2, anchor="w")
            ly += 17

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Scrapers
    # ══════════════════════════════════════════════════════════════════════════
    def _build_scrapers_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 8))
        tk.Label(th, text="Scrapers & Tools", font=FT_H1,
                 bg=C_BG, fg=C_TEXT).pack(side="left")

        body = tk.Frame(frame, bg=C_BG)
        body.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        # ── Left: tool cards ──────────────────────────────────────────────
        left = tk.Frame(body, bg=C_BG, width=430)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)

        canvas_scroll = tk.Canvas(left, bg=C_BG, highlightthickness=0)
        vsb = tk.Scrollbar(left, orient="vertical", command=canvas_scroll.yview)
        inner = tk.Frame(canvas_scroll, bg=C_BG)
        inner.bind("<Configure>",
                   lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
        canvas_scroll.create_window((0, 0), window=inner, anchor="nw")
        canvas_scroll.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas_scroll.pack(side="left", fill="both", expand=True)
        canvas_scroll.bind("<MouseWheel>",
                           lambda e: canvas_scroll.yview_scroll(-1*(e.delta//120), "units"))
        canvas_scroll.bind("<Button-4>", lambda e: canvas_scroll.yview_scroll(-1, "units"))
        canvas_scroll.bind("<Button-5>", lambda e: canvas_scroll.yview_scroll(1, "units"))

        self._tool_buttons = {}
        groups_seen = []
        for tool in TOOLS:
            g = tool["group"]
            if g not in groups_seen:
                groups_seen.append(g)
                # Group header
                gh = tk.Frame(inner, bg=C_BG)
                gh.pack(fill="x", pady=(10, 4))
                gc = GROUP_COLORS.get(g, C_MUTED)
                tk.Frame(gh, bg=gc, width=3, height=16).pack(side="left", padx=(0, 8))
                tk.Label(gh, text=g.upper(), font=FT_TINY,
                         bg=C_BG, fg=gc).pack(side="left")

            # Tool card
            outer, card = _card(inner)
            outer.pack(fill="x", pady=(0, 6))

            # Left color accent bar matching group color
            acbar = tk.Frame(card, bg=tool["color"], width=3)
            acbar.pack(side="left", fill="y")

            content = tk.Frame(card, bg=C_CARD)
            content.pack(side="left", fill="both", expand=True, padx=12, pady=10)

            top = tk.Frame(content, bg=C_CARD)
            top.pack(fill="x")
            tk.Label(top, text=f"{tool['icon']}  {tool['title']}",
                     font=FT_H3, bg=C_CARD, fg=C_TEXT, anchor="w"
                     ).pack(side="left", fill="x", expand=True)

            btn = _btn(top, "▶  Run", lambda t=tool: self._run_tool(t),
                       bg=tool["color"], fg=tool.get("btn_fg", "white"), padx=14, pady=5)
            btn.pack(side="right")
            self._tool_buttons[tool["id"]] = btn

            tk.Label(content, text=tool["desc"], font=FT_SMALL,
                     bg=C_CARD, fg=C_MUTED, wraplength=360,
                     justify="left", anchor="w").pack(anchor="w", pady=(5, 0))

        # ── Right: log panel ──────────────────────────────────────────────
        right = tk.Frame(body, bg=C_BG)
        right.pack(side="right", fill="both", expand=True)

        # Log header
        log_hdr = tk.Frame(right, bg=C_PANEL)
        log_hdr.pack(fill="x")
        tk.Frame(log_hdr, bg=C_LINE, height=1).pack(fill="x", side="bottom")

        self._log_title_lbl = tk.Label(log_hdr, text="  Output Log",
                                       font=FT_H3, bg=C_PANEL, fg=C_TEXT)
        self._log_title_lbl.pack(side="left", padx=8, pady=10)

        self._log_status_dot = tk.Label(log_hdr, text="●", font=FT_SMALL,
                                        bg=C_PANEL, fg=C_MUTED)
        self._log_status_dot.pack(side="left")

        ctl_row = tk.Frame(log_hdr, bg=C_PANEL)
        ctl_row.pack(side="right", padx=8)
        self._btn_stop = _btn(ctl_row, "■  Stop", self._stop_tool,
                              bg=B_RED, fg=B_RED_FG, padx=12, pady=5)
        self._btn_stop.config(state="disabled")
        self._btn_stop.pack(side="right", padx=(6, 0))
        _btn(ctl_row, "Clear", self._clear_log,
             bg=B_GHOST, fg=B_GHOST_FG, padx=10, pady=5).pack(side="right")

        # Log text widget
        log_body = tk.Frame(right, bg=C_LOG_BG)
        log_body.pack(fill="both", expand=True)

        self._log = tk.Text(log_body, font=FT_MONO, bg=C_LOG_BG, fg=C_LOG_FG,
                            insertbackground=C_GREEN, relief="flat",
                            wrap="word", state="disabled",
                            selectbackground=C_BORDER, padx=12, pady=8)
        lsb = tk.Scrollbar(log_body, orient="vertical", command=self._log.yview)
        self._log.configure(yscrollcommand=lsb.set)
        lsb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)

        tags = {
            "green":  C_GREEN,   "yellow": C_YELLOW, "red":    C_RED,
            "blue":   C_BLUE,    "purple": C_PURPLE,
            "muted":  C_MUTED,   "bold":   C_TEXT,
        }
        for tag, fg in tags.items():
            kw = {"foreground": fg}
            if tag == "bold":
                kw["font"] = (_FM, 10, "bold")
            self._log.tag_config(tag, **kw)

        return frame

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Analysis
    # ══════════════════════════════════════════════════════════════════════════
    def _build_analysis_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 8))
        tk.Label(th, text="Analysis & Opportunity Scores",
                 font=FT_H1, bg=C_BG, fg=C_TEXT).pack(side="left")

        # Filter / action bar
        fbar_outer, fbar = _card(frame)
        fbar_outer.pack(fill="x", padx=20, pady=(0, 8))

        # Row 1: search + filters
        r1 = tk.Frame(fbar, bg=C_CARD)
        r1.pack(fill="x", padx=12, pady=(10, 4))

        tk.Label(r1, text="🔍", font=FT_BODY, bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._apply_filter())
        se = tk.Entry(r1, textvariable=self._search_var, font=FT_BODY,
                      bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                      relief="flat", width=24)
        se.pack(side="left", padx=(4, 16))

        for lbl, var_name, opts in [
            ("Status:", "_filter_status", ["All", "Positive", "Negative", "Pending"]),
            ("Sort:",   "_sort_var",      ["Opp Score ↓", "Opp Score ↑",
                                           "Gig Count ↑", "Gig Count ↓",
                                           "Avg Queue ↓", "Avg Reviews ↓",
                                           "Sub-Niche A-Z"]),
        ]:
            tk.Label(r1, text=lbl, font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")
            var = tk.StringVar(value=opts[0])
            setattr(self, var_name, var)
            var.trace_add("write", lambda *_: self._apply_filter())
            cb = ttk.Combobox(r1, textvariable=var, values=opts,
                               state="readonly", width=14)
            cb.pack(side="left", padx=(2, 14))

        # Gig count range filter
        tk.Label(r1, text="Gigs:", font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._min_gigs_var = tk.StringVar()
        self._max_gigs_var = tk.StringVar()
        for var, placeholder in [(self._min_gigs_var, "min"), (self._max_gigs_var, "max")]:
            var.trace_add("write", lambda *_: self._apply_filter())
            e = tk.Entry(r1, textvariable=var, font=FT_SMALL,
                         bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                         relief="flat", width=6, highlightthickness=0)
            e.pack(side="left", padx=(2, 1))
            if placeholder == "min":
                tk.Label(r1, text="–", font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")

        # Row 2: action buttons
        r2 = tk.Frame(fbar, bg=C_CARD)
        r2.pack(fill="x", padx=12, pady=(0, 10))
        for text, bg, fg, cmd in [
            ("⟳ Load / Refresh",        B_GREEN,  B_GREEN_FG,  self._load_analysis_data),
            ("↓ Export CSV",            B_BLUE,   B_BLUE_FG,   self._export_csv),
            ("✎ Write Scores to Excel", B_PURPLE, B_PURPLE_FG, self._write_scores_to_excel),
        ]:
            _btn(r2, text, cmd, bg=bg, fg=fg, padx=12, pady=5).pack(side="left", padx=(0, 8))

        # Treeview
        tv_wrap = tk.Frame(frame, bg=C_BG)
        tv_wrap.pack(fill="both", expand=True, padx=20, pady=(0, 4))


        cols = ("sub","niche","status","gig","reviews","queue","score","entry")
        hdgs = ("Sub-Niche","Niche","Status","Gig Count",
                "Avg Reviews","Avg Queue","Opp Score","Entry Level")
        wids = (195, 130, 80, 90, 90, 85, 90, 90)

        self._tv = ttk.Treeview(tv_wrap, columns=cols, show="headings",
                                style="Hub.Treeview", selectmode="browse")
        for col, h, w in zip(cols, hdgs, wids):
            self._tv.heading(col, text=h, anchor="w")
            self._tv.column(col, width=w, anchor="w", minwidth=60)

        self._tv.tag_configure("positive", foreground=C_SUCCESS)
        self._tv.tag_configure("negative", foreground=C_RED)
        self._tv.tag_configure("na",       foreground=C_YELLOW)
        self._tv.tag_configure("score_hi", foreground=C_SUCCESS)
        self._tv.tag_configure("score_md", foreground=C_YELLOW)
        self._tv.tag_configure("score_lo", foreground=C_RED)

        vsb = ttk.Scrollbar(tv_wrap, orient="vertical",   command=self._tv.yview)
        hsb = ttk.Scrollbar(tv_wrap, orient="horizontal", command=self._tv.xview)
        self._tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tv_wrap.grid_rowconfigure(0, weight=1)
        tv_wrap.grid_columnconfigure(0, weight=1)

        self._tv.bind("<<TreeviewSelect>>", self._on_tv_select)

        # ── Detail panel ────────────────────────────────────────────────────
        det_outer, det = _card(frame)
        det_outer.pack(fill="x", padx=20, pady=(0, 12))

        det_top = tk.Frame(det, bg=C_CARD)
        det_top.pack(fill="x", padx=12, pady=(10, 4))
        tk.Label(det_top, text="Selected Row", font=FT_H3,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")
        self._det_copy_btn = _btn(det_top, "Copy Gig URL", self._copy_gig_url,
                                  bg=B_GHOST, fg=B_GHOST_FG, padx=10, pady=3)
        self._det_copy_btn.pack(side="right")
        self._det_open_btn = _btn(det_top, "Open in Browser", self._open_gig_url,
                                  bg=B_BLUE, fg=B_BLUE_FG, padx=10, pady=3)
        self._det_open_btn.pack(side="right", padx=(0, 6))

        det_body = tk.Frame(det, bg=C_CARD)
        det_body.pack(fill="x", padx=12, pady=(0, 10))

        self._det_vars = {}
        det_fields = [
            ("sub",     "Sub-Niche",    28),
            ("niche",   "Niche",        18),
            ("gig",     "Gig Count",    12),
            ("score",   "Opp Score",    12),
            ("entry",   "Entry Level",  12),
            ("price",   "Price",        10),
        ]
        for key, label, width in det_fields:
            col = tk.Frame(det_body, bg=C_CARD)
            col.pack(side="left", padx=(0, 20))
            tk.Label(col, text=label, font=FT_TINY, bg=C_CARD, fg=C_MUTED
                     ).pack(anchor="w")
            v = tk.StringVar(value="—")
            self._det_vars[key] = v
            tk.Label(col, textvariable=v, font=FT_H3,
                     bg=C_CARD, fg=C_TEXT, width=width, anchor="w"
                     ).pack(anchor="w")

        # Score bar in detail panel
        self._score_canvas = tk.Canvas(det_body, bg=C_CARD,
                                       width=160, height=44, highlightthickness=0)
        self._score_canvas.pack(side="right", padx=(0, 12))

        self._selected_url = ""
        self._analysis_info = tk.Label(frame, text="", font=FT_TINY,
                                       bg=C_PANEL, fg=C_MUTED, anchor="w")
        self._analysis_info.pack(fill="x", padx=20)

        return frame

    def _load_analysis_data(self):
        self._analysis_data = read_excel_data() or []
        self._apply_filter()

    def _apply_filter(self):
        search  = self._search_var.get().lower()
        status  = self._filter_status.get()
        sort_by = self._sort_var.get()
        data    = list(self._analysis_data)

        if status == "Positive": data = [r for r in data if r["verdict"].lower() == "positive"]
        elif status == "Negative": data = [r for r in data if r["verdict"].lower() == "negative"]
        elif status == "Pending":  data = [r for r in data
                                           if r["verdict"].lower() not in ("positive","negative")]
        if search:
            data = [r for r in data if search in r["sub"].lower()
                    or search in r["niche"].lower() or search in r["section"].lower()]

        # Gig count range filter
        try:
            min_g = int(self._min_gigs_var.get().strip())
        except Exception:
            min_g = None
        try:
            max_g = int(self._max_gigs_var.get().strip())
        except Exception:
            max_g = None
        if min_g is not None:
            data = [r for r in data if r["gig"] is not None and r["gig"] >= min_g]
        if max_g is not None:
            data = [r for r in data if r["gig"] is not None and r["gig"] <= max_g]

        def _k(r):
            if sort_by == "Opp Score ↓":    return -(r["score"] or -1)
            if sort_by == "Opp Score ↑":    return  (r["score"] or 999)
            if sort_by == "Gig Count ↑":    return  (r["gig"] or 999999)
            if sort_by == "Gig Count ↓":    return -(r["gig"] or 0)
            if sort_by == "Avg Queue ↓":    return -(r["queue"] or 0)
            if sort_by == "Avg Reviews ↓":  return -(r["reviews"] or 0)
            return r["sub"].lower()
        data.sort(key=_k)

        for row in self._tv.get_children():
            self._tv.delete(row)

        for r in data:
            v  = r["verdict"].lower()
            sc = r["score"]
            tag = ("positive" if v == "positive" else
                   "negative" if v == "negative" else "na")
            self._tv.insert("", "end", values=(
                r["sub"], r["niche"], r["verdict"] or "—",
                f"{r['gig']:,}" if r["gig"] is not None else "—",
                f"{r['reviews']:.1f}" if r["reviews"] is not None else "—",
                f"{r['queue']:.1f}"   if r["queue"]   is not None else "—",
                f"{sc:.0f}"           if sc            is not None else "—",
                r["entry"] or "—",
            ), tags=(tag,))

        total   = len(self._analysis_data)
        showing = len(data)
        pos_d   = [r for r in data if r["score"] is not None]
        avg_s   = (sum(r["score"] for r in pos_d) / len(pos_d)) if pos_d else 0
        self._analysis_info.config(
            text=f"  Showing {showing} / {total} rows  |  "
                 f"Avg Opportunity Score (positive, visible): {avg_s:.1f}")

    def _on_tv_select(self, _event):
        sel = self._tv.selection()
        if not sel:
            return
        vals = self._tv.item(sel[0], "values")
        # vals order: sub, niche, status, gig, reviews, queue, score, entry
        col_map = {"sub": 0, "niche": 1, "gig": 3, "score": 6, "entry": 7}
        for key, idx in col_map.items():
            v = vals[idx] if idx < len(vals) else "—"
            self._det_vars[key].set(v)

        # Find full row data for price + url
        sub_name = vals[0] if vals else ""
        found    = next((r for r in self._analysis_data if r["sub"] == sub_name), None)
        if found:
            self._det_vars["price"].set(found["price"] or "—")
            self._selected_url = found["url"]
        else:
            self._det_vars["price"].set("—")
            self._selected_url = ""

        # Score bar
        try:
            score_val = float(vals[6]) if vals[6] != "—" else None
        except Exception:
            score_val = None
        self._draw_score_bar(score_val)

    def _draw_score_bar(self, score):
        c = self._score_canvas
        c.delete("all")
        if score is None:
            c.create_text(80, 22, text="No Score", font=FT_TINY, fill=C_MUTED)
            return
        col = score_color(score)
        pct = score / 100.0
        c.create_text(80, 10, text=f"Score: {score:.0f} / 100",
                      font=FT_TINY, fill=C_MUTED)
        c.create_rectangle(0, 22, 160, 34, fill=_adj(C_BG, 1.5), outline="")
        bw = int(160 * pct)
        if bw > 0:
            c.create_rectangle(0, 22, bw, 34, fill=col, outline="")
        label = ("Excellent" if score >= 70 else "Good" if score >= 50 else
                 "Moderate" if score >= 35 else "Tough")
        c.create_text(80, 42, text=label, font=FT_TINY, fill=col)

    def _copy_gig_url(self):
        if self._selected_url:
            self.clipboard_clear()
            self.clipboard_append(self._selected_url)
            self._set_status("Gig URL copied to clipboard")
        else:
            self._set_status("No URL available for this row")

    def _open_gig_url(self):
        if self._selected_url:
            webbrowser.open(self._selected_url)
        else:
            self._set_status("No URL available — run Gig Details Scraper first")

    def _export_csv(self):
        if not self._analysis_data:
            messagebox.showwarning("No Data", "Load data first.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile="fiverr_analysis_export.csv", parent=self)
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Section","Niche","Sub-Niche","Status","Gig Count",
                        "Avg Reviews","Avg Queue","Opp Score","Entry Level",
                        "Gig URL","Price"])
            for r in self._analysis_data:
                w.writerow([r["section"], r["niche"], r["sub"], r["verdict"],
                            r["gig"], r["reviews"], r["queue"], r["score"],
                            r["entry"], r["url"], r["price"]])
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=self)

    def _write_scores_to_excel(self):
        if not self._analysis_data:
            messagebox.showwarning("No Data", "Load data first.", parent=self)
            return
        if not messagebox.askyesno(
            "Write Scores", "Write Opportunity Scores to Excel column N?", parent=self):
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            thin = Side(style="thin", color="BDBDBD")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
            h    = ws.cell(row=1, column=14)
            if not h.value:
                h.value     = "Opp Score"
                h.fill      = PatternFill("solid", fgColor="1DBF73")
                h.font      = Font(bold=True, color="FFFFFF", size=11)
                h.alignment = Alignment(horizontal="center", vertical="center")
                h.border    = brd
            for r in self._analysis_data:
                if r["score"] is not None:
                    cell = ws.cell(row=r["row"], column=14, value=round(r["score"], 1))
                    cell.border    = brd
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    sc = r["score"]
                    fg = "C8E6C9" if sc >= 70 else "FFFDE7" if sc >= 40 else "FFCDD2"
                    cell.fill = PatternFill("solid", fgColor=fg)
            wb.save(EXCEL_PATH)
            messagebox.showinfo("Done", "Opportunity scores written to Excel (col N).", parent=self)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Trends
    # ══════════════════════════════════════════════════════════════════════════
    def _build_trends_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 4))
        tk.Label(th, text="Trends & Market Insights", font=FT_H1,
                 bg=C_BG, fg=C_TEXT).pack(side="left")
        _btn(th, "↻  Refresh", self._refresh_trends_page,
             bg=B_GHOST, fg=B_GHOST_FG, padx=10, pady=5).pack(side="right")

        body = tk.Frame(frame, bg=C_BG)
        body.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        # ── Left column: snapshots + market insights ──────────────────────
        left = tk.Frame(body, bg=C_BG, width=340)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)

        # Snapshot card
        out_s, in_s = _card(left)
        out_s.pack(fill="x", pady=(0, 8))

        sh = tk.Frame(in_s, bg=C_CARD)
        sh.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(sh, text="Snapshots", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")
        _btn(sh, "📸  Save Now", self._save_snapshot,
             bg=B_GREEN, fg=B_GREEN_FG, padx=12, pady=5).pack(side="right")

        self._snap_count_lbl = tk.Label(in_s, text="No snapshots yet",
                                        font=FT_TINY, bg=C_CARD, fg=C_MUTED)
        self._snap_count_lbl.pack(anchor="w", padx=12)
        _sep(in_s, pady=4)

        list_f = tk.Frame(in_s, bg=C_BG)
        list_f.pack(fill="x", padx=8, pady=(0, 4))
        self._snap_listbox = tk.Listbox(list_f, font=FT_SMALL,
                                        bg=C_BG, fg=C_TEXT2,
                                        selectbackground=_adj(C_BORDER, 1.4),
                                        selectforeground=C_TEXT,
                                        relief="flat", borderwidth=0,
                                        height=6, activestyle="none", cursor="hand2")
        lsb2 = tk.Scrollbar(list_f, orient="vertical",
                             command=self._snap_listbox.yview)
        self._snap_listbox.configure(yscrollcommand=lsb2.set)
        lsb2.pack(side="right", fill="y")
        self._snap_listbox.pack(fill="both", expand=True)

        btns_row = tk.Frame(in_s, bg=C_CARD)
        btns_row.pack(fill="x", padx=12, pady=(4, 12))
        _btn(btns_row, "⇄  Compare Latest 2", self._compare_latest_two,
             bg=B_BLUE, fg=B_BLUE_FG, padx=12, pady=5).pack(side="left")
        _btn(btns_row, "🗑  Delete", self._delete_snapshot,
             bg=B_RED, fg=B_RED_FG, padx=12, pady=5).pack(side="right")

        # Market Insights card
        out_mi, in_mi = _card(left)
        out_mi.pack(fill="both", expand=True)

        mi_hdr = tk.Frame(in_mi, bg=C_CARD)
        mi_hdr.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(mi_hdr, text="Market Insights", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")

        self._insights_frame = tk.Frame(in_mi, bg=C_CARD)
        self._insights_frame.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        # ── Right column: comparison + trending results ───────────────────
        right = tk.Frame(body, bg=C_BG)
        right.pack(side="right", fill="both", expand=True)

        # Comparison card
        out_c, in_c = _card(right)
        out_c.pack(fill="both", expand=True, pady=(0, 8))

        comp_hdr = tk.Frame(in_c, bg=C_CARD)
        comp_hdr.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(comp_hdr, text="Snapshot Comparison", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")

        comp_body = tk.Frame(in_c, bg=C_LOG_BG)
        comp_body.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        self._comp_text = tk.Text(comp_body, font=FT_MONO, bg=C_LOG_BG, fg=C_LOG_FG,
                                  relief="flat", wrap="word", state="disabled",
                                  padx=12, pady=8)
        csb = tk.Scrollbar(comp_body, orient="vertical", command=self._comp_text.yview)
        self._comp_text.configure(yscrollcommand=csb.set)
        csb.pack(side="right", fill="y")
        self._comp_text.pack(fill="both", expand=True)

        for tag, fg in [("green", C_GREEN), ("red", C_RED),
                        ("yellow", C_YELLOW), ("blue", C_BLUE), ("muted", C_MUTED)]:
            self._comp_text.tag_config(tag, foreground=fg)
        self._comp_text.tag_config("header", font=(_FM, _FS - 1, "bold"), foreground=C_BLUE)

        # Trending results card (below comparison)
        out_tr, in_tr = _card(right)
        out_tr.pack(fill="x", pady=(0, 0))

        tr_hdr = tk.Frame(in_tr, bg=C_CARD)
        tr_hdr.pack(fill="x", padx=12, pady=(10, 4))
        tk.Label(tr_hdr, text="Fiverr Trending (latest scan)", font=FT_H2,
                 bg=C_CARD, fg=C_TEXT).pack(side="left")
        _btn(tr_hdr, "▶  Run Scanner", lambda: self._run_tool(
             next(t for t in TOOLS if t["id"] == "trending_scan")),
             bg=B_PURPLE, fg=B_PURPLE_FG, padx=10, pady=3).pack(side="right")

        self._trending_frame = tk.Frame(in_tr, bg=C_CARD)
        self._trending_frame.pack(fill="x", padx=12, pady=(0, 10))

        return frame

    def _refresh_trends_page(self):
        self._refresh_snapshot_list()
        self._refresh_market_insights()
        self._refresh_trending_results()

    def _refresh_market_insights(self):
        for w in self._insights_frame.winfo_children():
            w.destroy()

        rows = read_excel_data()
        if not rows:
            tk.Label(self._insights_frame, text="Run Niche Scraper first.",
                     font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(anchor="w", pady=4)
            return

        scored = sorted([r for r in rows if r["score"] is not None],
                        key=lambda x: -x["score"])

        # Top opportunities
        tk.Label(self._insights_frame, text="Top Opportunities",
                 font=FT_H3, bg=C_CARD, fg=C_TEXT).pack(anchor="w", pady=(4, 2))
        for i, r in enumerate(scored[:6], 1):
            row_f = tk.Frame(self._insights_frame, bg=C_CARD)
            row_f.pack(fill="x", pady=1)
            col   = score_color(r["score"])
            tk.Label(row_f, text=f"{i}.", font=FT_TINY,
                     bg=C_CARD, fg=col, width=2).pack(side="left")
            tk.Label(row_f, text=r["sub"][:28], font=FT_SMALL,
                     bg=C_CARD, fg=C_TEXT, anchor="w").pack(side="left", fill="x", expand=True)
            tk.Label(row_f, text=f"{r['score']:.0f}", font=(FT_SMALL[0], FT_SMALL[1], "bold"),
                     bg=C_CARD, fg=col, width=4).pack(side="right")

        _sep(self._insights_frame, pady=6)

        # Entry level distribution from seller mix data
        high   = sum(1 for r in rows if str(r.get("entry","")).lower() == "high")
        medium = sum(1 for r in rows if str(r.get("entry","")).lower() == "medium")
        low    = sum(1 for r in rows if str(r.get("entry","")).lower() == "low")

        if high + medium + low > 0:
            tk.Label(self._insights_frame, text="Entry Difficulty",
                     font=FT_H3, bg=C_CARD, fg=C_TEXT).pack(anchor="w", pady=(4, 2))
            for label, val, col in [("High (easy entry)", high, C_SUCCESS),
                                    ("Medium",            medium, C_YELLOW),
                                    ("Low (tough)",       low,    C_RED)]:
                if val == 0:
                    continue
                ef = tk.Frame(self._insights_frame, bg=C_CARD)
                ef.pack(fill="x", pady=1)
                tk.Label(ef, text=label, font=FT_SMALL,
                         bg=C_CARD, fg=C_TEXT2, anchor="w").pack(side="left", fill="x", expand=True)
                tk.Label(ef, text=str(val), font=(FT_SMALL[0], FT_SMALL[1], "bold"),
                         bg=C_CARD, fg=col, width=4).pack(side="right")

    def _refresh_trending_results(self):
        for w in self._trending_frame.winfo_children():
            w.destroy()

        tfile = os.path.join(BASE_DIR, "trending_results.json")
        if not os.path.exists(tfile):
            tk.Label(self._trending_frame,
                     text="No trending data yet. Click ▶ Run Scanner to fetch live data.",
                     font=FT_SMALL, bg=C_CARD, fg=C_MUTED,
                     wraplength=400, justify="left").pack(anchor="w", pady=6)
            return

        try:
            with open(tfile) as f:
                data = json.load(f)
        except Exception:
            return

        ts = data.get("timestamp", "")[:16].replace("T", "  ")
        tk.Label(self._trending_frame, text=f"Last scanned: {ts}",
                 font=FT_TINY, bg=C_CARD, fg=C_MUTED).pack(anchor="w", pady=(0, 4))

        items  = data.get("items", [])
        # Show popular/hot items first, then categories
        priority = [d for d in items if d.get("type") in ("popular", "hot")]
        others   = [d for d in items if d.get("type") not in ("popular", "hot")]
        show     = (priority + others)[:14]

        cols_map = {"popular": C_GREEN, "hot": C_YELLOW,
                    "category": C_BLUE, "gig": C_TEXT2}
        for d in show:
            rf = tk.Frame(self._trending_frame, bg=C_CARD)
            rf.pack(fill="x", pady=1)
            col  = cols_map.get(d.get("type",""), C_MUTED)
            badge = {"popular":"◉ HOT", "hot":"★ PICK",
                     "category":"▸ CAT", "gig":"♦ GIG"}.get(d.get("type",""), "·")
            tk.Label(rf, text=badge, font=FT_TINY,
                     bg=C_CARD, fg=col, width=7).pack(side="left")
            tk.Label(rf, text=d["text"][:46], font=FT_SMALL,
                     bg=C_CARD, fg=C_TEXT, anchor="w").pack(side="left", fill="x", expand=True)
            tk.Label(rf, text=d.get("section","")[:10], font=FT_TINY,
                     bg=C_CARD, fg=C_MUTED).pack(side="right")

    def _save_snapshot(self):
        rows = read_excel_data()
        if rows is None:
            messagebox.showwarning("No Excel", "Excel file not found.", parent=self)
            return
        os.makedirs(SNAP_DIR, exist_ok=True)
        ts   = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = os.path.join(SNAP_DIR, f"snapshot_{ts}.json")
        snap = {
            "timestamp": datetime.now().isoformat(),
            "total":     len(rows),
            "positive":  sum(1 for r in rows if r["verdict"].lower() == "positive"),
            "niches":    [{"sub": r["sub"], "niche": r["niche"], "section": r["section"],
                           "verdict": r["verdict"], "gig": r["gig"],
                           "reviews": r["reviews"], "queue": r["queue"],
                           "score": r["score"]}
                          for r in rows],
        }
        with open(path, "w") as f:
            json.dump(snap, f, indent=2)
        messagebox.showinfo("Snapshot Saved",
                            f"snapshot_{ts}.json\n\n"
                            f"Niches: {snap['total']}  |  Positive: {snap['positive']}",
                            parent=self)
        self._refresh_snapshot_list()

    def _refresh_snapshot_list(self):
        self._snap_listbox.delete(0, "end")
        os.makedirs(SNAP_DIR, exist_ok=True)
        files = sorted([f for f in os.listdir(SNAP_DIR) if f.endswith(".json")],
                       reverse=True)
        for fn in files:
            path = os.path.join(SNAP_DIR, fn)
            try:
                with open(path) as f:
                    d = json.load(f)
                ts = fn.replace("snapshot_","").replace(".json","").replace("_"," ")
                label = f"  {ts}   ({d['total']} niches, {d['positive']}✓)"
            except Exception:
                label = f"  {fn}"
            self._snap_listbox.insert("end", label)
        count = len(files)
        self._snap_count_lbl.config(
            text=f"{count} snapshot{'s' if count != 1 else ''} saved"
            if count else "No snapshots yet — click 'Save Now'")

    def _get_snapshot_files(self):
        os.makedirs(SNAP_DIR, exist_ok=True)
        return sorted([os.path.join(SNAP_DIR, f)
                       for f in os.listdir(SNAP_DIR) if f.endswith(".json")], reverse=True)

    def _compare_latest_two(self):
        files = self._get_snapshot_files()
        if len(files) < 2:
            messagebox.showwarning("Need 2 Snapshots",
                                   "Save at least two snapshots to compare.", parent=self)
            return
        try:
            with open(files[0]) as f: snap_new = json.load(f)
            with open(files[1]) as f: snap_old = json.load(f)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)
            return
        self._show_comparison(snap_old, snap_new,
                              os.path.basename(files[1]), os.path.basename(files[0]))

    def _show_comparison(self, old, new, old_name, new_name):
        t = self._comp_text
        t.config(state="normal")
        t.delete("1.0", "end")

        def w(text, tag=None):
            t.insert("end", text, tag) if tag else t.insert("end", text)

        w("SNAPSHOT COMPARISON\n", "header")
        w(f"  Old: {old_name}\n", "muted")
        w(f"  New: {new_name}\n\n", "muted")

        old_map = {r["sub"]: r for r in old.get("niches", [])}
        new_map = {r["sub"]: r for r in new.get("niches", [])}

        changed = improved = worsened = []
        changed, improved, worsened = [], [], []
        for sub, nr in new_map.items():
            or_ = old_map.get(sub)
            if not or_: continue
            ov = (or_.get("verdict") or "").lower()
            nv = (nr.get("verdict") or "").lower()
            if ov != nv and (ov or nv):
                changed.append((sub, ov, nv))
            og, ng = or_.get("gig"), nr.get("gig")
            if og and ng:
                pct = (ng - og) / og * 100
                if pct >= 15:   worsened.append((sub, og, ng, pct))
                elif pct <= -15: improved.append((sub, og, ng, abs(pct)))

        w("SUMMARY\n", "header")
        w(f"  Total niches:    {old['total']}  →  {new['total']}\n")
        w(f"  Positive niches: {old['positive']}  →  {new['positive']}\n")
        diff_pos = new["positive"] - old["positive"]
        if diff_pos > 0:   w(f"  Net new positives: +{diff_pos}\n", "green")
        elif diff_pos < 0: w(f"  Net lost positives: {diff_pos}\n", "red")

        if changed:
            w(f"\nSTATUS CHANGES  ({len(changed)})\n", "header")
            for sub, ov, nv in changed[:30]:
                arrow = "→"
                color = "green" if nv == "positive" else "red"
                w(f"  {sub[:44]:44s}  {ov or 'pending':10s} {arrow} ", "muted")
                w(f"{nv or 'pending'}\n", color)

        if improved:
            improved.sort(key=lambda x: x[3], reverse=True)
            w(f"\nCOMPETITION DROPPED  (gig count fell ≥15%)  — opportunities improved\n", "header")
            for sub, og, ng, pct in improved[:15]:
                w(f"  ↓  {sub[:44]:44s}  {og:,} → {ng:,}  (-{pct:.0f}%)\n", "green")

        if worsened:
            worsened.sort(key=lambda x: x[3], reverse=True)
            w(f"\nCOMPETITION GREW  (gig count rose ≥15%)  — harder to enter\n", "header")
            for sub, og, ng, pct in worsened[:15]:
                w(f"  ↑  {sub[:44]:44s}  {og:,} → {ng:,}  (+{pct:.0f}%)\n", "red")

        if not changed and not improved and not worsened:
            w("\n  No significant changes detected between the two snapshots.\n", "muted")
            w("  (Try comparing snapshots that are further apart in time.)\n", "muted")

        t.config(state="disabled")

    def _delete_snapshot(self):
        sel = self._snap_listbox.curselection()
        if not sel:
            messagebox.showwarning("Select One", "Select a snapshot to delete.", parent=self)
            return
        files = self._get_snapshot_files()
        idx   = sel[0]
        if idx >= len(files): return
        if messagebox.askyesno("Delete", f"Delete:\n{os.path.basename(files[idx])}?", parent=self):
            os.remove(files[idx])
            self._refresh_snapshot_list()

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Niche Manager
    # ══════════════════════════════════════════════════════════════════════════
    def _build_niche_manager_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        # Title bar
        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 8))
        tk.Label(th, text="Niche Manager", font=FT_H1,
                 bg=C_BG, fg=C_TEXT).pack(side="left")
        tk.Label(th, text="  Dynamic sub-niche list — add, import, discover, delete",
                 font=FT_SMALL, bg=C_BG, fg=C_MUTED).pack(side="left")
        _btn(th, "↻  Refresh", self._load_niche_data,
             bg=B_GHOST, fg=B_GHOST_FG, padx=10, pady=5).pack(side="right", padx=(4, 0))
        _btn(th, "⬆  Import CSV", self._import_niches_csv,
             bg=B_BLUE, fg=B_BLUE_FG, padx=10, pady=5).pack(side="right", padx=4)

        # Filter bar
        fbar_out, fbar = _card(frame)
        fbar_out.pack(fill="x", padx=20, pady=(0, 6))
        fb = tk.Frame(fbar, bg=C_CARD)
        fb.pack(fill="x", padx=12, pady=8)

        tk.Label(fb, text="Search:", font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._nm_search_var = tk.StringVar()
        self._nm_search_var.trace_add("write", lambda *_: self._apply_niche_filter_nm())
        tk.Entry(fb, textvariable=self._nm_search_var, font=FT_BODY,
                 bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                 relief="flat", width=22, highlightthickness=0
                 ).pack(side="left", padx=(4, 14))

        tk.Label(fb, text="Section:", font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._nm_section_var = tk.StringVar(value="All")
        self._nm_section_var.trace_add("write", lambda *_: self._apply_niche_filter_nm())
        self._nm_section_cb = ttk.Combobox(fb, textvariable=self._nm_section_var,
                                            values=["All"], state="readonly", width=18)
        self._nm_section_cb.pack(side="left", padx=(2, 14))

        tk.Label(fb, text="Status:", font=FT_SMALL, bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._nm_status_var = tk.StringVar(value="All")
        self._nm_status_var.trace_add("write", lambda *_: self._apply_niche_filter_nm())
        ttk.Combobox(fb, textvariable=self._nm_status_var,
                     values=["All", "Positive", "Negative", "Pending"],
                     state="readonly", width=10).pack(side="left", padx=(2, 0))

        self._nm_count_lbl = tk.Label(fb, text="", font=FT_TINY, bg=C_CARD, fg=C_MUTED)
        self._nm_count_lbl.pack(side="right")

        # Treeview
        tv_frame = tk.Frame(frame, bg=C_BG)
        tv_frame.pack(fill="both", expand=True, padx=20, pady=(0, 4))

        nm_cols = ("section","niche","sub","keyword","status","gigs","score")
        nm_hdgs = ("Section","Niche","Sub-Niche","Keyword","Status","Gig Count","Score")
        nm_wids = (110, 110, 185, 150, 80, 85, 60)

        self._nm_tv = ttk.Treeview(tv_frame, columns=nm_cols, show="headings",
                                    style="Hub.Treeview", selectmode="extended")
        for col, h, w in zip(nm_cols, nm_hdgs, nm_wids):
            self._nm_tv.heading(col, text=h, anchor="w",
                                command=lambda c=col: self._nm_sort_by(c))
            self._nm_tv.column(col, width=w, anchor="w", minwidth=50)

        self._nm_tv.tag_configure("positive", foreground=C_SUCCESS)
        self._nm_tv.tag_configure("negative", foreground=C_RED)
        self._nm_tv.tag_configure("pending",  foreground=C_YELLOW)

        nm_vsb = ttk.Scrollbar(tv_frame, orient="vertical",   command=self._nm_tv.yview)
        nm_hsb = ttk.Scrollbar(tv_frame, orient="horizontal", command=self._nm_tv.xview)
        self._nm_tv.configure(yscrollcommand=nm_vsb.set, xscrollcommand=nm_hsb.set)
        self._nm_tv.grid(row=0, column=0, sticky="nsew")
        nm_vsb.grid(row=0, column=1, sticky="ns")
        nm_hsb.grid(row=1, column=0, sticky="ew")
        tv_frame.grid_rowconfigure(0, weight=1)
        tv_frame.grid_columnconfigure(0, weight=1)

        # Action row
        act_out, act = _card(frame)
        act_out.pack(fill="x", padx=20, pady=(0, 6))
        act_row = tk.Frame(act, bg=C_CARD)
        act_row.pack(fill="x", padx=12, pady=8)
        _btn(act_row, "🗑  Delete Selected", self._delete_selected_nm,
             bg=B_RED, fg=B_RED_FG, padx=12, pady=5).pack(side="left", padx=(0, 8))
        _btn(act_row, "✎  Edit Keyword", self._edit_keyword_nm,
             bg=B_GHOST, fg=B_GHOST_FG, padx=12, pady=5).pack(side="left")
        self._nm_info_lbl = tk.Label(act_row, text="", font=FT_SMALL,
                                      bg=C_CARD, fg=C_MUTED)
        self._nm_info_lbl.pack(side="right")

        # Add new niche panel
        add_out, add = _card(frame)
        add_out.pack(fill="x", padx=20, pady=(0, 6))
        add_hdr = tk.Frame(add, bg=_adj(C_CARD, 1.25))
        add_hdr.pack(fill="x")
        tk.Frame(add_hdr, bg=B_GREEN, width=3).pack(side="left", fill="y")
        tk.Label(add_hdr, text="  + Add New Sub-Niche", font=FT_H3,
                 bg=_adj(C_CARD, 1.25), fg=C_TEXT, anchor="w",
                 pady=7).pack(side="left", fill="x")
        add_body = tk.Frame(add, bg=C_CARD)
        add_body.pack(fill="x", padx=12, pady=10)
        self._nm_add_vars = {}
        for label, key, width in [
            ("Section",   "section",  13),
            ("Niche",     "niche",    16),
            ("Sub-Niche *", "sub",    22),
            ("Keyword",   "keyword",  22),
        ]:
            col = tk.Frame(add_body, bg=C_CARD)
            col.pack(side="left", padx=(0, 10))
            tk.Label(col, text=label, font=FT_TINY, bg=C_CARD, fg=C_MUTED).pack(anchor="w")
            var = tk.StringVar()
            self._nm_add_vars[key] = var
            tk.Entry(col, textvariable=var, font=FT_BODY,
                     bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                     relief="flat", width=width, highlightthickness=0).pack()
        _btn(add_body, "+ Add", self._do_add_niche,
             bg=B_GREEN, fg=B_GREEN_FG, padx=16, pady=7).pack(side="left", padx=(10, 0))

        # Discover panel
        disc_out, disc = _card(frame)
        disc_out.pack(fill="x", padx=20, pady=(0, 12))
        disc_hdr = tk.Frame(disc, bg=_adj(C_CARD, 1.25))
        disc_hdr.pack(fill="x")
        tk.Frame(disc_hdr, bg=B_PURPLE, width=3).pack(side="left", fill="y")
        tk.Label(disc_hdr, text="  🔭 Discover New Niches from Fiverr",
                 font=FT_H3, bg=_adj(C_CARD, 1.25), fg=C_TEXT,
                 anchor="w", pady=7).pack(side="left", fill="x")
        disc_body = tk.Frame(disc, bg=C_CARD)
        disc_body.pack(fill="x", padx=12, pady=10)
        tk.Label(disc_body, text="Seed keyword:", font=FT_SMALL,
                 bg=C_CARD, fg=C_MUTED).pack(side="left")
        self._nm_disc_var = tk.StringVar()
        tk.Entry(disc_body, textvariable=self._nm_disc_var, font=FT_BODY,
                 bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                 relief="flat", width=28, highlightthickness=0
                 ).pack(side="left", padx=(6, 10))
        _btn(disc_body, "▶  Discover", self._run_niche_discovery,
             bg=B_PURPLE, fg=B_PURPLE_FG, padx=14, pady=5).pack(side="left")
        _btn(disc_body, "⬆  Import Results", self._import_discovery_results,
             bg=B_GHOST, fg=B_GHOST_FG, padx=12, pady=5).pack(side="left", padx=(8, 0))
        tk.Label(disc_body, text="  (results stream to Scrapers log)",
                 font=FT_TINY, bg=C_CARD, fg=C_MUTED).pack(side="left")

        self._nm_all_data  = []
        self._nm_sort_col  = None
        self._nm_sort_rev  = False
        return frame

    # ── Niche Manager logic ────────────────────────────────────────────────────
    def _load_niche_data(self):
        if not os.path.exists(EXCEL_PATH):
            messagebox.showwarning("No Excel",
                "Excel not found. Run 'Create / Regenerate Excel' first.", parent=self)
            return
        rows = read_excel_data()
        if rows is None:
            return
        self._nm_all_data = rows
        sections = sorted(set(r["section"] for r in rows if r["section"]))
        self._nm_section_cb["values"] = ["All"] + sections
        self._apply_niche_filter_nm()

    def _apply_niche_filter_nm(self):
        if not hasattr(self, "_nm_search_var"):
            return
        s_val   = self._nm_search_var.get().lower()
        section = self._nm_section_var.get()
        status  = self._nm_status_var.get()

        data = list(self._nm_all_data)
        if section != "All":
            data = [r for r in data if r["section"] == section]
        if status == "Positive":  data = [r for r in data if r["verdict"].lower() == "positive"]
        elif status == "Negative": data = [r for r in data if r["verdict"].lower() == "negative"]
        elif status == "Pending":  data = [r for r in data
                                           if r["verdict"].lower() not in ("positive","negative")]
        if s_val:
            data = [r for r in data if (s_val in r["sub"].lower()
                    or s_val in r["niche"].lower()
                    or s_val in r["section"].lower()
                    or s_val in (r.get("keyword","") or "").lower())]

        # Sort if a column is chosen
        if self._nm_sort_col:
            key_map = {
                "section": lambda r: r["section"].lower(),
                "niche":   lambda r: r["niche"].lower(),
                "sub":     lambda r: r["sub"].lower(),
                "keyword": lambda r: (r.get("keyword","") or "").lower(),
                "status":  lambda r: r["verdict"].lower(),
                "gigs":    lambda r: r["gig"] or 0,
                "score":   lambda r: r["score"] or 0,
            }
            fn = key_map.get(self._nm_sort_col, lambda r: r["sub"].lower())
            data.sort(key=fn, reverse=self._nm_sort_rev)

        for item in self._nm_tv.get_children():
            self._nm_tv.delete(item)

        for r in data:
            v   = r["verdict"].lower()
            sc  = r["score"]
            tag = "positive" if v == "positive" else "negative" if v == "negative" else "pending"
            self._nm_tv.insert("", "end", iid=str(r["row"]), values=(
                r["section"], r["niche"], r["sub"],
                r.get("keyword","") or r["sub"].lower(),
                r["verdict"] or "Pending",
                f"{r['gig']:,}" if r["gig"] is not None else "—",
                f"{sc:.0f}" if sc is not None else "—",
            ), tags=(tag,))

        total   = len(self._nm_all_data)
        showing = len(data)
        pos     = sum(1 for r in self._nm_all_data if r["verdict"].lower() == "positive")
        pending = sum(1 for r in self._nm_all_data if r["verdict"].lower()
                      not in ("positive","negative"))
        self._nm_count_lbl.config(text=f"Showing {showing} / {total}")
        self._nm_info_lbl.config(
            text=f"Total: {total}  |  Positive: {pos}  |  Pending: {pending}")

    def _nm_sort_by(self, col):
        if self._nm_sort_col == col:
            self._nm_sort_rev = not self._nm_sort_rev
        else:
            self._nm_sort_col = col
            self._nm_sort_rev = False
        self._apply_niche_filter_nm()

    def _do_add_niche(self):
        section = self._nm_add_vars["section"].get().strip()
        niche   = self._nm_add_vars["niche"].get().strip()
        sub     = self._nm_add_vars["sub"].get().strip()
        keyword = self._nm_add_vars["keyword"].get().strip()

        if not sub:
            messagebox.showwarning("Required", "Sub-Niche field is required.", parent=self)
            return
        if not section: section = "Custom"
        if not niche:   niche   = section
        if not keyword: keyword = sub.lower()

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side
            wb  = openpyxl.load_workbook(EXCEL_PATH)
            ws  = wb.active
            # Duplicate check
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 3).value and \
                   str(ws.cell(row, 3).value).strip().lower() == sub.lower():
                    messagebox.showwarning("Duplicate",
                        f"Sub-niche '{sub}' already exists (row {row}).", parent=self)
                    return
            thin = Side(style="thin", color="BDBDBD")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
            nr   = ws.max_row + 1
            for col, val in [(1, section), (2, niche), (3, sub), (5, keyword)]:
                c = ws.cell(row=nr, column=col, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = brd
            wb.save(EXCEL_PATH)
            for var in self._nm_add_vars.values():
                var.set("")
            messagebox.showinfo("Added", f"'{sub}' added to row {nr}.", parent=self)
            self._load_niche_data()
            self._set_status(f"Added: {sub}")
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)

    def _edit_keyword_nm(self):
        sel = self._nm_tv.selection()
        if not sel:
            messagebox.showwarning("Nothing Selected", "Select a row to edit.", parent=self)
            return
        row_idx  = int(sel[0])
        cur_vals = self._nm_tv.item(sel[0], "values")
        cur_kw   = cur_vals[3] if len(cur_vals) > 3 else ""

        dlg = tk.Toplevel(self)
        dlg.title("Edit Keyword")
        dlg.configure(bg=C_BG)
        dlg.geometry("420x160")
        dlg.grab_set()
        tk.Label(dlg, text=f"Sub-Niche: {cur_vals[2]}", font=FT_H3,
                 bg=C_BG, fg=C_TEXT).pack(padx=20, pady=(16, 4), anchor="w")
        tk.Label(dlg, text="Keyword (used for Fiverr search):", font=FT_SMALL,
                 bg=C_BG, fg=C_MUTED).pack(padx=20, anchor="w")
        var = tk.StringVar(value=cur_kw)
        tk.Entry(dlg, textvariable=var, font=FT_BODY,
                 bg=C_CARD, fg=C_TEXT, insertbackground=C_GREEN,
                 relief="flat", width=38, highlightthickness=0).pack(padx=20, pady=6)
        def save():
            kw = var.get().strip()
            if not kw:
                return
            try:
                import openpyxl
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active
                ws.cell(row=row_idx, column=5, value=kw)
                wb.save(EXCEL_PATH)
                dlg.destroy()
                self._load_niche_data()
                self._set_status("Keyword updated")
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)
        bf = tk.Frame(dlg, bg=C_BG)
        bf.pack(padx=20, fill="x")
        _btn(bf, "Save", save, bg=B_GREEN, fg=B_GREEN_FG, padx=14, pady=6).pack(side="left")
        _btn(bf, "Cancel", dlg.destroy, bg=B_GHOST, fg=B_GHOST_FG, padx=10, pady=6
             ).pack(side="left", padx=8)

    def _delete_selected_nm(self):
        sel = self._nm_tv.selection()
        if not sel:
            messagebox.showwarning("Nothing Selected",
                "Select one or more rows to delete.", parent=self)
            return
        rows_to_del = [int(iid) for iid in sel]
        names       = [self._nm_tv.item(iid, "values")[2] for iid in sel]
        preview     = "\n".join(f"  • {n}" for n in names[:6])
        if len(names) > 6:
            preview += f"\n  … and {len(names)-6} more"
        if not messagebox.askyesno("Delete",
                f"Permanently delete {len(rows_to_del)} sub-niche(s)?\n\n{preview}",
                parent=self):
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            for r in sorted(rows_to_del, reverse=True):
                ws.delete_rows(r)
            wb.save(EXCEL_PATH)
            messagebox.showinfo("Deleted",
                f"{len(rows_to_del)} sub-niche(s) deleted.", parent=self)
            self._load_niche_data()
            self._set_status(f"Deleted {len(rows_to_del)} rows")
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)

    def _import_niches_csv(self):
        path = filedialog.askopenfilename(
            title="Import Sub-Niches from CSV",
            filetypes=[("CSV files","*.csv"),("Text files","*.txt"),("All","*.*")],
            parent=self)
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side
            rows_in = []
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                next(reader, None)   # skip header
                for line in reader:
                    if len(line) >= 1 and any(line):
                        rows_in.append(line)
            if not rows_in:
                messagebox.showwarning("Empty", "No data rows found.", parent=self)
                return

            wb   = openpyxl.load_workbook(EXCEL_PATH)
            ws   = wb.active
            thin = Side(style="thin", color="BDBDBD")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

            existing = set()
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, 3).value
                if v: existing.add(str(v).strip().lower())

            added = skipped = 0
            for line in rows_in:
                section = (line[0].strip() if len(line) > 0 else "") or "Custom"
                niche   = (line[1].strip() if len(line) > 1 else "") or section
                sub     = (line[2].strip() if len(line) > 2 else "") or (line[0].strip() if line else "")
                keyword = (line[3].strip() if len(line) > 3 else "") or sub.lower()
                if not sub or sub.lower() in existing:
                    skipped += 1
                    continue
                nr = ws.max_row + 1
                for col, val in [(1, section), (2, niche), (3, sub), (5, keyword)]:
                    c = ws.cell(row=nr, column=col, value=val)
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    c.border    = brd
                existing.add(sub.lower())
                added += 1

            wb.save(EXCEL_PATH)
            messagebox.showinfo("Import Complete",
                f"Added: {added} new sub-niches\nSkipped (duplicates): {skipped}", parent=self)
            self._load_niche_data()
            self._set_status(f"Imported {added} niches")
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)

    def _run_niche_discovery(self):
        keyword = self._nm_disc_var.get().strip()
        if not keyword:
            messagebox.showwarning("No Keyword",
                "Enter a seed keyword to discover niches.", parent=self)
            return
        script = os.path.join(BASE_DIR, "fiverr_niche_discovery.py")
        if not os.path.exists(script):
            messagebox.showerror("Not Found",
                "fiverr_niche_discovery.py not found.", parent=self)
            return
        if self._proc and self._proc.poll() is None:
            messagebox.showwarning("Already Running",
                "A tool is currently running. Stop it first.", parent=self)
            return
        self._show_page("scrapers")
        self._active_tool = {"title": f"Niche Discovery: '{keyword}'",
                             "icon": "🔭", "id": "niche_discovery"}
        self._clear_log()
        self._log_write(f"{'─'*60}\n", "muted")
        self._log_write(f"  🔭  Niche Discovery — \"{keyword}\"\n", "purple")
        self._log_write(f"{'─'*60}\n\n", "muted")
        self._set_status(f"Discovering: {keyword}", running=True)
        self._log_status_dot.config(fg=C_GREEN)
        self._btn_stop.config(state="normal")
        self._disable_run_buttons()

        def _run():
            try:
                self._proc = subprocess.Popen(
                    [PYTHON, script, keyword], cwd=BASE_DIR,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, bufsize=1)
                for line in self._proc.stdout:
                    self._log_queue.put(("line", line))
                self._proc.wait()
                rc = self._proc.returncode
                self._log_queue.put(("done", None) if rc == 0 else ("error", rc))
            except Exception as e:
                self._log_queue.put(("exception", str(e)))
        threading.Thread(target=_run, daemon=True).start()

    def _import_discovery_results(self):
        rf = os.path.join(BASE_DIR, "discovery_results.json")
        if not os.path.exists(rf):
            messagebox.showwarning("No Results",
                "No discovery_results.json found.\nRun Discover first.", parent=self)
            return
        try:
            with open(rf) as f:
                data = json.load(f)
            discoveries = data.get("discoveries", [])
            if not discoveries:
                messagebox.showinfo("Empty", "No sub-niches in discovery results.", parent=self)
                return
            self._show_discovery_import_dialog(data.get("seed_keyword",""), discoveries)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)

    def _show_discovery_import_dialog(self, seed, discoveries):
        dlg = tk.Toplevel(self)
        dlg.title(f"Import Discovered Niches — \"{seed}\"")
        dlg.configure(bg=C_BG)
        dlg.geometry("700x540")
        dlg.grab_set()

        tk.Label(dlg, text=f"Select sub-niches to add  ({len(discoveries)} discovered)",
                 font=FT_H2, bg=C_BG, fg=C_TEXT).pack(padx=20, pady=(16,8), anchor="w")

        # Scrollable checklist
        list_outer = tk.Frame(dlg, bg=C_CARD, relief="flat")
        list_outer.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        canvas = tk.Canvas(list_outer, bg=C_CARD, highlightthickness=0)
        vsb    = ttk.Scrollbar(list_outer, orient="vertical", command=canvas.yview)
        inner  = tk.Frame(canvas, bg=C_CARD)
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        check_vars = []
        for d in discoveries:
            var = tk.BooleanVar(value=True)
            rf  = tk.Frame(inner, bg=C_CARD)
            rf.pack(fill="x", padx=10, pady=2)
            tk.Checkbutton(rf, variable=var, bg=C_CARD, fg=C_TEXT,
                           activebackground=C_CARD, selectcolor=C_BG,
                           highlightthickness=0).pack(side="left")
            tk.Label(rf, text=d["keyword"], font=FT_SMALL,
                     bg=C_CARD, fg=C_TEXT, width=36, anchor="w").pack(side="left")
            tk.Label(rf, text=f"[{d.get('source','')}]", font=FT_TINY,
                     bg=C_CARD, fg=C_MUTED).pack(side="left", padx=(8,0))
            check_vars.append((var, d))

        # Section / Niche override
        form = tk.Frame(dlg, bg=C_BG)
        form.pack(fill="x", padx=20, pady=(0, 8))
        tk.Label(form, text="Add checked as — Section:", font=FT_SMALL,
                 bg=C_BG, fg=C_MUTED).pack(side="left")
        sec_var = tk.StringVar(value="Custom")
        tk.Entry(form, textvariable=sec_var, font=FT_BODY,
                 bg=C_CARD, fg=C_TEXT, insertbackground=C_GREEN,
                 relief="flat", width=14, highlightthickness=0).pack(side="left", padx=(4,10))
        tk.Label(form, text="Niche:", font=FT_SMALL, bg=C_BG, fg=C_MUTED).pack(side="left")
        nic_var = tk.StringVar(value=seed.title() if seed else "Discovered")
        tk.Entry(form, textvariable=nic_var, font=FT_BODY,
                 bg=C_CARD, fg=C_TEXT, insertbackground=C_GREEN,
                 relief="flat", width=20, highlightthickness=0).pack(side="left", padx=4)

        # Buttons
        bf = tk.Frame(dlg, bg=C_BG)
        bf.pack(fill="x", padx=20, pady=(0, 16))

        def do_import():
            selected = [(v, d) for v, d in check_vars if v.get()]
            if not selected:
                messagebox.showwarning("None Selected",
                    "Check at least one niche.", parent=dlg)
                return
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Border, Side
                wb   = openpyxl.load_workbook(EXCEL_PATH)
                ws   = wb.active
                thin = Side(style="thin", color="BDBDBD")
                brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
                existing = set()
                for r in range(2, ws.max_row + 1):
                    v2 = ws.cell(r, 3).value
                    if v2: existing.add(str(v2).strip().lower())
                added = 0
                for _, d in selected:
                    sub = d["keyword"].strip()
                    if sub.lower() in existing:
                        continue
                    nr = ws.max_row + 1
                    for col, val in [(1, sec_var.get()), (2, nic_var.get()),
                                     (3, sub), (5, sub.lower())]:
                        c = ws.cell(row=nr, column=col, value=val)
                        c.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True)
                        c.border = brd
                    existing.add(sub.lower())
                    added += 1
                wb.save(EXCEL_PATH)
                dlg.destroy()
                messagebox.showinfo("Imported",
                    f"Added {added} new sub-niches.", parent=self)
                self._load_niche_data()
                self._set_status(f"Imported {added} discovered niches")
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        _btn(bf, "✓  Add Selected", do_import,
             bg=B_GREEN, fg=B_GREEN_FG, padx=16, pady=7, font=FT_H3).pack(side="left")
        _btn(bf, "Cancel", dlg.destroy,
             bg=B_GHOST, fg=B_GHOST_FG, padx=12, pady=7).pack(side="left", padx=10)

    # ══════════════════════════════════════════════════════════════════════════
    #  PAGE: Settings
    # ══════════════════════════════════════════════════════════════════════════
    def _build_settings_page(self):
        frame = tk.Frame(self._content, bg=C_BG)

        th = tk.Frame(frame, bg=C_BG)
        th.pack(fill="x", padx=20, pady=(18, 10))
        tk.Label(th, text="Settings", font=FT_H1,
                 bg=C_BG, fg=C_TEXT).pack(side="left")

        body = tk.Frame(frame, bg=C_BG)
        body.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        left  = tk.Frame(body, bg=C_BG)
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))
        right = tk.Frame(body, bg=C_BG)
        right.pack(side="right", fill="both", expand=True)

        self._setting_vars = {}

        def section(parent, title, icon=""):
            out, inn = _card(parent)
            out.pack(fill="x", pady=(0, 10))
            # Coloured header bar
            hdr = tk.Frame(inn, bg=_adj(C_CARD, 1.3))
            hdr.pack(fill="x")
            tk.Label(hdr, text=f"  {icon}  {title}", font=FT_H3,
                     bg=_adj(C_CARD, 1.3), fg=C_TEXT, anchor="w",
                     pady=10).pack(fill="x")
            _sep(inn, bg=C_LINE, pady=0)
            body_ = tk.Frame(inn, bg=C_CARD)
            body_.pack(fill="x", padx=14, pady=10)
            return body_

        def field(parent, label, key, tooltip="", wide=False):
            row = tk.Frame(parent, bg=C_CARD)
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label, font=FT_SMALL, bg=C_CARD, fg=C_TEXT2,
                     width=22 if not wide else 18, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(self._settings.get(key, "")))
            self._setting_vars[key] = var
            e = tk.Entry(row, textvariable=var, font=FT_BODY,
                         bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                         relief="flat", width=18 if not wide else 28)
            e.pack(side="left", padx=(4, 8))
            if tooltip:
                tk.Label(row, text=tooltip, font=FT_TINY, bg=C_CARD, fg=C_MUTED
                         ).pack(side="left")
            return var

        def path_field(parent, label, key):
            row = tk.Frame(parent, bg=C_CARD)
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label, font=FT_SMALL, bg=C_CARD, fg=C_TEXT2,
                     width=16, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(self._settings.get(key, "")))
            self._setting_vars[key] = var
            tk.Entry(row, textvariable=var, font=FT_TINY,
                     bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                     relief="flat").pack(side="left", fill="x", expand=True, padx=(4, 4))
            _btn(row, "Browse", lambda k=key, v=var: self._browse_file(k, v),
                 bg=B_GHOST, fg=B_GHOST_FG, padx=8, pady=2).pack(side="right")

        # ── AI Configuration ────────────────────────────────────────────────
        ai = section(left, "AI Configuration", "🤖")
        ai_info = tk.Frame(ai, bg=C_CARD)
        ai_info.pack(fill="x", pady=(0, 8))

        # Provider row
        row_prov = tk.Frame(ai_info, bg=C_CARD)
        row_prov.pack(fill="x", pady=2)
        tk.Label(row_prov, text="Provider", font=FT_SMALL, bg=C_CARD, fg=C_TEXT2,
                 width=16, anchor="w").pack(side="left")
        self._ai_provider_lbl = tk.Label(row_prov,
            text=get_provider_label() if _API_MANAGER_OK else "api_manager missing",
            font=FT_SMALL, bg=C_CARD, fg=C_TEXT, anchor="w")
        self._ai_provider_lbl.pack(side="left")

        # Key row
        row_key = tk.Frame(ai_info, bg=C_CARD)
        row_key.pack(fill="x", pady=2)
        tk.Label(row_key, text="API Key", font=FT_SMALL, bg=C_CARD, fg=C_TEXT2,
                 width=16, anchor="w").pack(side="left")
        self._ai_key_lbl = tk.Label(row_key,
            text=get_masked_key() if _API_MANAGER_OK else "—",
            font=FT_MONO_S, bg=C_CARD, fg=C_PURPLE, anchor="w")
        self._ai_key_lbl.pack(side="left")

        # Status indicator
        if _API_MANAGER_OK and has_api_key():
            status_text, status_col = "●  API key configured", C_SUCCESS
        else:
            status_text, status_col = "●  No API key — click Update to add one", C_YELLOW
        ai_status = tk.Label(ai_info, text=status_text, font=FT_TINY,
                             bg=C_CARD, fg=status_col, anchor="w")
        ai_status.pack(anchor="w", pady=(4, 0))

        _btn(ai, "🔑  Update API Key", lambda: self._show_api_key_dialog(initial=False),
             bg=B_BLUE, fg=B_BLUE_FG, padx=14, pady=7).pack(anchor="w", pady=(4, 0))

        # Scraping config
        sc = section(left, "Scraping Config", "⚙")
        field(sc, "Gig Count Threshold",    "threshold",  "(≤ this = Positive)")
        # Live threshold preview
        thr_row = tk.Frame(sc, bg=C_CARD)
        thr_row.pack(fill="x", pady=(0, 6))
        thr_note = tk.Label(thr_row,
            text="  Tip: use the Gigs min–max filter in Analysis to browse any range.",
            font=FT_TINY, bg=C_CARD, fg=C_MUTED, justify="left", wraplength=310)
        thr_note.pack(anchor="w", padx=4)
        field(sc, "Min Delay  (seconds)",   "min_delay",  "(between searches)")
        field(sc, "Max Delay  (seconds)",   "max_delay",  "(random range)")
        field(sc, "Slow-Down After N Rows", "slow_down",  "(rate-limit break)")

        # Paths
        ph = section(left, "File Paths", "📁")
        path_field(ph, "Excel File",  "excel_path")
        path_field(ph, "Images Dir",  "images_dir")

        # Save / reset
        btns = tk.Frame(left, bg=C_BG)
        btns.pack(fill="x", pady=4)
        _btn(btns, "💾  Save Settings", self._save_settings_from_ui,
             bg=B_GREEN, fg=B_GREEN_FG, padx=16, pady=8, font=FT_H3).pack(side="left", padx=(0, 10))
        _btn(btns, "Reset to Defaults", self._reset_settings,
             bg=B_GHOST, fg=B_GHOST_FG, padx=12, pady=8).pack(side="left")

        # ── Right: info + data management ─────────────────────────────────
        inf = section(right, "How Settings Are Used", "ℹ")
        notes = [
            ("Threshold",  C_BLUE,   "Niches with ≤ this many gigs → Positive.  Used by all new scrapers and the Opportunity Score calculation."),
            ("Delays",     C_YELLOW, "Random wait between Fiverr searches to avoid rate-limiting.  Note: the original fiverr_scraper.py has its own hardcoded delays — edit that file directly if needed."),
            ("Slow-Down",  C_GREEN,  "After N consecutive searches, the scraper pauses for 20–35 seconds to avoid triggering Fiverr's bot detection."),
            ("Config File",C_MUTED,  f"All settings are saved to hub_config.json in the project folder.  New scripts automatically read this file on startup."),
        ]
        for key, color, note in notes:
            nrow = tk.Frame(inf, bg=C_CARD)
            nrow.pack(fill="x", pady=(0, 8))
            hr = tk.Frame(nrow, bg=C_CARD)
            hr.pack(fill="x")
            tk.Frame(hr, bg=color, width=3, height=14).pack(side="left", padx=(0, 6))
            tk.Label(hr, text=key, font=FT_H3, bg=C_CARD, fg=C_TEXT).pack(side="left", anchor="w")
            tk.Label(nrow, text=note, font=FT_TINY, bg=C_CARD, fg=C_MUTED,
                     wraplength=310, justify="left", anchor="w").pack(anchor="w", padx=12)

        dm = section(right, "Data Management", "🗂")
        for text, bg, fg, cmd in [
            ("📋  Backup Excel Now",        B_BLUE,   B_BLUE_FG,   self._backup_excel),
            ("📂  Open Snapshots Folder",   B_GHOST,  B_GHOST_FG,  self._open_snapshots_folder),
            ("📊  Open Keyword Variations", B_PURPLE, B_PURPLE_FG, self._open_csv_variations),
        ]:
            _btn(dm, text, cmd, bg=bg, fg=fg, padx=14, pady=7).pack(anchor="w", pady=3)

        return frame

    def _browse_file(self, key, var):
        if "dir" in key:
            path = filedialog.askdirectory(parent=self)
        else:
            path = filedialog.askopenfilename(
                filetypes=[("Excel files","*.xlsx")], parent=self)
        if path:
            var.set(path)

    def _save_settings_from_ui(self):
        for key, var in self._setting_vars.items():
            val = var.get().strip()
            if key in ("threshold", "slow_down"):
                try:    self._settings[key] = int(val)
                except: messagebox.showwarning("Invalid", f"'{key}' must be an integer.", parent=self); return
            elif key in ("min_delay", "max_delay"):
                try:    self._settings[key] = float(val)
                except: messagebox.showwarning("Invalid", f"'{key}' must be a number.", parent=self); return
            else:
                self._settings[key] = val
        save_settings(self._settings)
        messagebox.showinfo("Saved", "Settings saved to hub_config.json", parent=self)
        self._set_status("Settings saved")

    def _reset_settings(self):
        if messagebox.askyesno("Reset", "Reset all settings to defaults?", parent=self):
            self._settings = dict(DEFAULT_SETTINGS)
            for key, var in self._setting_vars.items():
                var.set(str(self._settings.get(key, "")))
            save_settings(self._settings)

    def _backup_excel(self):
        if not os.path.exists(EXCEL_PATH):
            messagebox.showwarning("Not Found", "Excel file not found.", parent=self)
            return
        import shutil
        ts  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        bkp = EXCEL_PATH.replace(".xlsx", f"_backup_{ts}.xlsx")
        shutil.copy2(EXCEL_PATH, bkp)
        messagebox.showinfo("Backup Created", f"Saved:\n{os.path.basename(bkp)}", parent=self)

    def _open_snapshots_folder(self):
        os.makedirs(SNAP_DIR, exist_ok=True)
        self._open_path(SNAP_DIR)

    def _open_csv_variations(self):
        csv_path = os.path.join(BASE_DIR, "keyword_variations_results.csv")
        if os.path.exists(csv_path):
            self._open_path(csv_path)
        else:
            messagebox.showwarning("Not Found",
                                   "keyword_variations_results.csv not found.\n"
                                   "Run 'Keyword Variation Finder' first.", parent=self)

    # ══════════════════════════════════════════════════════════════════════════
    #  Tool runner
    # ══════════════════════════════════════════════════════════════════════════
    def _run_tool(self, tool):
        if self._proc and self._proc.poll() is None:
            messagebox.showwarning("Already Running",
                                   "A tool is currently running. Stop it first.", parent=self)
            return
        if tool.get("warn"):
            if not messagebox.askyesno("Confirm", tool["warn"], parent=self):
                return
        script = os.path.join(BASE_DIR, tool["script"])
        if not os.path.exists(script):
            messagebox.showerror("Not Found", f"Script not found:\n{script}", parent=self)
            return

        self._show_page("scrapers")
        self._active_tool = tool
        self._clear_log()
        self._log_write(f"{'─'*60}\n", "muted")
        self._log_write(f"  {tool['icon']}  {tool['title']}\n", "green")
        self._log_write(f"{'─'*60}\n\n", "muted")
        self._set_status(f"Running: {tool['title']}", running=True)
        self._log_status_dot.config(fg=C_GREEN)
        self._btn_stop.config(state="normal")
        self._disable_run_buttons()

        def _run():
            try:
                self._proc = subprocess.Popen(
                    [PYTHON, script], cwd=BASE_DIR,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, bufsize=1)
                for line in self._proc.stdout:
                    self._log_queue.put(("line", line))
                self._proc.wait()
                rc = self._proc.returncode
                self._log_queue.put(("done", None) if rc == 0 else ("error", rc))
            except Exception as e:
                self._log_queue.put(("exception", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _stop_tool(self):
        if self._proc and self._proc.poll() is None:
            self._proc.terminate()
            self._log_write("\n  [Stopped by user]\n", "yellow")
            self._set_status("Stopped")
            self._on_tool_done()

    def _on_tool_done(self):
        self._btn_stop.config(state="disabled")
        self._log_status_dot.config(fg=C_MUTED)
        self._enable_run_buttons()
        self._refresh_all()
        self._active_tool = None

    def _disable_run_buttons(self):
        for b in self._tool_buttons.values():
            b.config(state="disabled", bg=_adj(C_BG, 1.5), fg=C_MUTED)

    def _enable_run_buttons(self):
        for tool in TOOLS:
            self._tool_buttons[tool["id"]].config(
                state="normal", bg=tool["color"], fg=tool.get("btn_fg", "white"))

    # ── Log ────────────────────────────────────────────────────────────────────
    def _log_write(self, text, tag=None):
        self._log.config(state="normal")
        self._log.insert("end", text, tag) if tag else self._log.insert("end", text)
        self._log.see("end")
        self._log.config(state="disabled")

    def _clear_log(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")

    def _color_line(self, line):
        lo = line.lower()
        if any(x in lo for x in ["error","failed","captcha","warning","crash"]):  return "red"
        if any(x in lo for x in ["positive","saved","done","complete","✓","ok","success"]): return "green"
        if any(x in lo for x in ["skip","already","n/a","not found"]):            return "yellow"
        if any(x in lo for x in ["searching","opening","starting","scraping","analyzing"]): return "blue"
        if any(x in lo for x in ["variation","keyword","break","entry"]):          return "purple"
        if line.strip().startswith("─") or line.strip().startswith("="):           return "muted"
        return None

    def _poll_log(self):
        try:
            while True:
                kind, payload = self._log_queue.get_nowait()
                if kind == "line":
                    self._log_write(payload, self._color_line(payload))
                elif kind == "done":
                    self._log_write("\n  ✓  Finished successfully\n", "green")
                    self._set_status("Done")
                    self._on_tool_done()
                elif kind == "error":
                    self._log_write(f"\n  ✗  Exited with code {payload}\n", "red")
                    self._set_status(f"Exited with code {payload}")
                    self._on_tool_done()
                elif kind == "exception":
                    self._log_write(f"\n  ✗  Launch error: {payload}\n", "red")
                    self._set_status("Launch error")
                    self._on_tool_done()
        except queue.Empty:
            pass
        self.after(50, self._poll_log)

    # ── Cross-platform open ────────────────────────────────────────────────────
    def _open_path(self, path):
        """Open a file or folder in the system default application."""
        try:
            if _PLATFORM == 'Windows':
                os.startfile(path)
            elif _PLATFORM == 'Darwin':
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Open Error", str(e), parent=self)

    # ── Utilities ──────────────────────────────────────────────────────────────
    def _set_status(self, msg, running=False):
        self._status_var.set(f"  {msg}")
        if running:
            self._running_lbl.config(text="●  Running")
            self._animate_indicator()
        else:
            self._running_lbl.config(text="")

    def _animate_indicator(self):
        frames = ["●  Running", "●●  Running", "●●●  Running", "●  Running"]
        def _tick(i=0):
            if self._proc and self._proc.poll() is None:
                self._running_lbl.config(text=frames[i % 4])
                self.after(550, _tick, i + 1)
            else:
                self._running_lbl.config(text="")
        self.after(550, _tick)

    def _refresh_all(self):
        if self._current_page == "dashboard":     self._refresh_dashboard()
        elif self._current_page == "analysis":    self._load_analysis_data()
        elif self._current_page == "niche_manager": self._load_niche_data()
        elif self._current_page == "trends":      self._refresh_trends_page()

    def _open_excel(self):
        ep = self._settings.get("excel_path", EXCEL_PATH)
        if not os.path.exists(ep):
            messagebox.showwarning("Not Found",
                                   "Excel not found.\nRun 'Create / Regenerate Excel' first.",
                                   parent=self)
            return
        self._open_path(ep)
        self._set_status("Opened Excel")

    def _open_images(self):
        d = self._settings.get("images_dir", IMAGES_DIR)
        os.makedirs(d, exist_ok=True)
        self._open_path(d)
        self._set_status("Opened gig_images folder")


    # ══════════════════════════════════════════════════════════════════════════
    #  API Key Management
    # ══════════════════════════════════════════════════════════════════════════
    def _check_api_key_on_startup(self):
        if not _API_MANAGER_OK:
            return
        if not has_api_key():
            self._show_api_key_dialog(initial=True)
        else:
            self._refresh_ai_settings_display()

    def _show_api_key_dialog(self, initial=False):
        """Modal dialog to configure (or update) the AI API key."""
        dlg = tk.Toplevel(self)
        dlg.title("AI Configuration")
        dlg.configure(bg=C_PANEL)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        # Center on parent
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 460) // 2
        y = self.winfo_y() + (self.winfo_height() - 380) // 2
        dlg.geometry(f"460x380+{x}+{y}")

        # ── Header ──────────────────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg=C_GREEN, height=4)
        hdr.pack(fill="x")

        title_bar = tk.Frame(dlg, bg=C_PANEL)
        title_bar.pack(fill="x", padx=24, pady=(16, 0))
        tk.Label(title_bar, text="🤖  AI Configuration", font=FT_H2,
                 bg=C_PANEL, fg=C_TEXT).pack(side="left")

        if initial:
            tk.Label(dlg,
                     text="An AI API key is required to generate sub-niches.\n"
                          "Please choose your provider and enter your key.",
                     font=FT_SMALL, bg=C_PANEL, fg=C_TEXT2,
                     justify="left", wraplength=400).pack(padx=24, pady=(8, 0), anchor="w")

        # ── Form ────────────────────────────────────────────────────────────
        form = tk.Frame(dlg, bg=C_PANEL)
        form.pack(fill="x", padx=24, pady=16)

        # Provider
        tk.Label(form, text="AI Provider", font=FT_H3,
                 bg=C_PANEL, fg=C_TEXT, anchor="w").pack(anchor="w", pady=(0, 4))
        provider_var = tk.StringVar()
        if _API_MANAGER_OK:
            cfg = load_api_config()
            provider_var.set("Google Gemini" if cfg.get("provider","gemini") == "gemini" else "OpenAI ChatGPT")
        else:
            provider_var.set("Google Gemini")

        provider_combo = ttk.Combobox(form, textvariable=provider_var,
                                      values=["Google Gemini", "OpenAI ChatGPT"],
                                      state="readonly", font=FT_BODY, width=30)
        provider_combo.pack(anchor="w", pady=(0, 12))

        # API Key
        tk.Label(form, text="API Key", font=FT_H3,
                 bg=C_PANEL, fg=C_TEXT, anchor="w").pack(anchor="w", pady=(0, 4))

        key_row = tk.Frame(form, bg=C_PANEL)
        key_row.pack(fill="x", pady=(0, 6))

        if _API_MANAGER_OK:
            stored_key = load_api_config().get("api_key", "")
        else:
            stored_key = ""

        key_var = tk.StringVar(value=stored_key)
        key_entry = tk.Entry(key_row, textvariable=key_var, font=FT_BODY,
                             bg=C_BG, fg=C_TEXT, insertbackground=C_GREEN,
                             relief="flat", show="•", width=36)
        key_entry.pack(side="left", ipady=5, padx=(0, 6))

        show_var = tk.BooleanVar(value=False)
        def _toggle_show():
            key_entry.config(show="" if show_var.get() else "•")
        tk.Checkbutton(key_row, text="Show", variable=show_var, command=_toggle_show,
                       bg=C_PANEL, fg=C_TEXT2, selectcolor=C_BG,
                       activebackground=C_PANEL, activeforeground=C_TEXT,
                       font=FT_SMALL, relief="flat", bd=0).pack(side="left")

        # Status label
        status_lbl = tk.Label(form, text="", font=FT_SMALL,
                              bg=C_PANEL, fg=C_MUTED, anchor="w", wraplength=400)
        status_lbl.pack(anchor="w", pady=(0, 4))

        # ── Buttons ─────────────────────────────────────────────────────────
        btn_row = tk.Frame(dlg, bg=C_PANEL)
        btn_row.pack(fill="x", padx=24, pady=(0, 20))

        def _get_provider_id():
            return "gemini" if "Gemini" in provider_var.get() else "openai"

        def _test():
            k = key_var.get().strip()
            if not k:
                status_lbl.config(text="Please enter an API key.", fg=C_YELLOW)
                return
            status_lbl.config(text="Testing…", fg=C_MUTED)
            dlg.update()
            if _API_MANAGER_OK:
                ok, msg = test_api_key(_get_provider_id(), k)
                status_lbl.config(text=msg, fg=C_SUCCESS if ok else C_RED)
            else:
                status_lbl.config(text="api_manager not available.", fg=C_RED)

        def _save():
            k = key_var.get().strip()
            if not k:
                status_lbl.config(text="API key cannot be empty.", fg=C_YELLOW)
                return
            if _API_MANAGER_OK:
                save_api_config(_get_provider_id(), k)
                self._refresh_ai_settings_display()
            status_lbl.config(text="Saved!", fg=C_SUCCESS)
            dlg.after(600, dlg.destroy)

        _btn(btn_row, "Test Connection", _test,
             bg=B_GHOST, fg=B_GHOST_FG, padx=14, pady=8).pack(side="left", padx=(0, 8))
        _btn(btn_row, "💾  Save & Continue", _save,
             bg=B_GREEN, fg=B_GREEN_FG, padx=16, pady=8, font=FT_H3).pack(side="left")
        if not initial:
            _btn(btn_row, "Cancel", dlg.destroy,
                 bg=B_GHOST, fg=B_GHOST_FG, padx=14, pady=8).pack(side="right")

        key_entry.focus_set()

    def _refresh_ai_settings_display(self):
        """Refresh the AI config labels on the Settings page."""
        if hasattr(self, "_ai_provider_lbl"):
            self._ai_provider_lbl.config(text=get_provider_label())
        if hasattr(self, "_ai_key_lbl"):
            self._ai_key_lbl.config(text=get_masked_key())


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = FiverrHub()
    app.mainloop()
